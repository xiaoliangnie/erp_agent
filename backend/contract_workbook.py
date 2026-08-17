# -*- coding: utf-8 -*-
"""用 openpyxl 把合同模型写成采购单工作簿。

16 列 A–P、明细从第 9 行起、小计 ``=N{row}*L{row}``、合计 ``=SUM(O…)``。
国标码（E，商品条码）写文本并使用 ``@`` 格式；品名在 F，执行标准（G，GB/T…）紧随品名。
需方收货信息占 A6:A7 / B6:H7；一条明细时送货地址落在第 14 行。
明细备注写在 Q 列，不进 A–P 合同区。
"""
from __future__ import annotations

import tempfile
from collections import namedtuple
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.units import pixels_to_EMU
from PIL import Image as PILImage


SHEET_NAME = "合同"
ITEM_START = 9
COLUMN_COUNT = 16
COLUMN_WIDTHS = [
    13.34, 17, 17, 17, 16.58, 19.64, 16.5, 8.4,
    9.06, 15.2, 16.93, 12.34, 12.34, 12.7, 12.34, 8.68,
]
IMAGE_COL = 1  # B
IMAGE_WIDTH_PX = 86
IMAGE_HEIGHT_PX = 76
IMAGE_OFFSET_PX = 3
ITEM_ROW_HEIGHT = 82
NATIVE_IMAGE_SUFFIXES = {".png", ".jpg", ".jpeg", ".gif"}
DATE_FORMAT = 'yyyy"年"m"月"d"日"'
QTY_FORMAT = "#,##0.####"
MONEY_FORMAT = '"￥"#,##0.00'
TEXT_FORMAT = "@"
YAHEI = "Microsoft YaHei"
SIMSUN = "SimSun"

Layout = namedtuple(
    "Layout",
    "item_start item_end total_row packaging_row payment_row inspection_row "
    "address_row terms_row parties_row signatures_row final_row",
)

THIN = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)
FONT_BODY = Font(name=YAHEI, size=10, color="000000")
FONT_TITLE = Font(name=YAHEI, size=16, color="000000")
FONT_ITEM = Font(name=YAHEI, size=9, color="000000")
FONT_TERMS = Font(name=YAHEI, size=9, color="000000")
FONT_SIGN = Font(name=SIMSUN, size=11, bold=True, color="000000")
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True)
ALIGN_TOP_LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)
ALIGN_BODY = Alignment(vertical="center", wrap_text=True)


def layout_for(item_count):
    """按明细条数计算合计行、条款行、签字行。"""
    if item_count < 1:
        raise ValueError("采购合同至少需要一条商品明细")
    item_end = ITEM_START + item_count - 1
    total_row = item_end + 1
    packaging_row = total_row + 1
    payment_row = total_row + 2
    inspection_row = total_row + 3
    address_row = total_row + 4
    terms_row = total_row + 5
    parties_row = total_row + 6
    signatures_row = total_row + 7
    return Layout(
        item_start=ITEM_START,
        item_end=item_end,
        total_row=total_row,
        packaging_row=packaging_row,
        payment_row=payment_row,
        inspection_row=inspection_row,
        address_row=address_row,
        terms_row=terms_row,
        parties_row=parties_row,
        signatures_row=signatures_row,
        final_row=signatures_row,
    )


def format_tax_rate_text(value):
    """与原先 toLocaleString(zh-CN, maximumFractionDigits=2) 对齐。"""
    return f"{float(value):.2f}".rstrip("0").rstrip(".")


def unit_price_header(invoice):
    invoice = invoice or {}
    if not invoice.get("type"):
        return "单价"
    if invoice.get("type") == "no_invoice":
        return "单价（不开票）"
    rate = format_tax_rate_text(invoice.get("taxRate") or 0)
    label = invoice.get("label") or ""
    return f"单价（含{rate}%{label}税）"


def write_contract_workbook(model, output_path):
    """把合同模型写入 ``output_path``，返回 Path。"""
    items = list(model.get("items") or [])
    layout = layout_for(len(items))
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = SHEET_NAME
    sheet.sheet_view.showGridLines = False
    with tempfile.TemporaryDirectory(prefix="contract-wb-") as temp_dir:
        _build_sheet(sheet, model, items, layout, temp_dir)
        workbook.save(output_path)
    return output_path


def _build_sheet(sheet, model, items, layout, temp_dir):
    for index, width in enumerate(COLUMN_WIDTHS, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    for address in _merges(layout):
        sheet.merge_cells(address)

    _write_header(sheet, model)
    _write_items(sheet, items, layout, temp_dir, blank=bool(model.get("isTemplate")))
    _write_footer(sheet, model, layout)
    _apply_styles(sheet, layout)
    _place_local_remarks(sheet, layout)


def _merges(layout):
    return [
        "A1:P2", "A3:B3", "C3:P3",
        "B4:H4", "J4:P4", "B5:H5", "J5:P5",
        "A6:A7", "B6:H7", "J6:P6", "J7:P7",
        f"A{layout.total_row}:N{layout.total_row}",
        f"B{layout.packaging_row}:P{layout.packaging_row}",
        f"B{layout.payment_row}:P{layout.payment_row}",
        f"B{layout.inspection_row}:P{layout.inspection_row}",
        f"B{layout.address_row}:P{layout.address_row}",
        f"A{layout.terms_row}:P{layout.terms_row}",
        f"B{layout.parties_row}:I{layout.parties_row}",
        f"J{layout.parties_row}:K{layout.parties_row}",
        f"M{layout.parties_row}:P{layout.parties_row}",
        f"B{layout.signatures_row}:C{layout.signatures_row}",
        f"F{layout.signatures_row}:H{layout.signatures_row}",
        f"J{layout.signatures_row}:P{layout.signatures_row}",
    ]


def _write_header(sheet, model):
    buyer = model.get("buyer") or {}
    supplier = model.get("supplier") or {}
    invoice = model.get("invoice") or {}
    sheet["A1"] = "杭 州 无 际 云 帆 采 购 单"
    sheet["A3"] = "项目负责人/设计师"
    sheet["C3"] = _text(model.get("projectLead"))
    sheet["A4"] = "下单日期"
    sheet["B4"] = _excel_date(model.get("orderDate"))
    sheet["I4"] = "交货日期"
    sheet["J4"] = _excel_date(model.get("deliveryDate"))
    sheet["A5"] = "需方："
    sheet["B5"] = _text(buyer.get("company_name"))
    sheet["I5"] = "供方："
    sheet["J5"] = _text(supplier.get("legalName"))
    sheet["A6"] = "收货信息"
    sheet["B6"] = _text(model.get("receivingInfo") or model.get("deliveryAddress"))
    sheet["I6"] = "地址："
    sheet["J6"] = _text(supplier.get("address"))
    sheet["I7"] = "联系人："
    sheet["J7"] = _text(supplier.get("contact"))
    headers = [
        "序号", "图片", "款式编码", "商品编码", "国标码", "品名", "执行标准", "分类",
        "虚拟分类", "材质工艺", "包装方式", "数量", "单位", unit_price_header(invoice),
        "小计：元", "备注",
    ]
    for column, value in enumerate(headers, start=1):
        sheet.cell(8, column, value)
    sheet.cell(8, 17, "备注")


def _write_items(sheet, items, layout, temp_dir, *, blank=False):
    for index, item in enumerate(items):
        row = layout.item_start + index
        sheet.cell(row, 15, f"=N{row}*L{row}")
        if blank:
            continue
        sheet.cell(row, 1, index + 1)
        sheet.cell(row, 3, _text(item.get("styleCode")))
        sheet.cell(row, 4, _text(item.get("sku")))
        sheet.cell(row, 5, _text(item.get("nationalCode")))
        sheet.cell(row, 6, _text(item.get("name")))
        sheet.cell(row, 7, _text(item.get("gbStandard")))
        sheet.cell(row, 8, _text(item.get("category")))
        sheet.cell(row, 9, _text(item.get("virtualCategory")))
        sheet.cell(row, 10, _text(item.get("materialProcess")))
        sheet.cell(row, 11, _text(item.get("packaging")))
        sheet.cell(row, 12, item.get("quantity") if item.get("quantity") is not None else 0)
        sheet.cell(row, 13, _text(item.get("unit") or "个"))
        sheet.cell(row, 14, float(item.get("unitPrice") or 0))
        sheet.cell(row, 17, _text(item.get("remark")))
        _place_item_image(sheet, row, item.get("imagePath"), temp_dir)


def _write_footer(sheet, model, layout):
    buyer = model.get("buyer") or {}
    supplier = model.get("supplier") or {}
    terms = model.get("terms") or []
    sheet.cell(layout.total_row, 1, "合计金额：")
    sheet.cell(layout.total_row, 15, f"=SUM(O{layout.item_start}:O{layout.item_end})")
    sheet.cell(layout.packaging_row, 1, "包装")
    sheet.cell(layout.packaging_row, 2, _text(model.get("packagingTerms")))
    sheet.cell(layout.payment_row, 1, "付款方式")
    sheet.cell(layout.payment_row, 2, _text(model.get("paymentTerms")))
    sheet.cell(layout.inspection_row, 1, "检验标准")
    sheet.cell(layout.inspection_row, 2, _text(model.get("inspectionStandards")))
    sheet.cell(layout.address_row, 1, "送货地址")
    sheet.cell(layout.address_row, 2, _text(model.get("deliveryAddress")))
    sheet.cell(layout.terms_row, 1, "\n".join(str(term) for term in terms))
    sheet.cell(layout.parties_row, 1, "需方：")
    sheet.cell(layout.parties_row, 2, _text(buyer.get("company_name")))
    sheet.cell(layout.parties_row, 10, "供方：")
    sheet.cell(layout.parties_row, 12, _text(supplier.get("legalName")))
    sheet.cell(layout.signatures_row, 1, "申请人：")
    sheet.cell(layout.signatures_row, 2, _text(model.get("applicant")))
    sheet.cell(layout.signatures_row, 4, "直属领导签字：")
    sheet.cell(layout.signatures_row, 9, "副总经理签字：")


def _apply_styles(sheet, layout):
    _style(sheet, f"A1:P{layout.final_row}", font=FONT_BODY, alignment=ALIGN_BODY, border=THIN)
    _style(sheet, "A1:P2", font=FONT_TITLE, alignment=ALIGN_CENTER)
    _style(sheet, "B4:H4", number_format=DATE_FORMAT)
    _style(sheet, "J4:P4", number_format=DATE_FORMAT)
    _style(sheet, "A4:P7", alignment=ALIGN_CENTER)
    _style(sheet, "A8:P8", font=FONT_BODY, alignment=ALIGN_CENTER)
    item_range = f"A{layout.item_start}:P{layout.item_end}"
    _style(sheet, item_range, font=FONT_ITEM, alignment=ALIGN_CENTER)
    _style(sheet, f"C{layout.item_start}:G{layout.item_end}", number_format=TEXT_FORMAT)
    _style(sheet, f"L{layout.item_start}:L{layout.item_end}", number_format=QTY_FORMAT)
    _style(sheet, f"N{layout.item_start}:O{layout.item_end}", number_format=MONEY_FORMAT)
    _style(sheet, f"A{layout.total_row}:N{layout.total_row}", alignment=ALIGN_RIGHT)
    _style(sheet, f"O{layout.total_row}", number_format=MONEY_FORMAT)
    for row in (
        layout.packaging_row, layout.payment_row,
        layout.inspection_row, layout.address_row,
    ):
        _style(sheet, f"A{row}", alignment=ALIGN_CENTER)
        _style(sheet, f"B{row}:P{row}", alignment=ALIGN_LEFT)
        sheet.row_dimensions[row].height = 46
    _style(
        sheet, f"A{layout.terms_row}:P{layout.terms_row}",
        font=FONT_TERMS, alignment=ALIGN_TOP_LEFT,
    )
    _style(
        sheet, f"A{layout.signatures_row}:P{layout.signatures_row}",
        font=FONT_SIGN, alignment=ALIGN_CENTER,
    )
    for row in (1, 2):
        sheet.row_dimensions[row].height = 18
    for row in range(3, 6):
        sheet.row_dimensions[row].height = 23
    for row in (6, 7):
        sheet.row_dimensions[row].height = 32
    sheet.row_dimensions[8].height = 34
    _style(
        sheet, f"A{layout.parties_row}:P{layout.parties_row}",
        alignment=ALIGN_CENTER,
    )
    _style(sheet, "Q8", font=FONT_BODY, alignment=ALIGN_CENTER)
    if layout.item_end >= layout.item_start:
        _style(
            sheet, f"Q{layout.item_start}:Q{layout.item_end}",
            font=FONT_ITEM, alignment=ALIGN_LEFT, number_format=TEXT_FORMAT,
        )
    for row in range(layout.item_start, layout.item_end + 1):
        sheet.row_dimensions[row].height = ITEM_ROW_HEIGHT
    sheet.row_dimensions[layout.total_row].height = 23
    sheet.row_dimensions[layout.terms_row].height = 270
    sheet.row_dimensions[layout.parties_row].height = 24
    sheet.row_dimensions[layout.signatures_row].height = 24


def _place_local_remarks(sheet, layout):
    """备注写在可见的 Q 列，打印区仍只到 P，预览按 A–P 出图。"""
    sheet.cell(8, 17, "备注")
    sheet.column_dimensions["Q"].hidden = False
    sheet.column_dimensions["Q"].width = 28
    sheet.print_area = f"A1:P{layout.final_row}"


def _place_item_image(sheet, row, image_path, temp_dir):
    if not image_path:
        sheet.cell(row, 2, "暂无图片")
        return
    try:
        embed_path = _embeddable_image_path(image_path, temp_dir)
        picture = ExcelImage(embed_path)
        picture.width = IMAGE_WIDTH_PX
        picture.height = IMAGE_HEIGHT_PX
        picture.anchor = OneCellAnchor(
            _from=AnchorMarker(
                col=IMAGE_COL,
                colOff=pixels_to_EMU(IMAGE_OFFSET_PX),
                row=row - 1,
                rowOff=pixels_to_EMU(IMAGE_OFFSET_PX),
            ),
            ext=XDRPositiveSize2D(
                cx=pixels_to_EMU(IMAGE_WIDTH_PX),
                cy=pixels_to_EMU(IMAGE_HEIGHT_PX),
            ),
        )
        sheet.add_image(picture)
    except Exception:
        sheet.cell(row, 2, "图片读取失败")


def _embeddable_image_path(image_path, temp_dir):
    """xlsx 只稳定支持 png/jpeg/gif；webp 等先转成 png。"""
    path = Path(image_path)
    if not path.is_file():
        raise FileNotFoundError(str(path))
    if path.suffix.lower() in NATIVE_IMAGE_SUFFIXES:
        return str(path.resolve())
    converted = Path(temp_dir) / f"{path.stem}.png"
    with PILImage.open(path) as image:
        image.load()
        image.save(converted, format="PNG")
    return str(converted)


def _style(sheet, coord, *, font=None, alignment=None, number_format=None, border=None):
    for cell in _cells(sheet, coord):
        if font is not None:
            cell.font = font
        if alignment is not None:
            cell.alignment = alignment
        if number_format is not None:
            cell.number_format = number_format
        if border is not None:
            cell.border = border


def _cells(sheet, coord):
    region = sheet[coord]
    if not isinstance(region, tuple):
        yield region
        return
    for row in region:
        if not isinstance(row, tuple):
            yield row
            continue
        for cell in row:
            yield cell


def _excel_date(value):
    text = str(value or "").strip()
    if not text:
        return None
    return datetime.strptime(text[:10], "%Y-%m-%d")


def _text(value):
    if value is None:
        return ""
    return str(value)
