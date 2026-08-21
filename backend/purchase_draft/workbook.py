# -*- coding: utf-8 -*-
"""采购单草稿 Excel：按供应商拆行，给员工进 ERP 建单用。"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..paths import TEMPLATES_DIR, local_dir

SHEET1_COLUMNS = [
    "供应商", "商品编码", "款式编码", "商品名称", "颜色及规格",
    "数量", "单价", "交货日期", "仓库", "备注",
]
SHEET2_COLUMNS = ["供应商", "行数", "数量合计"]
TEXT_COLUMNS = {"商品编码", "款式编码", "交货日期"}
HEADER_FONT = Font(name="宋体", size=11)
HEADER_FILL = PatternFill("solid", fgColor="E6E6E6")
HEADER_BORDER = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)
COLUMN_WIDTHS = {
    "供应商": 22,
    "商品编码": 20,
    "款式编码": 18,
    "商品名称": 28,
    "颜色及规格": 18,
    "备注": 24,
}
TEMPLATE_NAME = "采购单模板.xlsx"


def purchase_template_path(*, root=None) -> Path:
    if root is None:
        return TEMPLATES_DIR / TEMPLATE_NAME
    return local_dir("templates", root=root) / TEMPLATE_NAME


def _write_sheet(book, title, columns, rows, *, freeze="A2"):
    sheet = book.active if title == "Sheet1" else book.create_sheet(title)
    sheet.title = title
    for col, header in enumerate(columns, start=1):
        cell = sheet.cell(1, col, header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = HEADER_BORDER
        cell.alignment = Alignment(vertical="center")
        letter = get_column_letter(col)
        sheet.column_dimensions[letter].width = COLUMN_WIDTHS.get(header, 12)
    for index, row in enumerate(rows, start=2):
        for col, header in enumerate(columns, start=1):
            value = row.get(header)
            if value == "":
                value = None
            cell = sheet.cell(index, col, value)
            cell.alignment = Alignment(vertical="center")
            if header in TEXT_COLUMNS:
                cell.number_format = "@"
            elif header == "单价" and value is not None:
                cell.number_format = "0.####"
    last_row = max(1, len(rows) + 1)
    last_col = get_column_letter(len(columns))
    sheet.auto_filter.ref = f"A1:{last_col}{last_row}"
    if freeze:
        sheet.freeze_panes = freeze
    return sheet


def lines_to_rows(lines: list[dict]) -> list[dict]:
    rows = []
    for line in lines:
        rows.append({
            "供应商": line.get("supplier") or "",
            "商品编码": line.get("sku") or "",
            "款式编码": line.get("styleId") or "",
            "商品名称": line.get("name") or "",
            "颜色及规格": line.get("spec") or "",
            "数量": line.get("qty"),
            "单价": line.get("price"),
            "交货日期": line.get("deliveryDate") or "",
            "仓库": line.get("warehouse") or "",
            "备注": line.get("remark") or "",
        })
    return rows


def group_rows(groups: list[dict]) -> list[dict]:
    rows = []
    for group in groups:
        rows.append({
            "供应商": group.get("supplier") or "未对上供应商",
            "行数": group.get("lines"),
            "数量合计": group.get("qty"),
        })
    return rows


def write_purchase_draft_workbook(draft: dict, path) -> Path:
    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)
    book = Workbook()
    _write_sheet(book, "Sheet1", SHEET1_COLUMNS, lines_to_rows(draft.get("lines") or []))
    _write_sheet(book, "Sheet2", SHEET2_COLUMNS, group_rows(draft.get("groups") or []), freeze="")
    book.save(target)
    return target


def write_blank_purchase_template(*, root=None) -> Path:
    path = purchase_template_path(root=root)
    path.parent.mkdir(parents=True, exist_ok=True)
    write_purchase_draft_workbook({"lines": [], "groups": []}, path)
    return path
