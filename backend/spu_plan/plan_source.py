# -*- coding: utf-8 -*-
"""订货表上传：校验工作簿 → 原子落盘 → 后台重生成生产计划表。

入口两个：网页 `/api/spu/plan-source` 上传、钉钉给机器人发 xlsx 文件。
都汇到 `PlanSourceUpdater.update`，同一把锁保证同时只有一次更新。
"""
from __future__ import annotations

import io
import logging
import threading
from pathlib import Path

from openpyxl import load_workbook

from ..business_time import business_timestamp

logger = logging.getLogger(__name__)

REQUIRED_SHEETS = ("重点产品订货", "爆品订货", "生产计划表", "库存")
MAX_UPLOAD_BYTES = 40 * 1024 * 1024


class PlanSourceError(ValueError):
    """订货表校验失败或更新忙；消息可直接回给员工。"""


def validate_plan_workbook(data: bytes) -> dict:
    """确认上传的是订货表工作簿：四张表都在、订货表里有款式编码。"""
    if not data:
        raise PlanSourceError("文件是空的")
    if len(data) > MAX_UPLOAD_BYTES:
        raise PlanSourceError("文件超过 40MB，看起来不是订货表")
    try:
        book = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:
        raise PlanSourceError("不是有效的 xlsx 文件") from exc
    try:
        missing = [name for name in REQUIRED_SHEETS if name not in book.sheetnames]
        if missing:
            raise PlanSourceError(
                "缺少工作表：" + "、".join(missing) + "；请发完整的订货表工作簿"
            )
        key = book["重点产品订货"]
        # 这份工作簿的 dimension 元数据是坏的（read_only 会误读成 1×1），忽略它重扫
        if hasattr(key, "reset_dimensions"):
            key.reset_dimensions()
        styles = 0
        for row in key.iter_rows(min_row=4, min_col=2, max_col=2, values_only=True):
            if row and row[0]:
                styles += 1
        if styles == 0:
            raise PlanSourceError("「重点产品订货」里没有款式编码，文件不对")
        return {"styles": styles, "sheets": list(book.sheetnames)}
    finally:
        book.close()


def save_plan_source(data: bytes, path) -> None:
    dest = Path(path)
    dest.parent.mkdir(parents=True, exist_ok=True)
    tmp = dest.with_suffix(dest.suffix + ".tmp")
    tmp.write_bytes(data)
    tmp.replace(dest)


class PlanSourceUpdater:
    """收文件 → 校验 → 覆盖 `SPU_PLAN_SOURCE_XLSX` → 后台重生成（约 3 分钟）。"""

    def __init__(self, *, env_path: str, source_path: str):
        self.env_path = env_path
        self.source_path = str(source_path or "").strip()
        self.sender = None
        self.audit = None
        self.alert_enabled = True
        self._lock = threading.Lock()
        self.state = {"updatedAt": "", "origin": "", "lastError": "", "running": False}

    def status(self) -> dict:
        return {**self.state, "sourcePath": self.source_path}

    def update(self, data: bytes, *, origin: str, notify=None) -> dict:
        """校验并落盘后立刻返回；重生成在后台线程做完再回调 notify(text)。"""
        if not self.source_path:
            raise PlanSourceError("未配置订货表路径（SPU_PLAN_SOURCE_XLSX）")
        checked = validate_plan_workbook(data)
        if not self._lock.acquire(blocking=False):
            raise PlanSourceError("上一次订货表更新还在重生成，稍等几分钟再发")
        try:
            save_plan_source(data, self.source_path)
            self.state.update(
                updatedAt=business_timestamp(), origin=origin, lastError="", running=True,
            )
        except Exception:
            self._lock.release()
            raise

        def run():
            from .production_plan import build_production_plan

            try:
                result = build_production_plan(self.source_path, self.env_path)
                self.state.update(lastError="", running=False)
                logger.info(
                    "订货表更新（%s）→ 生产计划表重生成完成：%s 款", origin, result["styles"],
                )
                alerts = result.get("alerts") or {}
                if self.alert_enabled:
                    from .alerts import push_plan_workbook
                    try:
                        push_plan_workbook(
                            result.get("xlsx") or "",
                            sender=self.sender, audit=self.audit,
                            force=True, operator=origin,
                            today=result.get("today") or "",
                        )
                    except Exception:
                        logger.exception("订货表更新后发送生产计划表失败")
                if notify is not None:
                    added = result.get("added") or []
                    dropped = result.get("dropped") or []
                    extra = ""
                    if added:
                        extra += f"；新进表 {len(added)} 款"
                    if dropped:
                        extra += f"；退出 {len(dropped)} 款"
                    replenish = len(alerts.get("replenish") or [])
                    inbound = len(alerts.get("inbound") or [])
                    extra += f"；当月需补货 {replenish}、及时入库 {inbound}"
                    notify(
                        f"【订货表已更新】生产计划表重生成完成：{result['styles']} 款{extra}。"
                        f"文件：{result['xlsx']}"
                    )
            except Exception as exc:
                message = f"{type(exc).__name__}: {exc}"
                self.state.update(lastError=message[:500], running=False)
                logger.exception("订货表更新后的重生成失败（%s）", origin)
                if notify is not None:
                    notify(f"【订货表更新失败】文件已保存，但重生成报错：{message[:200]}")
            finally:
                self._lock.release()

        threading.Thread(target=run, name="plan-source-update", daemon=True).start()
        return {"ok": True, "started": True, **checked}
