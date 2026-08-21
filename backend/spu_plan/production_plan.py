# -*- coding: utf-8 -*-
"""生产计划表：镜像圈款 + 月底订货量 + 每天跟库存刷新进度。

三层各管各的，不要混：
- 名册：镜像商品资料标签「重点产品」，当天打标/摘标就进退表。
- 需求：员工月底上传工作簿里「重点产品订货」「爆品订货」的最终需求。
  系统不生成订货量；两次上传之间需求数冻结。
- 进度：每天用镜像现势库存/在途/当月净销量重算「生产计划表」。

员工原表 11 月需求错引了 12 月列（FD/FE），这里按月份对齐修正。
"""
from __future__ import annotations

import calendar
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule, Rule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils import column_index_from_string, get_column_letter

from ..business_time import business_today
from ..database import REALTIME_PRODUCT_TABLE, connect
from ..forecast.dataset import load_in_transit
from ..paths import local_dir
from .alerts import collect_month_alerts
from .roster import CATEGORY_LINES, parse_product_attrs
from .service import SALES_OUT_TABLE, SALES_STATUSES_EXCLUDED

SALES_RETURN_TABLE = "realtime_sales_returns"
# 仅退款/投诉/补发/维修没有实收入库，不能抵净销。换货有 r_qty 要扣。
RETURN_TYPES_EXCLUDED = ("仅退款", "投诉", "补发", "维修")
RETURN_STATUSES_EXCLUDED = ("Cancelled", "Cancel", "作废", "Merged")

# 名册规则（2026-08-20 拍板）：商品资料标签带「重点产品」的款才进生产计划表
ROSTER_TAG = "重点产品"

# 两张订货表的「最终需求」列块都从 2026 年 4 月起顺排
DEMAND_BASE_MONTH = (2026, 4)
# 出库镜像从 2026-06-19 起才有完整数据；这个月份之后的净销量用镜像算，更早沿用员工手贴值。
# 售后入库日还没覆盖到该月月末时，不拿「出库 − 0 退货」去算售罄率，继续用手贴净销。
MIRROR_NET_FROM = (2026, 7)
KEY_SHEET = "重点产品订货"
HOT_SHEET = "爆品订货"
PLAN_SHEET = "生产计划表"
STOCK_SHEET = "库存"
KEY_FINAL_FIRST, KEY_FINAL_COUNT = column_index_from_string("EW"), 11
HOT_FINAL_FIRST, HOT_FINAL_COUNT = column_index_from_string("DK"), 14
PAST_MONTHS = 3
FUTURE_MONTHS = 5


def month_seq(base: tuple[int, int], count: int) -> list[tuple[int, int]]:
    year, month = base
    out = []
    for _ in range(count):
        out.append((year, month))
        month += 1
        if month > 12:
            year, month = year + 1, 1
    return out


def shift_month(anchor: tuple[int, int], delta: int) -> tuple[int, int]:
    index = anchor[0] * 12 + (anchor[1] - 1) + delta
    return index // 12, index % 12 + 1


def _num(value):
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def read_source_plan(path) -> dict:
    """读源工作簿：计划表名册 + 历史月净销量、期初库存、两张订货表最终需求。"""
    book = load_workbook(Path(path), data_only=True)
    try:
        demands: dict[str, dict[tuple[int, int], float]] = {}
        for sheet, first, count in (
            (KEY_SHEET, KEY_FINAL_FIRST, KEY_FINAL_COUNT),
            (HOT_SHEET, HOT_FINAL_FIRST, HOT_FINAL_COUNT),
        ):
            ws = book[sheet]
            months = month_seq(DEMAND_BASE_MONTH, count)
            for row in range(4, ws.max_row + 1):
                style = str(ws.cell(row, 2).value or "").strip()
                if not style:
                    continue
                per = demands.setdefault(style, {})
                for offset, key in enumerate(months):
                    value = _num(ws.cell(row, first + offset).value)
                    if value is not None:
                        per[key] = per.get(key, 0.0) + value

        ws = book[STOCK_SHEET]
        opening: dict[str, tuple[float, float]] = {}
        for row in range(3, ws.max_row + 1):
            style = str(ws.cell(row, 1).value or "").strip()
            if style:
                opening[style] = (
                    _num(ws.cell(row, 2).value) or 0.0,
                    _num(ws.cell(row, 3).value) or 0.0,
                )
        opening_label = str(book[PLAN_SHEET].cell(2, 6).value or "期初库存").strip()

        ws = book[PLAN_SHEET]
        styles = []
        for row in range(6, ws.max_row + 1):
            style = str(ws.cell(row, 3).value or "").strip()
            if not style:
                continue
            styles.append({
                "styleId": style,
                "owner": str(ws.cell(row, 1).value or "").strip(),
                "line": str(ws.cell(row, 2).value or "").strip(),
                "name": str(ws.cell(row, 4).value or "").strip(),
                "node": str(ws.cell(row, 5).value or "").strip(),
                # 员工手贴的历史月净销量（O/R/U 列），带不动就空着
                "pastNet": [
                    _num(ws.cell(row, 15).value),
                    _num(ws.cell(row, 18).value),
                    _num(ws.cell(row, 21).value),
                ],
            })
        return {
            "styles": styles,
            "demands": demands,
            "opening": opening,
            "openingLabel": opening_label,
        }
    finally:
        book.close()


def load_tagged_styles(env_path: str) -> list[dict]:
    """镜像里带「重点产品」标签的启用款：款式编码 / 品类线 / 品名。"""
    sql = (
        f"SELECT sku_id, i_id, name, labels, enabled, category, "
        f"sale_price, cost_price, source_payload "
        f"FROM `{REALTIME_PRODUCT_TABLE}` "
        f"WHERE labels LIKE %s AND COALESCE(i_id, '') <> '' AND enabled = 1"
    )
    with connect(env_path, autocommit=True) as conn:
        with conn.cursor() as cursor:
            cursor.execute(sql, (f"%{ROSTER_TAG}%",))
            rows = cursor.fetchall()
    styles: dict[str, dict] = {}
    for row in rows:
        attrs = parse_product_attrs(row)
        if ROSTER_TAG not in "，".join(attrs["labels"]):
            continue
        item = styles.setdefault(attrs["styleId"], {
            "styleId": attrs["styleId"], "line": "", "name": "",
        })
        if not item["line"] and attrs["categoryLine"]:
            item["line"] = attrs["categoryLine"]
        if not item["name"]:
            item["name"] = attrs.get("productName") or attrs["name"]
    return list(styles.values())


def merge_roster(tagged: list[dict], source_styles: list[dict]) -> dict:
    """标签款为准；员工表提供责任人/订货节点/历史净销量。返回名册与增减清单。"""
    by_style = {item["styleId"]: item for item in source_styles}
    line_order = {name: index for index, name in enumerate(CATEGORY_LINES)}
    merged = []
    for item in sorted(
        tagged, key=lambda s: (line_order.get(s["line"], 99), s["styleId"]),
    ):
        meta = by_style.get(item["styleId"]) or {}
        merged.append({
            "styleId": item["styleId"],
            "owner": meta.get("owner") or "",
            "line": meta.get("line") or item["line"],
            "name": meta.get("name") or item["name"],
            "node": meta.get("node") or "",
            "pastNet": meta.get("pastNet") or [None] * PAST_MONTHS,
        })
    tagged_ids = {item["styleId"] for item in tagged}
    return {
        "styles": merged,
        "added": sorted(tagged_ids - set(by_style)),
        "dropped": sorted(set(by_style) - tagged_ids),
    }


def _chunks(items: list, size: int = 800):
    for start in range(0, len(items), size):
        yield items[start:start + size]


def apply_monthly_returns(
    outbound: dict[tuple[int, int], float],
    returns: dict[tuple[int, int], float],
) -> dict[tuple[int, int], float]:
    """出库 − 实收退货。缺月按 0，不截断负数。"""
    keys = set(outbound) | set(returns)
    return {
        key: float(outbound.get(key) or 0) - float(returns.get(key) or 0)
        for key in keys
    }


def _as_date(value) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    try:
        return date.fromisoformat(text[:10])
    except ValueError:
        return None


def returns_cover_month(coverage_end, key: tuple[int, int]) -> bool:
    """售后入库日至少到该月最后一天，才认为这个月的退货齐了。"""
    end = _as_date(coverage_end)
    if end is None:
        return False
    last = date(key[0], key[1], calendar.monthrange(key[0], key[1])[1])
    return end >= last


def pick_past_net(
    key: tuple[int, int],
    by_month: dict,
    staff,
    coverage_end,
):
    """7 月起用镜像净销；售后未覆盖到该月月末则回退员工手贴值。"""
    if key >= MIRROR_NET_FROM and returns_cover_month(coverage_end, key):
        return by_month.get(key)
    return staff


def counts_as_net_return(after_type: str, item_type: str, status: str, r_qty) -> bool:
    """生产计划净销只扣已入库实收。仅退款等类型即使有数字也不扣。"""
    try:
        qty = float(r_qty or 0)
    except (TypeError, ValueError):
        return False
    if qty <= 0:
        return False
    if (after_type or "").strip() in RETURN_TYPES_EXCLUDED:
        return False
    if (item_type or "").strip() in RETURN_TYPES_EXCLUDED:
        return False
    if (status or "").strip() in RETURN_STATUSES_EXCLUDED:
        return False
    return True


def _mirror_table_exists(cursor, name: str) -> bool:
    cursor.execute(
        "SELECT 1 AS ok FROM information_schema.tables "
        "WHERE table_schema = DATABASE() AND table_name = %s LIMIT 1",
        (name,),
    )
    return cursor.fetchone() is not None


def load_live_by_style(env_path: str, style_ids: list[str], *, today: date,
                       months_back: int = PAST_MONTHS) -> dict:
    """镜像现势：按款汇总库存/在途，以及近几个月净销（出库 − 实收退货）。

    出库表几十万行，直接和商品表 join 会超时；先拿款→SKU 映射，
    再按 `idx_salesout_sku_date` 分块聚合。退货表不存在则只算出库。
    """
    wanted = sorted({s for s in style_ids if s})
    if not wanted:
        return {}
    anchor = (today.year, today.month)
    earliest = shift_month(anchor, -months_back)
    window_start = f"{date(earliest[0], earliest[1], 1).isoformat()} 00:00:00"
    live = {
        style: {"qty": 0.0, "transit": 0.0, "byMonth": {}} for style in wanted
    }
    sku_to_style: dict[str, str] = {}
    sku_transit_missing: dict[str, list[str]] = {}
    status_marks = ",".join(["%s"] * len(SALES_STATUSES_EXCLUDED))
    with connect(env_path, autocommit=True) as conn:
        with conn.cursor() as cursor:
            for chunk in _chunks(wanted):
                marks = ",".join(["%s"] * len(chunk))
                cursor.execute(
                    f"SELECT sku_id, i_id FROM `{REALTIME_PRODUCT_TABLE}` "
                    f"WHERE i_id IN ({marks})",
                    chunk,
                )
                for row in cursor.fetchall():
                    sku_to_style[str(row["sku_id"])] = str(row["i_id"])
            skus = sorted(sku_to_style)
            for chunk in _chunks(skus):
                marks = ",".join(["%s"] * len(chunk))
                cursor.execute(
                    f"SELECT sku_id, qty, purchase_qty FROM realtime_inventory "
                    f"WHERE sku_id IN ({marks})",
                    chunk,
                )
                for row in cursor.fetchall():
                    sku = str(row["sku_id"])
                    style = sku_to_style[sku]
                    live[style]["qty"] += float(row["qty"] or 0)
                    transit = float(row["purchase_qty"] or 0)
                    if transit > 0:
                        live[style]["transit"] += transit
                    else:
                        sku_transit_missing.setdefault(style, []).append(sku)
            for chunk in _chunks(skus):
                marks = ",".join(["%s"] * len(chunk))
                cursor.execute(
                    f"SELECT sku_id, LEFT(io_date, 7) AS ym, SUM(qty) AS qty "
                    f"FROM `{SALES_OUT_TABLE}` "
                    f"WHERE sku_id IN ({marks}) AND io_date >= %s "
                    f"AND COALESCE(status, '') NOT IN ({status_marks}) "
                    f"GROUP BY sku_id, ym",
                    (*chunk, window_start, *SALES_STATUSES_EXCLUDED),
                )
                for row in cursor.fetchall():
                    ym = str(row["ym"] or "")
                    try:
                        key = (int(ym[:4]), int(ym[5:7]))
                    except ValueError:
                        continue
                    style = sku_to_style[str(row["sku_id"])]
                    per = live[style]["byMonth"]
                    per[key] = per.get(key, 0.0) + float(row["qty"] or 0)
            if skus and _mirror_table_exists(cursor, SALES_RETURN_TABLE):
                type_marks = ",".join(["%s"] * len(RETURN_TYPES_EXCLUDED))
                status_marks = ",".join(["%s"] * len(RETURN_STATUSES_EXCLUDED))
                for chunk in _chunks(skus):
                    marks = ",".join(["%s"] * len(chunk))
                    cursor.execute(
                        f"SELECT sku_id, LEFT(receive_date, 7) AS ym, SUM(r_qty) AS qty "
                        f"FROM `{SALES_RETURN_TABLE}` "
                        f"WHERE sku_id IN ({marks}) AND receive_date >= %s "
                        f"AND r_qty > 0 "
                        f"AND COALESCE(after_type, '') NOT IN ({type_marks}) "
                        f"AND COALESCE(item_type, '') NOT IN ({type_marks}) "
                        f"AND COALESCE(status, '') NOT IN ({status_marks}) "
                        f"GROUP BY sku_id, ym",
                        (
                            *chunk, window_start,
                            *RETURN_TYPES_EXCLUDED, *RETURN_TYPES_EXCLUDED,
                            *RETURN_STATUSES_EXCLUDED,
                        ),
                    )
                    for row in cursor.fetchall():
                        ym = str(row["ym"] or "")
                        try:
                            key = (int(ym[:4]), int(ym[5:7]))
                        except ValueError:
                            continue
                        style = sku_to_style.get(str(row["sku_id"]))
                        if not style:
                            continue
                        per = live[style]["byMonth"]
                        per[key] = per.get(key, 0.0) - float(row["qty"] or 0)
    # 在途为 0 的 SKU 回退采购明细（数量−已入库），与鞋服 SPU 表同口径
    fallback_skus = [sku for skus in sku_transit_missing.values() for sku in skus]
    if fallback_skus:
        fallback = load_in_transit(env_path, keys=fallback_skus)
        for style, skus in sku_transit_missing.items():
            live[style]["transit"] += sum(float(fallback.get(sku) or 0) for sku in skus)
    return live


def load_returns_through(env_path: str) -> date | None:
    """售后镜像里最晚一笔实收入库日。表不存在或还是空的则 None。"""
    with connect(env_path, autocommit=True) as conn:
        with conn.cursor() as cursor:
            if not _mirror_table_exists(cursor, SALES_RETURN_TABLE):
                return None
            cursor.execute(
                f"SELECT MAX(receive_date) AS mx FROM `{SALES_RETURN_TABLE}` "
                f"WHERE r_qty > 0"
            )
            row = cursor.fetchone() or {}
    return _as_date(row.get("mx"))


def month_label(key: tuple[int, int]) -> str:
    return f"{key[1]}月"


def plan_workbook_filename(today: date | None = None) -> str:
    day = today or business_today()
    return f"{day.strftime('%y%m')}-生产计划表.xlsx"


def plan_workbook_path(*, root=None, today: date | None = None) -> Path:
    return local_dir("outputs", root=root) / "spu" / plan_workbook_filename(today)


# 版式对齐员工原表（2026-08-20 从源文件逐项扒出）
THIN = Side(style="thin")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
RED = "FFFE0300"
FILL_HEAD = PatternFill("solid", fgColor="E4EDDB")      # 浅绿：常规表头
FILL_OPENING = PatternFill("solid", fgColor="FFF3CE")   # 淡黄：期初库存块
FILL_LIVE = PatternFill("solid", fgColor="FFD966")      # 琥珀：更新数据块表头
FILL_GAP = PatternFill("solid", fgColor="C5DFB4")       # 浅绿深：缺口列表头
FILL_PROGRESS = PatternFill("solid", fgColor="F7B07F")  # 橙：当月进度格
FILL_NODE = PatternFill("solid", fgColor="D9E1F4")      # 蓝紫：订货节点表头
# 判定三色 / 是否两色 / 售罄率警戒。Excel 条件格式认 bgColor，写 fgColor 打不开底色。
def _cf_dxf(bg: str, font_color: str) -> DifferentialStyle:
    return DifferentialStyle(
        fill=PatternFill(bgColor=bg),
        font=Font(size=9, color=font_color),
    )


CF_DXF_AMBER = _cf_dxf("FCE3A1", "813D1A")
CF_DXF_PINK = _cf_dxf("FBC8CF", "BD101E")
CF_DXF_GREEN = _cf_dxf("C9EDD1", "25B059")


def _contains_text_rule(text: str, dxf: DifferentialStyle, anchor: str) -> Rule:
    """与员工原表相同的「单元格包含」规则。"""
    rule = Rule(type="containsText", operator="containsText", text=text, dxf=dxf)
    rule.formula = [f'NOT(ISERROR(SEARCH("{text}",{anchor})))']
    return rule


COLUMN_WIDTHS = {
    "fixed": (5.25, 10.12, 14.52, 21.98, 10.38),
    "opening_qty": 8.27, "opening_transit": 7.52,
    "live_qty": 7.4, "live_transit": 7.15, "month_out": 9.65,
    "f_progress": 7.12, "f_remain": 10.11, "f_enough": 10.15,
    "demand": 7.0, "past_net": 7.0, "f_sellthrough": 6.8,
    "f_verdict_now": 11.12, "f_gap_now": 8.62,
    "f_verdict": 11.12, "f_gap": 8.5,
}


def _collapse_columns(ws, letters: dict, detail_indexes: list[int], summary_index: int) -> None:
    """Excel 列分组：summaryRight=False，+ 号在明细左侧，默认折叠。"""
    if not detail_indexes:
        return
    for index in detail_indexes:
        dim = ws.column_dimensions[letters[index]]
        dim.outlineLevel = 1
        dim.hidden = True
    ws.column_dimensions[letters[summary_index]].collapsed = True


def write_production_plan(
    source: dict, live: dict, path, *, today: date | None = None,
    returns_through: date | None = None,
) -> Path:
    """写「生产计划表」一张：数据是值，判定列是表内公式，员工改需求数会自动重算。"""
    today = today or business_today()
    anchor = (today.year, today.month)
    days_in_month = calendar.monthrange(*anchor)[1]
    progress = round(max(0, today.day - 1) / days_in_month, 4)
    past = [shift_month(anchor, -delta) for delta in range(PAST_MONTHS, 0, -1)]
    future = [shift_month(anchor, delta) for delta in range(1, FUTURE_MONTHS + 1)]

    dest = Path(path)
    dest.parent.mkdir(parents=True, exist_ok=True)
    book = Workbook()
    ws = book.active
    ws.title = PLAN_SHEET

    rule = (
        "补货周期规则：（严格执行）\n"
        "1.补货周期在半个月以内的产品按照半个月循环备货，确保上月底库存数满足上半个月的销售需求，"
        "本月15号的库存数满足下半个月的销售需求。\n"
        "2.补货周期在半个月以上的，确保每月月底的库存数满足下一个月的需求数。\n"
        "需综合考虑面辅料生产周期、工厂生产排期、生产时间、起订量等提前做每个款的生产计划。"
    )
    ws.cell(1, 1, rule).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A1:E1")
    ws.row_dimensions[1].height = 58
    note = f"库存/在途/净销量：镜像 {today.isoformat()}"
    uncovered = [
        key for key in past
        if key >= MIRROR_NET_FROM and not returns_cover_month(returns_through, key)
    ]
    if uncovered:
        months = "、".join(f"{key[1]}月" for key in uncovered)
        note += f"；{months}售后未齐，售罄率用手贴净销"
    ws.cell(1, 8, note)

    fixed = ["责任人", "品类线", "款式编码", "名称", "电商部订货节点"]
    group_row, leaf_row, subtotal_row, first_data = 2, 3, 4, 5

    columns: list[dict] = []
    for label in fixed:
        columns.append({"group": "", "leaf": label, "kind": "fixed"})
    columns.append({"group": source.get("openingLabel") or "期初库存", "span": 2,
                    "leaf": "库存", "kind": "opening_qty"})
    columns.append({"group": "", "leaf": "在途", "kind": "opening_transit"})
    columns.append({"group": f"更新数据：库存{today.month}.{today.day}", "span": 3,
                    "leaf": "库存", "kind": "live_qty"})
    columns.append({"group": "", "leaf": "在途", "kind": "live_transit"})
    columns.append({"group": "", "leaf": "净销量", "kind": "month_out"})
    columns.append({"group": "", "leaf": "销售进度", "kind": "f_progress"})
    columns.append({"group": "", "leaf": "当月剩余需求量", "kind": "f_remain"})
    columns.append({"group": "", "leaf": "在仓库存是否满足", "kind": "f_enough"})
    for key in past:
        columns.append({"group": month_label(key), "span": 3,
                        "leaf": "需求数", "kind": "demand", "month": key})
        columns.append({"group": "", "leaf": "净销量", "kind": "past_net", "month": key})
        columns.append({"group": "", "leaf": "售罄率", "kind": "f_sellthrough", "month": key})
    columns.append({"group": month_label(anchor), "span": 3,
                    "leaf": "需求数", "kind": "demand", "month": anchor, "current": True})
    columns.append({"group": "", "leaf": "计划入库数", "kind": "f_verdict_now"})
    columns.append({"group": "", "leaf": f"{anchor[1]}月缺口数", "kind": "f_gap_now"})
    for key in future:
        columns.append({"group": month_label(key), "span": 3,
                        "leaf": "需求数", "kind": "demand", "month": key})
        columns.append({"group": "", "leaf": "计划入库数", "kind": "f_verdict"})
        columns.append({"group": "", "leaf": f"{key[1]}月缺口数", "kind": "f_gap"})

    letters = {}
    for index, column in enumerate(columns, start=1):
        letters[index] = get_column_letter(index)
    def letter_of(kind: str, month=None) -> str:
        for index, column in enumerate(columns, start=1):
            if column["kind"] == kind and (month is None or column.get("month") == month):
                return letters[index]
        raise KeyError(kind)

    # 表头分块配色照员工原表：期初=淡黄、更新数据=琥珀、缺口=浅绿深、其余浅绿
    def header_fill(kind: str):
        if kind in ("opening_qty", "opening_transit"):
            return FILL_OPENING
        if kind in ("live_qty", "live_transit", "month_out",
                    "f_progress", "f_remain", "f_enough"):
            return FILL_LIVE
        if kind in ("f_gap_now", "f_gap"):
            return FILL_GAP
        return FILL_HEAD

    header_red = {"opening_qty", "opening_transit", "f_enough", "f_gap_now", "f_gap"}
    center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
    spans = [
        (index, index + column["span"] - 1)
        for index, column in enumerate(columns, start=1)
        if column.get("span")
    ]
    for index, column in enumerate(columns, start=1):
        kind = column["kind"]
        font = Font(size=9, bold=True, color=RED if kind in header_red else None)
        group_cell = ws.cell(group_row, index)
        leaf_cell = ws.cell(leaf_row, index)
        if column["group"]:
            group_cell.value = column["group"]
        leaf_cell.value = column["leaf"]
        fill = header_fill(kind)
        for cell in (group_cell, leaf_cell):
            cell.font = font
            cell.fill = fill
            cell.border = BORDER
            cell.alignment = center_wrap
        # 原表「更新数据」组头是淡黄、列名行才是琥珀
        if kind in ("live_qty", "live_transit", "month_out"):
            group_cell.fill = FILL_OPENING
        if kind == "fixed" and column["leaf"] == "电商部订货节点":
            for cell in (group_cell, leaf_cell):
                cell.fill = FILL_NODE
        if kind == "fixed":
            width = COLUMN_WIDTHS["fixed"][index - 1]
        else:
            width = COLUMN_WIDTHS.get(kind, 8.5)
        ws.column_dimensions[letters[index]].width = width
    for begin, end in spans:
        ws.merge_cells(
            start_row=group_row, start_column=begin, end_row=group_row, end_column=end,
        )
    # 没有组头的列，组头行和列名行合并成一个高格（原表是 R3:R4 合并的样子）。
    # 销售进度/剩余需求/是否满足三列除外：原表把当月进度百分比嵌在它们的组头行。
    grouped = {i for begin, end in spans for i in range(begin, end + 1)}
    progress_kinds = {"f_progress", "f_remain", "f_enough"}
    for index, column in enumerate(columns, start=1):
        if index not in grouped and not column["group"] and column["kind"] not in progress_kinds:
            ws.cell(group_row, index).value = column["leaf"]
            ws.cell(leaf_row, index).value = None
            ws.merge_cells(
                start_row=group_row, start_column=index, end_row=leaf_row, end_column=index,
            )
    # 原表 K2=已过进度（橙底红字）、L2=1-K2；公式引用 $L$2
    idx_progress = next(i for i, c in enumerate(columns, start=1) if c["kind"] == "f_progress")
    idx_remain = idx_progress + 1
    progress_cell = ws.cell(group_row, idx_progress, progress)
    progress_cell.number_format = "0.0%"
    progress_cell.fill = FILL_PROGRESS
    progress_cell.font = Font(size=9, bold=True, color=RED)
    remain_cell = ws.cell(group_row, idx_remain, f"=1-{letters[idx_progress]}{group_row}")
    remain_cell.number_format = "0.0%"
    # 「更新数据」「期初库存」组标题红字
    for kind in ("live_qty", "opening_qty"):
        idx = next(i for i, c in enumerate(columns, start=1) if c["kind"] == kind)
        ws.cell(group_row, idx).font = Font(size=9, bold=True, color=RED)
    ws.row_dimensions[group_row].height = 17
    ws.row_dimensions[leaf_row].height = 27

    styles = source["styles"]
    demands = source["demands"]
    opening = source["opening"]
    data_font = Font(size=9)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    l_live_qty = letter_of("live_qty")
    l_live_transit = letter_of("live_transit")
    l_month_out = letter_of("month_out")
    l_open_qty = letter_of("opening_qty")
    l_open_transit = letter_of("opening_transit")
    l_now = letter_of("demand", anchor)
    future_letters = [letter_of("demand", key) for key in future]

    for offset, item in enumerate(styles):
        row = first_data + offset
        style = item["styleId"]
        live_row = live.get(style) or {}
        by_month = live_row.get("byMonth") or {}
        open_qty, open_transit = opening.get(style, (None, None))
        per_demand = demands.get(style) or {}
        r = str(row)
        for index, column in enumerate(columns, start=1):
            kind = column["kind"]
            cell = ws.cell(row, index)
            if kind == "fixed":
                cell.value = [item["owner"], item["line"], style, item["name"], item["node"]][index - 1]
            elif kind == "opening_qty":
                cell.value = open_qty
            elif kind == "opening_transit":
                cell.value = open_transit
            elif kind == "live_qty":
                cell.value = live_row.get("qty")
                cell.fill = FILL_OPENING  # 原表实时库存/在途列带淡黄底
            elif kind == "live_transit":
                cell.value = live_row.get("transit")
                cell.fill = FILL_OPENING
            elif kind == "month_out":
                cell.value = by_month.get(anchor)
            elif kind == "demand":
                value = per_demand.get(column["month"])
                cell.value = value
            elif kind == "past_net":
                key = column["month"]
                position = past.index(key)
                staff = item["pastNet"][position] if position < len(item["pastNet"]) else None
                cell.value = pick_past_net(key, by_month, staff, returns_through)
            elif kind == "f_progress":
                cell.value = f"=IFERROR({l_month_out}{r}/{l_now}{r},\"\")"
                cell.number_format = "0.0%"
            elif kind == "f_remain":
                cell.value = f"=IF({l_now}{r}=\"\",\"\",{l_now}{r}-{l_month_out}{r})"
            elif kind == "f_enough":
                remain = letter_of("f_remain")
                cell.value = f"=IFERROR(IF(({l_live_qty}{r}-{remain}{r})>=0,\"是\",\"否\"),\"\")"
            elif kind == "f_sellthrough":
                demand_letter = letters[index - 2]
                net_letter = letters[index - 1]
                cell.value = f"=IFERROR({net_letter}{r}/{demand_letter}{r},\"\")"
                cell.number_format = "0.0%"
            elif kind == "f_verdict_now":
                cell.value = (
                    f"=IF({l_now}{r}=\"\",\"\",IF({l_open_qty}{r}>{l_now}{r},\"库存满足\","
                    f"IF({l_open_qty}{r}+{l_open_transit}{r}>{l_now}{r},\"及时入库\",\"需补货\")))"
                )
            elif kind == "f_gap_now":
                cell.value = (
                    f"=IFERROR(IF({l_now}{r}-{l_open_qty}{r}-{l_open_transit}{r}>0,"
                    f"{l_now}{r}-{l_open_qty}{r}-{l_open_transit}{r},\"\"),\"\")"
                )
                cell.number_format = "#,##0"
            elif kind in ("f_verdict", "f_gap"):
                position = future.index(columns[index - 2 if kind == "f_verdict" else index - 3]["month"])
                mine = future_letters[position]
                consumed = "".join(f"-{fl}{r}" for fl in future_letters[: position + 1])
                consumed_before = "".join(f"-{fl}{r}" for fl in future_letters[:position])
                base = f"{l_live_qty}{r}-{l_now}{r}*${letter_of('f_remain')}${group_row}"
                if kind == "f_verdict":
                    cell.value = (
                        f"=IFERROR(IF({base}{consumed}>0,\"库存满足\","
                        f"IF({base}+{l_live_transit}{r}{consumed}>0,\"及时入库\",\"需补货\")),\"\")"
                    )
                else:
                    cell.value = (
                        f"=IFERROR(IF({mine}{r}-({base}+{l_live_transit}{r}{consumed_before})>0,"
                        f"{mine}{r}-({base}+{l_live_transit}{r}{consumed_before}),\"\"),\"\")"
                    )
                    cell.number_format = "#,##0"
            cell.font = data_font
            cell.border = BORDER
            if kind == "fixed" and column["leaf"] in ("款式编码", "名称"):
                cell.alignment = left
            else:
                cell.alignment = center

    last_row = first_data + len(styles) - 1
    subtotal_font = Font(size=9, color=RED)
    ws.cell(subtotal_row, 1, "小计").font = subtotal_font
    for index, column in enumerate(columns, start=1):
        cell = ws.cell(subtotal_row, index)
        cell.border = BORDER
        cell.alignment = center
        if column["kind"] in ("opening_qty", "opening_transit", "live_qty", "live_transit",
                              "month_out", "demand", "past_net"):
            cell.value = f"=SUBTOTAL(9,{letters[index]}{first_data}:{letters[index]}{last_row})"
            cell.font = subtotal_font

    # 历史月份折叠照原表：最老一个月整组收起；近两个月只露售罄率看警戒色。
    # 订货节点折进名称。+ 号在左侧（summaryRight=False）。
    ws.sheet_properties.outlinePr.summaryRight = False
    ws.sheet_properties.outlinePr.summaryBelow = False
    ws.sheet_format.outlineLevelCol = 1
    name_idx = next(i for i, c in enumerate(columns, start=1) if c["leaf"] == "名称")
    node_idx = next(i for i, c in enumerate(columns, start=1) if c["leaf"] == "电商部订货节点")
    _collapse_columns(ws, letters, [node_idx], name_idx)
    past_by_month: dict[tuple[int, int], list[tuple[int, dict]]] = {}
    for index, column in enumerate(columns, start=1):
        month = column.get("month")
        if month in past:
            past_by_month.setdefault(month, []).append((index, column))
    past_months = [key for key in past if key in past_by_month]
    if past_months:
        first_group = [index for index, _ in past_by_month[past_months[0]]]
        if len(past_months) >= 2:
            for month in past_months[1:-1]:
                first_group.extend(
                    index for index, col in past_by_month[month] if col["kind"] != "f_sellthrough"
                )
            last_detail = [
                index for index, col in past_by_month[past_months[-1]]
                if col["kind"] != "f_sellthrough"
            ]
        else:
            last_detail = []
            first_group = [
                index for index, col in past_by_month[past_months[0]]
                if col["kind"] != "f_sellthrough"
            ]
        if first_group:
            _collapse_columns(ws, letters, first_group, min(first_group) - 1)
        if last_detail:
            _collapse_columns(ws, letters, last_detail, min(last_detail) - 1)

    # 条件格式照原表：单元格包含 + bgColor（判定三色、是/否、售罄率 0.9~1.1）
    verdict_states = (("库存满足", CF_DXF_GREEN), ("及时入库", CF_DXF_AMBER), ("需补货", CF_DXF_PINK))
    for index, column in enumerate(columns, start=1):
        letter = letters[index]
        anchor = f"{letter}{subtotal_row}"
        rng = f"{anchor}:{letter}{last_row}"
        if column["kind"] in ("f_verdict_now", "f_verdict"):
            for text, dxf in verdict_states:
                ws.conditional_formatting.add(rng, _contains_text_rule(text, dxf, anchor))
        elif column["kind"] == "f_enough":
            for text, dxf in (("是", CF_DXF_GREEN), ("否", CF_DXF_PINK)):
                ws.conditional_formatting.add(rng, _contains_text_rule(text, dxf, anchor))
        elif column["kind"] == "f_sellthrough":
            ws.conditional_formatting.add(rng, CellIsRule(
                operator="between", formula=["0.9", "1.1"],
                fill=CF_DXF_PINK.fill, font=CF_DXF_PINK.font,
            ))

    ws.freeze_panes = f"{letter_of('f_progress')}{first_data}"
    ws.auto_filter.ref = f"A{leaf_row}:{letters[len(columns)]}{last_row}"
    book.save(dest)
    return dest


def build_production_plan(source_path, env_path: str, output_path=None, *,
                          today: date | None = None) -> dict:
    today = today or business_today()
    source = read_source_plan(source_path)
    roster = merge_roster(load_tagged_styles(env_path), source["styles"])
    source["styles"] = roster["styles"]
    styles = [item["styleId"] for item in source["styles"]]
    live = load_live_by_style(env_path, styles, today=today)
    dest = Path(output_path) if output_path else plan_workbook_path(today=today)
    write_production_plan(
        source, live, dest, today=today,
        returns_through=load_returns_through(env_path),
    )
    alerts = collect_month_alerts(source, live, today)
    missing_demand = [
        item["styleId"] for item in source["styles"]
        if not (source["demands"].get(item["styleId"]) or {})
    ]
    return {
        "ok": True,
        "styles": len(styles),
        "added": roster["added"],
        "dropped": roster["dropped"],
        "missingDemand": missing_demand,
        "xlsx": str(dest),
        "today": today.isoformat(),
        "alerts": alerts,
    }
