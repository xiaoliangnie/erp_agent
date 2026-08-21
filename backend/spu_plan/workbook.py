# -*- coding: utf-8 -*-
"""写出员工总表列顺序与版式的鞋服 SPU Excel（对齐示例表格式）。"""
from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, Rule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils import get_column_letter

from ..business_time import business_today
from ..paths import local_dir

TOTAL_COLUMNS = [
    "品类线", "款式编码", "品名", "是否活动", "是否淘汰", "生产模式",
    "SKU", "年份", "商品季节", "分类", "基本售价", "成本价",
    "昨天销量", "3天销量", "7天销量", "15天销量", "30天销量", "45天销量", "60天销量",
    # 百货总表在 7 天后插 14 天，30 天后插 7/15/30 渠道，见 columns_for()
    "前7天合计", "近7天合计", "周环比销量", "已断码数量", "即将缺码数量",
    "周转天数（包含采购在途）", "缺货风险", "补货建议数", "日平均销量", "周转/天",
    "总库存", "实际库存", "订单占有", "采购在途", "备注", "标签",
]

FONT_NAME = "微软雅黑"
RED = "FFFF0000"
HEADER_FILL = PatternFill("solid", fgColor="F4B482")
# 示例表头里红字强调的列
HEADER_RED_COLUMNS = {
    "已断码数量", "即将缺码数量", "缺货风险",
    "日平均销量", "周转/天", "总库存", "标签",
}
# 数据行整列红字的列
DATA_RED_COLUMNS = {"缺货风险", "标签"}
# 表头与数据行左对齐（其余居中）
LEFT_COLUMNS = {"品名", "商品名称", "是否活动", "是否淘汰"}
NUMBER_FORMATS = {
    "周环比销量": "0.0%",
    "已断码数量": "0_ ",
    "即将缺码数量": "0_ ",
    "周转天数（包含采购在途）": "0.0_ ",
    "补货建议数": "0.0_ ",
    "日平均销量": "0.0_ ",
    "周转/天": "0.0_ ",
}
THIN = Side(style="thin")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
# 列宽照抄示例，但按列名映射，加减列不会错位；没写的列用示例默认 8.73
COLUMN_WIDTHS = {
    "品类线": 10.25, "款式编码": 19.38, "品名": 32.0, "是否活动": 4.12,
    "是否淘汰": 9.42, "生产模式": 7.12, "SKU": 4.46, "年份": 4.75,
    "商品季节": 7.18, "分类": 8.62, "成本价": 5.75,
    "7天销量": 9.0, "15天销量": 8.0, "30天销量": 8.25, "45天销量": 9.88,
    "14天销量": 8.25,
    "线上7天": 8.0, "线下7天": 8.0,
    "线上15天": 8.0, "线下15天": 8.0,
    "线上30天": 8.25, "线下30天": 8.25,
    "进货仓库存": 8.25,
    "前7天合计": 8.62, "近7天合计": 8.62, "周环比销量": 8.62,
    "已断码数量": 6.0, "即将缺码数量": 6.62,
    "周转天数（包含采购在途）": 7.75, "缺货风险": 8.09, "补货建议数": 8.09,
    "日平均销量": 8.09, "总库存": 8.09, "备注": 11.0, "标签": 15.0,
}
DEFAULT_COLUMN_WIDTH = 8.73
HEADER_ROW_HEIGHT = 47.0
DATA_ROW_HEIGHT = 20.0
# 条件格式配色沿用示例（Excel 内置「浅红/浅黄」）
CF_RED_FILL = PatternFill("solid", fgColor="FFC7CE")
CF_RED_FONT = Font(name=FONT_NAME, size=8, color="9C0006")
CF_YELLOW_FILL = PatternFill("solid", fgColor="FFEB9C")
CF_YELLOW_FONT = Font(name=FONT_NAME, size=8, color="9C6500")
TURNOVER_HIGHLIGHT_DAYS = 35


def columns_for(board: str = "apparel") -> list[str]:
    """鞋服列不动；百货用商品编码/名称，加 14 天和 7/15/30 渠道，去掉断码/缺码。"""
    if board == "baihuo":
        columns = list(TOTAL_COLUMNS)
        columns[columns.index("款式编码")] = "商品编码"
        columns[columns.index("品名")] = "商品名称"
        columns.insert(columns.index("7天销量") + 1, "14天销量")
        after_30 = columns.index("30天销量") + 1
        columns[after_30:after_30] = [
            "线上7天", "线下7天", "线上15天", "线下15天", "线上30天", "线下30天",
        ]
        columns.insert(columns.index("采购在途") + 1, "进货仓库存")
        for header in ("已断码数量", "即将缺码数量"):
            columns.remove(header)
        return columns
    return list(TOTAL_COLUMNS)


def style_workbook_filename(today: date | None = None, *, board: str = "apparel") -> str:
    day = today or business_today()
    suffix = "自营百货总表" if board == "baihuo" else "鞋服SPU总表"
    return f"{day.strftime('%y%m%d')}-{suffix}.xlsx"


def style_workbook_path(*, root=None, today: date | None = None, board: str = "apparel") -> Path:
    return local_dir("outputs", root=root) / "spu" / style_workbook_filename(today, board=board)


def _row_values(item: dict, columns: list[str]) -> list:
    turnover = item.get("turnoverDisplay")
    if turnover is None:
        turnover = "-" if item.get("turnoverDays") is None else item.get("turnoverDays")
    by_header = {
        "品类线": item.get("categoryLine") or "",
        "款式编码": item.get("styleId") or "",
        "商品编码": item.get("styleId") or "",
        "品名": item.get("name") or "",
        "商品名称": item.get("name") or "",
        "是否活动": "",
        "是否淘汰": item.get("obsoleteLabel") or "",
        "生产模式": item.get("productionMode") or "",
        "SKU": item.get("skuCount") or 0,
        "年份": item.get("year") or "",
        "商品季节": item.get("season") or "",
        "分类": item.get("category") or "",
        "基本售价": item.get("salePrice") or 0,
        "成本价": item.get("costPrice") or 0,
        "昨天销量": item.get("sales1") or 0,
        "3天销量": item.get("sales3") or 0,
        "7天销量": item.get("sales7") or 0,
        "14天销量": item.get("sales14") or 0,
        "15天销量": item.get("sales15") or 0,
        "30天销量": item.get("sales30") or 0,
        "45天销量": item.get("sales45") or 0,
        "60天销量": item.get("sales60") or 0,
        "线上7天": item.get("sales7Online") or 0,
        "线下7天": item.get("sales7Offline") or 0,
        "线上15天": item.get("sales15Online") or 0,
        "线下15天": item.get("sales15Offline") or 0,
        "线上30天": item.get("sales30Online") or 0,
        "线下30天": item.get("sales30Offline") or 0,
        "前7天合计": item.get("salesPrev7") or 0,
        "近7天合计": item.get("sales7") or 0,
        "周环比销量": "-" if item.get("wowRatio") is None else item.get("wowRatio"),
        "已断码数量": item.get("brokenSkus") or 0,
        "即将缺码数量": item.get("shortSkus") or 0,
        "周转天数（包含采购在途）": turnover,
        "缺货风险": item.get("stockoutLabel") or "",
        "补货建议数": item.get("replenishQty"),
        "日平均销量": item.get("dailyAvg") or 0,
        "周转/天": "" if item.get("turnoverDays") is None else item.get("turnoverDays"),
        "总库存": item.get("onHand") or 0,
        "实际库存": item.get("qty") or 0,
        "订单占有": item.get("occupy") or 0,
        "采购在途": item.get("inbound") or 0,
        "进货仓库存": item.get("inQty") or 0,
        "备注": item.get("remark") or "",
        "标签": "，".join(item.get("labels") or []),
    }
    return [by_header.get(header, "") for header in columns]


def _column_letter(header: str, columns: list[str]) -> str:
    return get_column_letter(columns.index(header) + 1)


def _apply_conditional_formats(sheet, last_row: int, columns: list[str]) -> None:
    if last_row < 2:
        return
    red = {"fill": CF_RED_FILL, "font": CF_RED_FONT}
    for header in ("已断码数量", "即将缺码数量"):
        if header not in columns:
            continue
        letter = _column_letter(header, columns)
        sheet.conditional_formatting.add(
            f"{letter}2:{letter}{last_row}",
            CellIsRule(operator="greaterThan", formula=["0"], **red),
        )
    turn = _column_letter("周转天数（包含采购在途）", columns)
    sheet.conditional_formatting.add(
        f"{turn}2:{turn}{last_row}",
        CellIsRule(
            operator="lessThan", formula=[str(TURNOVER_HIGHLIGHT_DAYS)],
            fill=CF_YELLOW_FILL, font=CF_YELLOW_FONT,
        ),
    )
    id_header = "商品编码" if "商品编码" in columns else "款式编码"
    style_col = _column_letter(id_header, columns)
    sheet.conditional_formatting.add(
        f"{style_col}2:{style_col}{last_row}",
        Rule(
            type="duplicateValues",
            dxf=DifferentialStyle(fill=CF_RED_FILL, font=CF_RED_FONT),
        ),
    )


def write_style_workbook(result: dict, path) -> Path:
    """只写「总表」一张，版式对齐员工示例。手填列留空。"""
    dest = Path(path)
    dest.parent.mkdir(parents=True, exist_ok=True)
    board = "baihuo" if result.get("board") == "baihuo" else "apparel"
    columns = columns_for(board)
    book = Workbook()
    sheet = book.active
    sheet.title = "总表"

    for col, header in enumerate(columns, start=1):
        cell = sheet.cell(1, col, header)
        cell.font = Font(
            name=FONT_NAME, size=9, bold=True,
            color=RED if header in HEADER_RED_COLUMNS else "FF000000",
        )
        cell.fill = HEADER_FILL
        cell.border = BORDER
        cell.alignment = Alignment(
            horizontal="left" if header in LEFT_COLUMNS else "center",
            vertical="center", wrap_text=True,
        )
        letter = get_column_letter(col)
        sheet.column_dimensions[letter].width = COLUMN_WIDTHS.get(header, DEFAULT_COLUMN_WIDTH)
    sheet.row_dimensions[1].height = HEADER_ROW_HEIGHT

    styles = result.get("styles") or []
    black_font = Font(name=FONT_NAME, size=8)
    red_font = Font(name=FONT_NAME, size=8, color=RED)
    center = Alignment(horizontal="center", vertical="center")
    left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for row_index, item in enumerate(styles, start=2):
        values = _row_values(item, columns)
        for col, (header, value) in enumerate(zip(columns, values), start=1):
            cell = sheet.cell(row_index, col, value)
            cell.font = red_font if header in DATA_RED_COLUMNS else black_font
            cell.border = BORDER
            cell.alignment = left_wrap if header in LEFT_COLUMNS else center
            fmt = NUMBER_FORMATS.get(header)
            if fmt and value is not None and value != "" and not isinstance(value, str):
                cell.number_format = fmt
        sheet.row_dimensions[row_index].height = DATA_ROW_HEIGHT

    last_row = max(1, 1 + len(styles))
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{last_row}"
    sheet.freeze_panes = "Q2"
    _apply_conditional_formats(sheet, last_row, columns)
    book.save(dest)
    return dest
