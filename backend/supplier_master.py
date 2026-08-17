# -*- coding: utf-8 -*-
"""本机维护的供应商主数据。只读 Excel，不上库。

ERP 采购单的 seller 是简称。合同要的全称、地址、联系人、票种税率从这张表补。
员工覆盖导出文件即可，服务按修改时间重读，不必重启。缺文件时回退
``files/config/suppliers.json``，方便离线用例。
"""
from __future__ import annotations

import json
import os
import re
from dataclasses import dataclass, field
from datetime import datetime
from functools import lru_cache
from pathlib import Path

from openpyxl import load_workbook

from .contract_mappings import load_mappings
from .database import load_all_env
from .paths import CONFIG_DIR, ROOT, local_dir, resolve_repo_path

DEFAULT_XLSX = CONFIG_DIR / "供应商管理.xlsx"
FALLBACK_JSON = CONFIG_DIR / "suppliers.json"
INTERNAL_JSON = CONFIG_DIR / "internal_suppliers.json"
INTERNAL_ADDRESS = "内部往来"
INTERNAL_CONTACT = "内部"
PROJECT_ENV = load_all_env(ROOT / ".env") if (ROOT / ".env").exists() else {}
SHEET_NAME = "供应商管理"
REQUIRED_FIELDS = ("legal_name", "address", "contact_name", "contact_phone")
REQUIRED_HEADERS = (
    "编码", "全称", "简称", "是否冻结", "发票类型", "结算方式",
    "联系人", "联系电话", "公司地址",
)
PLACEHOLDERS = {"", "1", "-", "/", "无", "none", "null"}
INVOICE_LABELS = {
    "no_invoice": "不开票",
    "normal_invoice": "普票",
    "special_invoice": "专票",
}
RATE_RE = re.compile(r"([0-9]+(?:\.[0-9]+)?)\s*%")


def supplier_master_setting(name, default=""):
    return os.environ.get(name, PROJECT_ENV.get(name, default))


def resolve_supplier_master_path(path=None, root=None) -> Path:
    """显式路径 → 环境变量 → ``<root>/files/config/供应商管理.xlsx`` → JSON 回退。"""
    if path:
        resolved = Path(path)
        if not resolved.is_absolute():
            resolved = resolve_repo_path(path, root=root)
        if not resolved.is_file():
            raise FileNotFoundError(f"找不到供应商主数据：{resolved}")
        return resolved
    if root is None:
        configured = str(supplier_master_setting("SUPPLIER_MASTER_XLSX") or "").strip()
        if configured:
            resolved = Path(configured)
            if not resolved.is_absolute():
                resolved = resolve_repo_path(configured)
            if not resolved.is_file():
                raise FileNotFoundError(f"SUPPLIER_MASTER_XLSX 指向的文件不存在：{resolved}")
            return resolved
    config_dir = local_dir("config", root=root)
    xlsx = config_dir / "供应商管理.xlsx"
    fallback = config_dir / "suppliers.json"
    if xlsx.is_file():
        return xlsx
    if fallback.is_file():
        return fallback
    raise FileNotFoundError(
        "供应商主数据未找到：请把 ERP 导出的「供应商管理」表放到 "
        f"{xlsx}（不上库，本机维护）"
    )


def cell_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def meaningful(value) -> str:
    text = cell_text(value)
    if text.lower() in PLACEHOLDERS:
        return ""
    return text


def parse_created_at(value) -> datetime | None:
    """供应商表「创建时间」。认 datetime 或常见字符串；认不出当没有。"""
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        try:
            from openpyxl.utils.datetime import from_excel
            return from_excel(value)
        except Exception:
            return None
    text = cell_text(value)
    if not text:
        return None
    for size, fmt in ((19, "%Y-%m-%d %H:%M:%S"), (16, "%Y-%m-%d %H:%M"), (10, "%Y-%m-%d")):
        if len(text) >= size:
            try:
                return datetime.strptime(text[:size], fmt)
            except ValueError:
                continue
    return None


def is_newer_record(incoming: dict, existing: dict) -> bool:
    """创建时间更近的覆盖旧行；都没有时间则后读到的覆盖。"""
    incoming_at = incoming.get("created_at")
    existing_at = existing.get("created_at")
    if incoming_at and existing_at:
        return incoming_at >= existing_at
    if incoming_at and not existing_at:
        return True
    if existing_at and not incoming_at:
        return False
    return True


def dedupe_supplier_records(records: list[dict]) -> list[dict]:
    """同一简称只留创建时间最近的一条。"""
    chosen: dict[str, dict] = {}
    for record in records:
        key = record.get("short_name") or ""
        if not key:
            continue
        previous = chosen.get(key)
        if previous is None or is_newer_record(record, previous):
            chosen[key] = record
    return list(chosen.values())


def parse_invoice_type(raw, aliases=None) -> tuple[str | None, float | None]:
    """把「专用发票(13%)」收成票种键和税率。认不出票种就只回税率，不猜。"""
    text = cell_text(raw)
    if not text:
        return None, None
    rate = None
    matched = RATE_RE.search(text)
    if matched:
        rate = float(matched.group(1))
    if aliases is None:
        aliases = load_mappings()["invoice_types"]
    kind = None
    for label, key in aliases.items():
        if label in text:
            kind = key
            break
    return kind, rate


def invoice_rates_for(kind, rate) -> dict:
    rates = {key: None for key in INVOICE_LABELS}
    if kind is not None and rate is not None:
        rates[kind] = rate
    return rates


def row_record(row: dict, aliases=None) -> dict | None:
    short = meaningful(row.get("简称"))
    legal = meaningful(row.get("全称"))
    if not short and not legal:
        return None
    kind, rate = parse_invoice_type(row.get("发票类型"), aliases)
    address = meaningful(row.get("公司地址")) or meaningful(row.get("地址电话"))
    return {
        "code": meaningful(row.get("编码")),
        "short_name": short or legal,
        "legal_name": legal,
        "address": address,
        "contact_name": meaningful(row.get("联系人")),
        "contact_phone": meaningful(row.get("联系电话")),
        "bank_account_name": meaningful(row.get("付款账户名")),
        "bank_name": meaningful(row.get("开户行")),
        "bank_account": meaningful(row.get("账户")),
        "created_at": parse_created_at(row.get("创建时间")),
        "frozen": cell_text(row.get("是否冻结")) == "冻结",
        "invoice_label": cell_text(row.get("发票类型")),
        "settlement": meaningful(row.get("结算方式")),
        "erp_price_mode": kind,
        "invoice_rates": invoice_rates_for(kind, rate),
        "source": "excel",
    }


def missing_supplier_fields(supplier: dict | None) -> list[str]:
    if not supplier:
        return list(REQUIRED_FIELDS)
    if supplier.get("internal"):
        return []
    return [key for key in REQUIRED_FIELDS if not meaningful(supplier.get(key))]


def supplier_issue(short_name: str, supplier: dict | None) -> str:
    if not supplier:
        return f"供应商简称“{short_name}”尚未维护完整信息"
    if supplier.get("internal"):
        return ""
    if supplier.get("frozen"):
        return f"供应商“{short_name}”已冻结，不能生成合同"
    missing = missing_supplier_fields(supplier)
    labels = {
        "legal_name": "全称",
        "address": "公司地址",
        "contact_name": "联系人",
        "contact_phone": "联系电话",
    }
    if missing:
        return f"供应商“{short_name}”缺少字段：{', '.join(labels[key] for key in missing)}"
    return ""


@dataclass
class SupplierBook:
    path: Path
    source: str
    by_short: dict = field(default_factory=dict)
    by_legal: dict = field(default_factory=dict)
    by_code: dict = field(default_factory=dict)

    def lookup(self, seller: str):
        key = cell_text(seller)
        if not key:
            return None
        for index in (self.by_short, self.by_legal, self.by_code):
            hit = index.get(key)
            if hit:
                return hit
        stripped = key.lstrip("&")
        alt = f"&{stripped}" if not key.startswith("&") else stripped
        candidates = []
        for candidate in dict.fromkeys((stripped, alt)):
            if candidate and candidate != key and candidate in self.by_short:
                candidates.append(self.by_short[candidate])
        if len(candidates) == 1:
            return candidates[0]
        return None

    def names(self) -> set[str]:
        """品控解析用的简称集合；带 & 的同时收一份去掉前缀的键。"""
        out = set()
        for key in self.by_short:
            out.add(key)
            stripped = key.lstrip("&")
            if stripped:
                out.add(stripped)
        return out

    def as_dict(self) -> dict:
        return dict(self.by_short)


def _index_record(book: SupplierBook, record: dict, *, collide: str):
    short = record["short_name"]
    if short in book.by_short:
        raise ValueError(f"{collide}简称“{short}”重复")
    book.by_short[short] = record
    legal = record.get("legal_name") or ""
    if legal and legal not in book.by_legal:
        book.by_legal[legal] = record
    code = record.get("code") or ""
    if code and code not in book.by_code:
        book.by_code[code] = record


def _load_json_book(path: Path) -> SupplierBook:
    payload = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(payload, dict):
        raise ValueError(f"{path.name} 必须是对象")
    book = SupplierBook(path=path, source="json")
    for key, raw in payload.items():
        short = cell_text(key)
        if not short or not isinstance(raw, dict):
            continue
        record = {
            "code": "",
            "short_name": short,
            "legal_name": meaningful(raw.get("legal_name")),
            "address": meaningful(raw.get("address")),
            "contact_name": meaningful(raw.get("contact_name")),
            "contact_phone": meaningful(raw.get("contact_phone")),
            "bank_account_name": meaningful(raw.get("bank_account_name")),
            "bank_name": meaningful(raw.get("bank_name")),
            "bank_account": meaningful(raw.get("bank_account")),
            "created_at": parse_created_at(raw.get("created_at")),
            "frozen": bool(raw.get("frozen")),
            "invoice_label": cell_text(raw.get("invoice_label")),
            "settlement": meaningful(raw.get("settlement")),
            "erp_price_mode": raw.get("erp_price_mode"),
            "invoice_rates": dict(raw.get("invoice_rates") or {}),
            "source": "json",
        }
        _index_record(book, record, collide=f"{path.name} ")
    return book


def _header_map(headers) -> dict[str, int]:
    mapping = {}
    for index, header in enumerate(headers):
        name = cell_text(header)
        if name:
            mapping[name] = index
    missing = [name for name in REQUIRED_HEADERS if name not in mapping]
    if missing:
        raise ValueError("供应商管理表缺少列：" + "、".join(missing))
    return mapping


def _load_excel_book(path: Path) -> SupplierBook:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[SHEET_NAME] if SHEET_NAME in workbook.sheetnames else workbook[workbook.sheetnames[0]]
        rows = sheet.iter_rows(values_only=True)
        try:
            headers = next(rows)
        except StopIteration as exc:
            raise ValueError(f"{path.name} 是空表") from exc
        columns = _header_map(headers)
        book = SupplierBook(path=path, source="excel")
        aliases = load_mappings()["invoice_types"]
        records = []
        for raw in rows:
            row = {name: raw[index] if index < len(raw) else None for name, index in columns.items()}
            record = row_record(row, aliases)
            if record:
                records.append(record)
        for record in dedupe_supplier_records(records):
            _index_record(book, record, collide=f"{path.name} ")
    finally:
        workbook.close()
    return book


def _internal_path(path=None, root=None) -> Path:
    if root is not None:
        return local_dir("config", root=root) / "internal_suppliers.json"
    if path is not None:
        return Path(path).resolve().parent / "internal_suppliers.json"
    return INTERNAL_JSON


def _internal_record(short: str, raw: dict | None) -> dict:
    raw = raw if isinstance(raw, dict) else {}
    label = meaningful(raw.get("label")) or short
    return {
        "code": "",
        "short_name": short,
        "legal_name": label,
        "address": INTERNAL_ADDRESS,
        "contact_name": INTERNAL_CONTACT,
        "contact_phone": "",
        "bank_account_name": "",
        "bank_name": "",
        "bank_account": "",
        "created_at": None,
        "frozen": False,
        "invoice_label": "内部往来",
        "settlement": "",
        "erp_price_mode": None,
        "invoice_rates": {},
        "source": "internal",
        "internal": True,
    }


def _load_internal_records(path: Path) -> list[dict]:
    if not path.is_file():
        return []
    payload = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(payload, dict):
        raise ValueError(f"{path.name} 必须是对象，键为 ERP 供应商简称")
    records = []
    for key, raw in payload.items():
        short = cell_text(key)
        if short:
            records.append(_internal_record(short, raw if isinstance(raw, dict) else {}))
    return records


def _apply_internal(book: SupplierBook, records: list[dict]):
    """内部户覆盖同名外部行：不写出收付款信息，也不要求 Excel 全称。"""
    for record in records:
        book.by_short[record["short_name"]] = record


@lru_cache(maxsize=8)
def _cached_book(resolved: str, mtime: float, internal_path: str, internal_mtime: float) -> SupplierBook:
    path = Path(resolved)
    if path.suffix.lower() == ".json":
        book = _load_json_book(path)
    else:
        book = _load_excel_book(path)
    if internal_path:
        _apply_internal(book, _load_internal_records(Path(internal_path)))
    return book


def load_supplier_book(path=None, root=None) -> SupplierBook:
    resolved = resolve_supplier_master_path(path, root=root)
    internal = _internal_path(path, root)
    internal_mtime = internal.stat().st_mtime if internal.is_file() else 0.0
    return _cached_book(
        str(resolved.resolve()), resolved.stat().st_mtime,
        str(internal.resolve()) if internal.is_file() else "", internal_mtime,
    )


def load_suppliers(path=None, root=None) -> dict:
    """兼容旧调用方：简称 → 记录。查找请用 ``lookup_supplier``。"""
    return load_supplier_book(path, root=root).as_dict()


def lookup_supplier(seller: str, path=None, book: SupplierBook | None = None, root=None):
    current = book or load_supplier_book(path, root=root)
    return current.lookup(seller)


def load_supplier_names(root: Path | None = None) -> set[str]:
    """品控指令解析用的简称集合。"""
    try:
        return load_supplier_book(root=root).names()
    except FileNotFoundError:
        return set()


def clear_supplier_cache():
    _cached_book.cache_clear()
