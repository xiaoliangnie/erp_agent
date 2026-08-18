# -*- coding: utf-8 -*-
"""每日代发 Excel 调度。发送时刻前先抓表，到点只发；不覆盖已填表。"""
from __future__ import annotations

import logging
import re
import threading
from datetime import datetime, timezone
from pathlib import Path

from ..business_time import business_now, business_today
from .workbook import dropship_output_path


logger = logging.getLogger(__name__)


def dropship_file_has_rows(path: Path) -> bool:
    """当日文件已有明细则视为已填，定时任务不得覆盖。"""
    return dropship_row_count(path) > 0


def dropship_row_count(path: Path) -> int:
    target = Path(path)
    if not target.exists() or target.stat().st_size <= 0:
        return 0
    from openpyxl import load_workbook
    book = load_workbook(target, read_only=True, data_only=True)
    try:
        sheet = book[book.sheetnames[0]]
        count = 0
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if any(cell not in (None, "") for cell in row):
                count += 1
    finally:
        book.close()
    return count


def dropship_delivery_key(today: str) -> str:
    return f"dropship-file-{today}"


def dropship_group_key(today: str) -> str:
    return f"dropship-file-group-{today}"


def _split_names(value) -> tuple[str, ...]:
    text = str(value or "").strip()
    if not text:
        return ()
    return tuple(part for part in re.split(r"[,，、;；\s]+", text) if part)


class DailyDropshipScheduler:
    def __init__(self, *, runtime, root, env_path="", send_time: str = "14:00",
                 poll_seconds: int = 30, enabled: bool = False, sender=None,
                 audit=None, conversation_id: str = "", directory=None,
                 oto_buyers: str = "安安", oto_user_ids: str = "",
                 prepare_lead_minutes: int = 30,
                 max_attempts_per_day: int = 3, retry_interval_seconds: int = 15 * 60):
        self.runtime = runtime
        self.root = Path(root)
        self.env_path = env_path
        self.send_time = self._parse_time(send_time)
        self.prepare_lead_minutes = max(0, int(prepare_lead_minutes or 0))
        self.prepare_time = self._shift_time(self.send_time, -self.prepare_lead_minutes)
        self.poll_seconds = max(5, int(poll_seconds))
        self.enabled = bool(enabled)
        self.sender = sender
        self.audit = audit
        self.directory = directory
        self.conversation_id = str(conversation_id or "").strip()
        self.oto_buyers = _split_names(oto_buyers if oto_buyers is not None else "安安")
        self.oto_user_ids = _split_names(oto_user_ids)
        self.max_attempts_per_day = max(1, int(max_attempts_per_day))
        self.retry_interval_seconds = max(0, int(retry_interval_seconds))
        self._thread: threading.Thread | None = None
        self._stop = threading.Event()
        self.last_run = ""
        self.last_error = ""

    @staticmethod
    def _parse_time(value) -> tuple:
        try:
            hour, minute = str(value or "14:00").split(":", 1)
            return max(0, min(int(hour), 23)), max(0, min(int(minute), 59))
        except (TypeError, ValueError):
            return 14, 0

    @staticmethod
    def _shift_time(value: tuple, minutes: int) -> tuple:
        total = max(0, min(value[0] * 60 + value[1] + int(minutes), 23 * 60 + 59))
        return total // 60, total % 60

    def status(self) -> dict:
        return {
            "enabled": self.enabled,
            "running": bool(self._thread and self._thread.is_alive()),
            "prepareTime": f"{self.prepare_time[0]:02d}:{self.prepare_time[1]:02d}",
            "sendTime": f"{self.send_time[0]:02d}:{self.send_time[1]:02d}",
            "lastRun": self.last_run,
            "lastError": self.last_error,
            "dingtalk": bool(self._can_send()),
            "otoBuyers": list(self.oto_buyers),
        }

    def start(self) -> None:
        if not self.enabled:
            return
        if self._thread and self._thread.is_alive():
            return
        self._stop.clear()
        self._thread = threading.Thread(target=self._loop, name="dropship-export", daemon=True)
        self._thread.start()

    def stop(self) -> None:
        self._stop.set()

    def _loop(self) -> None:
        while not self._stop.wait(self.poll_seconds):
            try:
                self.tick()
            except Exception as exc:
                self.last_error = f"{type(exc).__name__}: {exc}"
                logger.exception("Dropship scheduler tick failed")

    def tick(self, *, now: datetime | None = None) -> dict:
        current = now or business_now()
        today = current.date().isoformat()
        clock = (current.hour, current.minute)
        if clock < self.prepare_time:
            return {"skipped": True, "reason": "未到准备时间", "today": today}
        if clock < self.send_time:
            return self._prepare_tick(today=current.date())
        if self.last_run == today:
            return {"skipped": True, "reason": "今日已成功发送", "today": today}
        decision = self._retry_decision(today, current=current)
        if not decision["allowed"]:
            return {"skipped": True, "reason": decision["reason"], "today": today}
        result = self.run_once(today=current.date(), send=True)
        if result.get("sent") or (result.get("ok") and not result.get("prepared")):
            self.last_run = today
            self.last_error = ""
        elif result.get("skipped") and result.get("reason") in {
            "当日已填表，不覆盖", "同一天的代发表已经发过",
        }:
            if result.get("reason") == "同一天的代发表已经发过":
                self.last_run = today
            self.last_error = ""
        if result.get("failed"):
            self.last_error = str(result.get("reason") or "代发发送失败")
        return result

    def _prepare_tick(self, *, today) -> dict:
        path = dropship_output_path(root=self.root, today=today)
        if dropship_file_has_rows(path):
            self.last_error = ""
            return {
                "skipped": True, "reason": "表已备好，等待发送",
                "today": today.isoformat() if hasattr(today, "isoformat") else str(today),
                "path": str(path),
            }
        result = self.run_once(today=today, send=False)
        if result.get("failed"):
            self.last_error = str(result.get("reason") or "代发备表失败")
        else:
            self.last_error = ""
        return result

    def run_once(self, *, today=None, operator="scheduler", send: bool = True) -> dict:
        day = today or business_today()
        if hasattr(day, "isoformat") and not isinstance(day, str):
            stamp = day
        else:
            from datetime import date as date_cls
            stamp = date_cls.fromisoformat(str(day)[:10])
        today_key = stamp.isoformat()
        path = dropship_output_path(root=self.root, today=stamp)
        reused = dropship_file_has_rows(path)
        if not reused:
            if self.runtime is None:
                return {"failed": True, "reason": "ERP Digital Worker 未装配", "today": today_key}
            from .export import export_today_dropship, public_export_result
            try:
                payload = export_today_dropship(
                    self.runtime, path=path, root=self.root, env_path=self.env_path,
                )
            except Exception as exc:
                logger.exception("Dropship scheduled export failed")
                return {"failed": True, "reason": str(exc), "today": today_key}
            public = public_export_result(payload)
            path = Path(public.get("path") or path)
        if not send:
            return {
                "ok": True, "prepared": True, "sent": False, "today": today_key,
                "operator": operator, "filename": path.name, "path": str(path),
                "reused": reused, "reason": "表已备好，等待发送",
            }
        if not self._can_send():
            if reused:
                return {
                    "skipped": True, "reason": "当日已填表，不覆盖",
                    "today": today_key, "path": str(path),
                }
            return {
                "ok": True, "sent": False, "today": today_key, "operator": operator,
                "filename": path.name, "path": str(path),
                "reason": "已导出，钉钉未配置企业机器人或群会话",
            }
        return self._send_file(path, today=today_key, operator=operator, reused=reused)

    def _can_send(self) -> bool:
        sender = self.sender
        conversation_id = self.conversation_id or getattr(sender, "group_conversation_id", "")
        return bool(
            sender
            and getattr(sender, "app_ready", False)
            and conversation_id
            and hasattr(sender, "upload_media")
            and hasattr(sender, "send_file")
        )

    def _oto_targets(self) -> tuple[list[str], list[str]]:
        """返回 (钉钉 userId, 尚未绑定的采购员名)。没有身份目录时不拦群发。"""
        extra = [item for item in self.oto_user_ids if item]
        if self.directory is None:
            return extra, []
        if not self.oto_buyers:
            return extra, []
        resolve = getattr(self.directory, "resolve", None)
        if resolve is None:
            return extra, list(self.oto_buyers)
        result = resolve(self.oto_buyers)
        ids = list(extra)
        for user_id in result.get("userIds") or []:
            if user_id and user_id not in ids:
                ids.append(user_id)
        return ids, list(result.get("unbound") or [])

    def _record_failed(self, key: str, *, target: str, detail: dict, error: str) -> None:
        if self.audit is None:
            return
        self.audit.record_delivery(
            channel="dingtalk", target=target, kind="dropship_file",
            status="failed", detail=detail, error=error,
            idempotency_key=self.audit.next_attempt_key(key),
        )

    def _record_sent(self, key: str, *, target: str, detail: dict) -> None:
        if self.audit is None:
            return
        if self.audit.has_successful_delivery(key):
            return
        self.audit.release_unsuccessful_key(key)
        self.audit.record_delivery(
            channel="dingtalk", target=target, kind="dropship_file",
            status="sent", detail=detail, idempotency_key=key,
        )

    def _send_file(self, path: Path, *, today: str, operator: str, reused: bool) -> dict:
        key = dropship_delivery_key(today)
        group_key = dropship_group_key(today)
        if self.audit is not None and self.audit.has_successful_delivery(key):
            return {
                "skipped": True, "reason": "同一天的代发表已经发过",
                "today": today, "path": str(path),
            }
        rows = dropship_row_count(path)
        if not path.exists() or path.stat().st_size <= 0:
            return {"failed": True, "reason": "当日代发表不存在", "today": today}
        title = f"代发未安排 · {today}"
        if rows:
            text = f"今日代发未安排 {rows} 行，文件：{path.name}"
        else:
            text = f"今日无代发未安排（{today}）。文件：{path.name}"
        detail = {
            "today": today, "operator": operator, "path": str(path),
            "filename": path.name, "rows": rows, "reused": reused,
        }
        conversation_id = self.conversation_id or getattr(self.sender, "group_conversation_id", "") or "group"
        oto_ids, unbound = self._oto_targets()
        oto_required = bool(self.directory is not None and self.oto_buyers)
        markdown = {}
        file_info = {}
        oto_info = {}
        group_already = self.audit is not None and self.audit.has_successful_delivery(group_key)
        try:
            if not group_already:
                markdown = self.sender.send_markdown(title, text)
                if rows:
                    media = self.sender.upload_media(path, filetype="file")
                    file_info = self.sender.send_file(
                        conversation_id, media["mediaId"], path.name, file_type="xlsx",
                    )
                self._record_sent(
                    group_key, target=conversation_id,
                    detail={**detail, "channel": "app"},
                )
            if oto_required and unbound and not oto_ids:
                reason = f"{'、'.join(unbound)}还没绑定钉钉，群已发，私聊未发出"
                self._record_failed(key, target="oto", detail=detail, error=reason)
                return {"failed": True, "reason": reason, "today": today, "path": str(path)}
            if oto_ids:
                oto_info = {"markdown": self.sender.send_oto_markdown(title, text, user_ids=oto_ids)}
                if rows:
                    media = self.sender.upload_media(path, filetype="file")
                    oto_info["file"] = self.sender.send_oto_file(
                        oto_ids, media["mediaId"], path.name, file_type="xlsx",
                    )
        except Exception as exc:
            logger.exception("Dropship DingTalk send failed")
            self._record_failed(key, target=conversation_id, detail=detail, error=str(exc))
            return {"failed": True, "reason": str(exc), "today": today, "path": str(path)}
        self._record_sent(
            key, target=conversation_id,
            detail={**detail, "channel": "app", "oto": bool(oto_ids)},
        )
        return {
            "ok": True, "sent": True, "today": today, "operator": operator,
            "filename": path.name, "path": str(path), "rows": rows,
            "reused": reused, "markdown": markdown, "file": file_info,
            "oto": oto_info, "otoUserCount": len(oto_ids),
        }

    def _retry_decision(self, today: str, *, current: datetime) -> dict:
        list_fn = getattr(self.audit, "list_deliveries", None) if self.audit is not None else None
        if not list_fn:
            return {"allowed": True, "reason": ""}
        attempts = list_fn(key_prefix=f"{dropship_delivery_key(today)}-attempt-", status="failed")
        if len(attempts) >= self.max_attempts_per_day:
            return {"allowed": False, "reason": f"今日已失败 {len(attempts)} 次，等次日再试"}
        if attempts and self.retry_interval_seconds:
            last_at = _parse_created_at(attempts[-1].get("createdAt"))
            if last_at is not None:
                elapsed = (current.astimezone(timezone.utc) - last_at).total_seconds()
                if elapsed < self.retry_interval_seconds:
                    return {"allowed": False, "reason": "距上次失败不足重试间隔"}
        return {"allowed": True, "reason": ""}


def _parse_created_at(value) -> datetime | None:
    text = str(value or "").strip()
    if not text:
        return None
    if text.endswith("Z"):
        text = text[:-1] + "+00:00"
    try:
        parsed = datetime.fromisoformat(text)
    except ValueError:
        return None
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=timezone.utc)
    return parsed.astimezone(timezone.utc)
