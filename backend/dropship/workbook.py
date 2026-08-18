# -*- coding: utf-8 -*-
"""代发订单 Excel：列名对齐 8.15代发.xlsx。

Sheet1 是完整导出列，样例里隐藏的列这里同样隐藏；
Sheet2 是给供应商的收货 + 商品子集。
"""
from __future__ import annotations

import json
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..business_time import BUSINESS_TIMEZONE, business_now, business_today
from ..paths import TEMPLATES_DIR, local_dir

SHEET1_COLUMNS = [
    "内部订单号", "线上订单号", "店铺名称", "店铺简称", "店铺分组",
    "商品编码", "店铺主账号", "订单类型", "平台站点", "付款日期",
    "收货人", "省份", "城市", "区县", "地址(包含省市区)", "手机",
    "买家留言", "订单备注", "供应商", "标准商品名", "供应商款号",
    "供应商商品编码", "颜色及规格", "快递公司", "快递单号", "数量",
    "商品裸价", "成本价",
]
SHEET2_COLUMNS = [
    "内部订单号", "收货人", "省份", "城市", "区县", "地址(包含省市区)",
    "手机", "买家留言", "订单备注", "供应商", "标准商品名", "供应商款号",
    "供应商商品编码", "颜色及规格", "快递公司", "快递单号", "数量",
]
SHEET1_HIDDEN = {"线上订单号", "省份", "买家留言", "标准商品名"}
COLUMNS = SHEET1_COLUMNS
TEXT_COLUMNS = {
    "内部订单号", "线上订单号", "商品编码", "手机", "快递单号",
    "供应商商品编码", "供应商款号",
}
NUMBER_COLUMNS = {"数量", "商品裸价", "成本价"}
HEADER_FONT = Font(name="宋体", size=11)
HEADER_FILL = PatternFill("solid", fgColor="E6E6E6")
HEADER_BORDER = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)
COLUMN_WIDTHS = {
    "内部订单号": 18,
    "线上订单号": 18,
    "收货人": 18,
    "省份": 18,
    "地址(包含省市区)": 18,
    "买家留言": 18,
    "供应商": 18,
    "标准商品名": 18,
    "快递单号": 18,
    "数量": 6,
    "商品裸价": 18,
}
TEMPLATE_NAME = "代发订单模板.xlsx"


def dropship_filename(today: date | None = None) -> str:
    """东八区业务日，形如 260817-代发.xlsx。"""
    day = today or business_today()
    return f"{day.strftime('%y%m%d')}-代发.xlsx"


def dropship_output_path(*, root=None, today: date | None = None) -> Path:
    return local_dir("outputs", root=root) / "dropship" / dropship_filename(today)


def dropship_meta_path(workbook_path) -> Path:
    target = Path(workbook_path)
    return target.with_name(target.stem + ".meta.json")


def format_dropship_cutoff(moment=None) -> str:
    stamp = moment or business_now()
    if getattr(stamp, "tzinfo", None) is None:
        stamp = stamp.replace(tzinfo=BUSINESS_TIMEZONE)
    else:
        stamp = stamp.astimezone(BUSINESS_TIMEZONE)
    return stamp.strftime("%Y-%m-%d %H:%M")


def _clean_oids(values) -> list[str]:
    found = []
    for raw in values or []:
        oid = str(raw or "").strip()
        if oid and oid not in found:
            found.append(oid)
    return found


def write_dropship_meta(workbook_path, *, cutoff=None, rate_limited=None,
                       stats=None, rest_complete=None) -> dict:
    """抓表完成时记下截止时间和揭收货限流单号，发送通知用。不含收货明文。"""
    current = read_dropship_meta(workbook_path)
    stamp = cutoff or current.get("dataCutoff") or format_dropship_cutoff()
    limited = _clean_oids(
        rate_limited if rate_limited is not None else current.get("rateLimited")
    )
    if rest_complete is None:
        complete = bool(current.get("restComplete"))
    else:
        complete = bool(rest_complete)
    payload = {
        "dataCutoff": stamp,
        "rateLimited": limited,
        "restComplete": complete,
    }
    source = stats if stats is not None else current.get("stats")
    if isinstance(source, dict) and source:
        payload["stats"] = {
            key: source.get(key)
            for key in ("orders", "lines", "收货人", "手机", "地址")
            if source.get(key) is not None
        }
    meta = dropship_meta_path(workbook_path)
    meta.parent.mkdir(parents=True, exist_ok=True)
    meta.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return payload


def write_dropship_cutoff(workbook_path, *, cutoff=None, rate_limited=None,
                          stats=None, rest_complete=None) -> str:
    return write_dropship_meta(
        workbook_path, cutoff=cutoff, rate_limited=rate_limited,
        stats=stats, rest_complete=rest_complete,
    ).get("dataCutoff") or ""


def read_dropship_meta(workbook_path) -> dict:
    target = Path(workbook_path)
    meta = dropship_meta_path(target)
    payload = {}
    if meta.exists():
        try:
            loaded = json.loads(meta.read_text(encoding="utf-8"))
            if isinstance(loaded, dict):
                payload = loaded
        except (OSError, json.JSONDecodeError, TypeError):
            payload = {}
    cutoff = str(payload.get("dataCutoff") or "").strip()
    if not cutoff and target.exists():
        cutoff = format_dropship_cutoff(
            datetime.fromtimestamp(target.stat().st_mtime, BUSINESS_TIMEZONE)
        )
    if not cutoff:
        cutoff = format_dropship_cutoff()
    return {
        "dataCutoff": cutoff,
        "rateLimited": _clean_oids(payload.get("rateLimited")),
        "restComplete": bool(payload.get("restComplete")),
        "stats": payload.get("stats") if isinstance(payload.get("stats"), dict) else {},
    }


def read_dropship_cutoff(workbook_path) -> str:
    return str(read_dropship_meta(workbook_path).get("dataCutoff") or "")


def dropship_template_path(*, root=None) -> Path:
    if root is None:
        return TEMPLATES_DIR / TEMPLATE_NAME
    return local_dir("templates", root=root) / TEMPLATE_NAME


def _write_sheet(book, title, columns, rows, *, hidden=(), freeze="A2"):
    sheet = book.active if title == "Sheet1" else book.create_sheet(title)
    sheet.title = title
    for col, header in enumerate(columns, start=1):
        cell = sheet.cell(1, col, header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = HEADER_BORDER
        cell.alignment = Alignment(vertical="center")
        letter = get_column_letter(col)
        sheet.column_dimensions[letter].width = COLUMN_WIDTHS.get(header, 13)
        if header in hidden:
            sheet.column_dimensions[letter].hidden = True
    for index, row in enumerate(rows, start=2):
        for col, header in enumerate(columns, start=1):
            value = row.get(header)
            if value == "":
                value = None
            cell = sheet.cell(index, col, value)
            cell.alignment = Alignment(
                vertical="center",
                wrap_text=header == "地址(包含省市区)",
            )
            if header in TEXT_COLUMNS:
                cell.number_format = "@"
            elif header == "商品裸价" and value is not None:
                cell.number_format = "0.0000"
            elif header == "成本价" and value is not None:
                cell.number_format = "0.####"
    last_row = max(1, len(rows) + 1)
    last_col = get_column_letter(len(columns))
    sheet.auto_filter.ref = f"A1:{last_col}{last_row}"
    if freeze:
        sheet.freeze_panes = freeze
    return sheet


def write_dropship_workbook(rows: list[dict] | None, path, *, blank: bool = False) -> Path:
    """写入 8.15 两表；blank=True 时只留表头。"""
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    lines = [] if blank else list(rows or [])
    book = Workbook()
    _write_sheet(book, "Sheet1", SHEET1_COLUMNS, lines, hidden=SHEET1_HIDDEN, freeze="A2")
    _write_sheet(book, "Sheet2", SHEET2_COLUMNS, lines, hidden=(), freeze=None)
    book.save(path)
    return path


def write_blank_dropship_template(*, root=None, today: date | None = None, include_dated: bool = False) -> dict:
    """刷新空白母版。默认不覆盖当日已填的 YYMMDD-代发.xlsx。"""
    template = write_dropship_workbook([], dropship_template_path(root=root), blank=True)
    result = {"template": str(template), "filename": Path(template).name}
    if include_dated:
        dated = write_dropship_workbook([], dropship_output_path(root=root, today=today), blank=True)
        result["dated"] = str(dated)
        result["filename"] = dated.name
    return result
