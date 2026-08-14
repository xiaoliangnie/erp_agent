# -*- coding: utf-8 -*-
"""品控日报工作簿与摘要。"""
from __future__ import annotations

from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


HEADERS = [
    "编号", "登记时间", "登记人", "供应商", "采购单号", "SKU",
    "严重度", "问题描述", "状态", "处理备注",
]
STATUS_LABEL = {"open": "未关闭", "resolved": "已关闭", "cancelled": "已撤销"}


def build_quality_workbook(issues: list[dict], path) -> Path:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    book = Workbook()
    sheet = book.active
    sheet.title = "品控台账"
    for col, header in enumerate(HEADERS, start=1):
        cell = sheet.cell(1, col, header)
        cell.font = Font(name="Microsoft YaHei", bold=True, size=10)
    for index, item in enumerate(issues, start=2):
        values = [
            item.get("id") or "",
            item.get("createdAt") or "",
            item.get("reporter") or "",
            item.get("supplier") or "",
            item.get("poId") or "",
            item.get("sku") or "",
            item.get("severity") or "",
            item.get("description") or "",
            STATUS_LABEL.get(item.get("status"), item.get("status") or ""),
            item.get("resolution") or "",
        ]
        for col, value in enumerate(values, start=1):
            cell = sheet.cell(index, col, value)
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            if col in (1, 5, 6):
                cell.number_format = "@"
    stats_row = len(issues) + 3
    by_supplier = Counter(item.get("supplier") or "未填供应商" for item in issues)
    open_count = sum(1 for item in issues if item.get("status") == "open")
    sheet.cell(stats_row, 1, "统计")
    sheet.cell(
        stats_row, 2,
        f"共 {len(issues)} 条，未关闭 {open_count} 条；"
        + "；".join(f"{name} {count}" for name, count in sorted(by_supplier.items())),
    )
    widths = [12, 22, 12, 14, 14, 18, 10, 40, 10, 24]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    book.save(path)
    return path


def quality_report_markdown(issues: list[dict], *, today: str, historic_open: int = 0) -> str:
    by_supplier = Counter(item.get("supplier") or "未填供应商" for item in issues)
    open_count = sum(1 for item in issues if item.get("status") == "open")
    lines = [
        f"### 品控日报 · {today}",
        "",
        f"当日登记 **{len(issues)}** 条（未关闭 {open_count}）。",
    ]
    if by_supplier:
        lines.append("按供应商：" + "、".join(f"{name} {count}" for name, count in sorted(by_supplier.items())))
    if historic_open:
        lines.append(f"另有 {historic_open} 条历史问题未关闭，回复「品控查询 未关闭」查看。")
    return "\n".join(lines)
