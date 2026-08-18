# -*- coding: utf-8 -*-
"""品控台账：解析、幂等、状态机、日报、调度抗错。全程离线。"""
import tempfile
import unittest
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

from backend.agent.actions import ActionError
from backend.agent.audit import AuditLog
from backend.agent.store import AgentStore
from backend.agent.work_items import WorkItems
from backend.staff_names import VIEWER_WRITE_DENIED, WEB_OPERATOR_UNBOUND
from backend.business_time import BUSINESS_TIMEZONE, business_today
from backend.dingtalk.sender import encode_multipart
from backend.quality import (
    DailyQualityReportScheduler,
    QualityError,
    QualityLedger,
    build_quality_workbook,
    parse_quality_command,
    parse_quality_fields,
    report_link_sig,
    report_link_valid,
)
from openpyxl import load_workbook


class FakeSender:
    def __init__(self, *, fail=False, app=False):
        self.calls = []
        self.files = []
        self.fail = fail
        self.configured = True
        self.app_ready = app
        self.group_conversation_id = "cid" if app else ""

    def send_markdown(self, title, text, **kwargs):
        self.calls.append({"title": title, "text": text, **kwargs})
        if self.fail:
            raise RuntimeError("simulated send failure")
        return {"channel": "app" if self.app_ready else "webhook"}

    def upload_media(self, path, filetype="file"):
        self.files.append({"upload": str(path), "type": filetype})
        return {"mediaId": "mid-1"}

    def send_file(self, conversation_id, media_id, file_name, file_type="xlsx"):
        self.files.append({
            "conversation": conversation_id, "mediaId": media_id,
            "name": file_name, "type": file_type,
        })
        return {"ok": True}


class ParseTests(unittest.TestCase):
    def test_commands(self):
        self.assertEqual("record", parse_quality_command("品控 佰特 开胶")["action"])
        self.assertEqual("record", parse_quality_command("品控登记 色差")["action"])
        self.assertEqual("resolve", parse_quality_command("品控关闭 abcdef 已赔")["action"])
        self.assertEqual("cancel", parse_quality_command("撤销品控 abcdef")["action"])
        self.assertEqual("今天", parse_quality_command("品控查询")["query"])
        self.assertIsNone(parse_quality_command("催一下逾期单"))

    def test_agent_intent_maps_close_and_cancel(self):
        from backend.agent.intents import classify_intent
        closed = classify_intent("品控关闭 abcdef 已赔")
        self.assertEqual("resolve_quality_issue", closed.tool)
        self.assertEqual("abcdef", closed.arguments["issue_id"])
        self.assertEqual("已赔", closed.arguments["resolution"])
        cancelled = classify_intent("撤销品控 abcdef")
        self.assertEqual("cancel_quality_issue", cancelled.tool)
        self.assertEqual({"issue_id": "abcdef"}, cancelled.arguments)

    def test_fields_hit_and_do_not_guess(self):
        fields = parse_quality_fields(
            "佰特 604264 XZ25401308-101 鞋垫开胶 3 双",
            suppliers={"佰特"},
            lookup_po=lambda po_id: po_id == "604264",
        )
        self.assertEqual("佰特", fields["supplier"])
        self.assertEqual("604264", fields["po_id"])
        self.assertEqual("XZ25401308-101", fields["sku"])
        self.assertEqual("鞋垫开胶 3 双", fields["description"])

        unknown = parse_quality_fields(
            "未知厂 999999 鞋垫开胶",
            suppliers={"佰特"},
            lookup_po=lambda po_id: False,
        )
        self.assertEqual("", unknown["supplier"])
        self.assertEqual("", unknown["po_id"])
        self.assertIn("999999", unknown["description"])


class LedgerTests(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.store = AgentStore(Path(self.tmp.name) / "agent.sqlite3")
        self.ledger = QualityLedger(
            self.store, suppliers={"佰特"}, lookup_po=lambda po_id: po_id == "604264",
        )

    def tearDown(self):
        self.tmp.cleanup()

    def test_message_id_is_idempotent(self):
        first = self.ledger.record(description="开胶", message_id="msg-1", supplier="佰特")
        second = self.ledger.record(description="开胶又来一次", message_id="msg-1", supplier="佰特")
        self.assertEqual(first["id"], second["id"])
        self.assertEqual("开胶", second["description"])

    def test_cancelled_not_in_report(self):
        open_item = self.ledger.record(description="色差")
        gone = self.ledger.record(description="少件")
        self.ledger.cancel(gone["id"])
        today = business_today().isoformat()
        ids = [item["id"] for item in self.ledger.list_for_report(today)]
        self.assertIn(open_item["id"], ids)
        self.assertNotIn(gone["id"], ids)

    def test_cannot_resolve_cancelled(self):
        item = self.ledger.record(description="开胶")
        self.ledger.cancel(item["id"])
        with self.assertRaisesRegex(QualityError, "已撤销"):
            self.ledger.resolve(item["id"], "已赔")

    def test_handle_text_record_and_query(self):
        text = self.ledger.handle_text("品控 佰特 604264 鞋垫开胶", reporter="利特")
        self.assertIn("已登记品控", text)
        self.assertIn("604264", text)
        listed = self.ledger.handle_text("品控查询 今天")
        self.assertIn("鞋垫开胶", listed)


class FakeDirectory:
    def __init__(self, names, role="buyer"):
        self.names = set(names)
        self.role = role

    def known_operator(self, operator):
        return str(operator or "") in self.names

    def find_binding(self, *, operator="", actor_id=""):
        del actor_id
        if str(operator or "") not in self.names:
            return {}
        return {"role": self.role, "buyerName": operator}


class WorkbenchQualityTests(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.store = AgentStore(Path(self.tmp.name) / "agent.sqlite3")
        self.ledger = QualityLedger(self.store, suppliers={"佰特"})
        self.items = WorkItems(self.store)
        self.directory = FakeDirectory({"张三"})

    def tearDown(self):
        self.tmp.cleanup()

    def test_reporter_can_resolve_from_workbench(self):
        issue = self.ledger.record(description="开胶", reporter="张三")
        item = self.items.decide_quality(
            self.ledger, issue_id=issue["id"], decision="resolve",
            operator="张三", resolution="已赔", directory=self.directory,
        )
        self.assertEqual("resolved", item["status"])
        self.assertEqual("resolved", self.ledger.get(issue["id"])["status"])

    def test_other_operator_cannot_decide(self):
        issue = self.ledger.record(description="开胶", reporter="张三")
        other = FakeDirectory({"李四"})
        with self.assertRaisesRegex(ActionError, "自己登记"):
            self.items.decide_quality(
                self.ledger, issue_id=issue["id"], decision="cancel",
                operator="李四", directory=other,
            )
        self.assertEqual("open", self.ledger.get(issue["id"])["status"])

    def test_viewer_and_unbound_are_denied(self):
        issue = self.ledger.record(description="开胶", reporter="张三")
        viewer = FakeDirectory({"张三"}, role="viewer")
        with self.assertRaisesRegex(ActionError, VIEWER_WRITE_DENIED):
            self.items.decide_quality(
                self.ledger, issue_id=issue["id"], decision="resolve",
                operator="张三", directory=viewer,
            )
        with self.assertRaisesRegex(ActionError, WEB_OPERATOR_UNBOUND):
            self.items.decide_quality(
                self.ledger, issue_id=issue["id"], decision="resolve",
                operator="路人", directory=self.directory,
            )


class WorkbookTests(unittest.TestCase):
    def test_headers_and_text_format(self):
        tmp = tempfile.TemporaryDirectory()
        path = Path(tmp.name) / "q.xlsx"
        build_quality_workbook([{
            "id": "ab12cd", "createdAt": "2026-08-13 17:00:00", "reporter": "利特",
            "supplier": "佰特", "poId": "604264", "sku": "XZ1", "severity": "",
            "description": "开胶", "status": "open", "resolution": "",
        }], path)
        book = load_workbook(path)
        sheet = book.active
        self.assertEqual(
            ["编号", "登记时间", "登记人", "供应商", "采购单号", "SKU",
             "严重度", "问题描述", "状态", "处理备注"],
            [sheet.cell(1, col).value for col in range(1, 11)],
        )
        self.assertEqual("@", sheet.cell(2, 1).number_format)
        self.assertEqual("@", sheet.cell(2, 5).number_format)
        self.assertEqual("@", sheet.cell(2, 6).number_format)
        self.assertEqual("未关闭", sheet.cell(2, 9).value)
        tmp.cleanup()


class SchedulerTests(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.store = AgentStore(Path(self.tmp.name) / "agent.sqlite3")
        self.audit = AuditLog(self.store)
        self.ledger = QualityLedger(self.store, suppliers={"佰特"}, lookup_po=lambda _: True)
        self.out = Path(self.tmp.name) / "outputs" / "quality"

    def tearDown(self):
        self.tmp.cleanup()

    def _scheduler(self, sender, **kwargs):
        return DailyQualityReportScheduler(
            ledger=self.ledger, sender=sender, audit=self.audit,
            output_dir=self.out, send_time="17:30", **kwargs,
        )

    def test_empty_day_skips(self):
        result = self._scheduler(FakeSender()).run_once(today="2026-08-13")
        self.assertTrue(result.get("skipped"))
        self.assertEqual("今日无品控登记", result["reason"])

    def test_empty_day_notice(self):
        sender = FakeSender()
        result = self._scheduler(sender, empty_mode="notice").run_once(today="2026-08-13")
        self.assertTrue(result.get("sent"))
        self.assertIn("今日无品控登记", sender.calls[0]["text"])

    def test_idempotent_key(self):
        self.ledger.record(description="开胶")
        today = business_today().isoformat()
        sender = FakeSender()
        first = self._scheduler(sender).run_once(today=today)
        second = self._scheduler(sender).run_once(today=today)
        self.assertTrue(first.get("sent"))
        self.assertTrue(second.get("skipped"))
        self.assertEqual(1, len(sender.calls))

    def test_tick_survives_send_error(self):
        self.ledger.record(description="开胶")
        scheduler = self._scheduler(FakeSender(fail=True), retry_interval_seconds=0)
        now = datetime(2026, 8, 13, 17, 31, tzinfo=ZoneInfo("Asia/Shanghai"))
        # 用业务今天的日期，避免 last_run 与 report_date 错位
        today = business_today().isoformat()
        now = datetime.fromisoformat(today + "T17:31:00").replace(tzinfo=BUSINESS_TIMEZONE)
        result = scheduler.tick(now=now)
        self.assertTrue(result.get("failed"))
        self.assertTrue(scheduler.last_error)
        scheduler.sender = FakeSender()
        again = scheduler.tick(now=now)
        self.assertTrue(again.get("sent") or again.get("skipped") or again.get("failed"))

    def test_webhook_signed_url(self):
        self.ledger.record(description="开胶")
        sender = FakeSender()
        scheduler = self._scheduler(
            sender, link_secret="secret", public_base="http://127.0.0.1:8777",
        )
        today = business_today().isoformat()
        scheduler.run_once(today=today)
        compact = today.replace("-", "")
        sig = report_link_sig("secret", compact)
        self.assertIn(f"/api/quality/reports/{compact}/{sig}.xlsx", sender.calls[0]["text"])
        self.assertTrue(report_link_valid("secret", compact, sig, today=business_today()))
        self.assertFalse(report_link_valid("secret", compact, "deadbeefdeadbeef"))

    def test_app_robot_uploads_file(self):
        self.ledger.record(description="开胶")
        sender = FakeSender(app=True)
        today = business_today().isoformat()
        self._scheduler(sender).run_once(today=today)
        self.assertTrue(any(item.get("mediaId") == "mid-1" for item in sender.files))


class MultipartTests(unittest.TestCase):
    def test_encode_multipart_structure(self):
        body, content_type = encode_multipart(
            [("type", "file")],
            [("media", "a.xlsx", b"PK", "application/octet-stream")],
        )
        self.assertIn("multipart/form-data; boundary=", content_type)
        boundary = content_type.split("boundary=", 1)[1].encode("utf-8")
        self.assertIn(b'name="type"', body)
        self.assertIn(b'name="media"', body)
        self.assertIn(b'filename="a.xlsx"', body)
        self.assertIn(b"PK", body)
        self.assertTrue(body.startswith(b"--" + boundary))
        self.assertTrue(body.endswith(b"--" + boundary + b"--\r\n"))


if __name__ == "__main__":
    unittest.main()
