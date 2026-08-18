# -*- coding: utf-8 -*-
"""代发订单：空白模板列名、文件名，以及 DigitalRuntime 打开列表。"""
from __future__ import annotations

import tempfile
import unittest
from datetime import date
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

from backend.dropship.collect import prepare_dropship_list
from backend.dropship.export import fill_stats, public_export_result, rows_from_orders
from backend.dropship.page import (
    LIST_URL,
    QUESTION_TYPE,
    ensure_epaas_order_page,
    filter_unscheduled_dropship,
)
from backend.dropship.products import apply_sku_facts
from backend.dropship.workbook import (
    SHEET1_COLUMNS,
    SHEET1_HIDDEN,
    SHEET2_COLUMNS,
    dropship_filename,
    dropship_output_path,
    write_blank_dropship_template,
    write_dropship_workbook,
)
from backend.erp import DigitalRuntime, ErpError


class RecordingSender:
    app_ready = True
    group_conversation_id = "cid-group"

    def __init__(self):
        self.markdowns = []
        self.files = []

    def send_markdown(self, title, text, **kwargs):
        self.markdowns.append((title, text))
        return {"channel": "app"}

    def upload_media(self, path, filetype="file"):
        return {"mediaId": "mid-1"}

    def send_file(self, conversation_id, media_id, file_name, file_type="xlsx"):
        self.files.append({"cid": conversation_id, "media": media_id, "name": file_name})
        return {"ok": True}

    def send_oto_markdown(self, title, text, *, user_ids=()):
        self.markdowns.append((title, text, list(user_ids)))
        return {"channel": "oto", "atUserIds": list(user_ids)}

    def send_oto_file(self, user_ids, media_id, file_name, file_type="xlsx"):
        self.files.append({"oto": list(user_ids), "media": media_id, "name": file_name})
        return {"channel": "oto", "atUserIds": list(user_ids)}


class WorkbookTests(unittest.TestCase):
    def test_filename_uses_business_yymmdd(self):
        self.assertEqual("260817-代发.xlsx", dropship_filename(date(2026, 8, 17)))

    def test_output_path_under_files_outputs(self):
        path = dropship_output_path(root=Path("."), today=date(2026, 8, 17))
        self.assertEqual("260817-代发.xlsx", path.name)
        self.assertIn("outputs", path.parts)
        self.assertIn("dropship", path.parts)

    def test_blank_workbook_has_two_sheets_and_no_rows(self):
        tmp = tempfile.TemporaryDirectory()
        path = Path(tmp.name) / "260817-代发.xlsx"
        write_dropship_workbook([], path, blank=True)
        book = load_workbook(path)
        self.assertEqual(["Sheet1", "Sheet2"], book.sheetnames)
        sheet1 = book["Sheet1"]
        sheet2 = book["Sheet2"]
        self.assertEqual(SHEET1_COLUMNS, [sheet1.cell(1, col).value for col in range(1, 29)])
        self.assertEqual(SHEET2_COLUMNS, [sheet2.cell(1, col).value for col in range(1, 18)])
        self.assertIsNone(sheet1.cell(2, 1).value)
        self.assertEqual("A2", sheet1.freeze_panes)
        self.assertIsNone(sheet2.freeze_panes)
        self.assertEqual("A1:AB1", sheet1.auto_filter.ref)
        hidden = {
            sheet1.cell(1, col).value
            for col in range(1, 29)
            if sheet1.column_dimensions[sheet1.cell(1, col).column_letter].hidden
        }
        self.assertEqual(SHEET1_HIDDEN, hidden)
        tmp.cleanup()

    def test_filled_row_maps_to_supplier_sheet(self):
        tmp = tempfile.TemporaryDirectory()
        path = Path(tmp.name) / "row.xlsx"
        write_dropship_workbook([{
            "内部订单号": "11554549",
            "线上订单号": "SO-1",
            "收货人": "测",
            "手机": "13800138000",
            "数量": 2,
            "商品裸价": 10,
            "成本价": 1.5,
            "快递单号": "SF123",
        }], path)
        book = load_workbook(path)
        sheet1 = book["Sheet1"]
        sheet2 = book["Sheet2"]
        self.assertEqual("11554549", sheet1.cell(2, 1).value)
        self.assertEqual("@", sheet1.cell(2, 1).number_format)
        self.assertEqual("测", sheet1.cell(2, 11).value)
        self.assertEqual("13800138000", sheet1.cell(2, 16).value)
        self.assertEqual("@", sheet1.cell(2, 16).number_format)
        self.assertEqual("SF123", sheet1.cell(2, 25).value)
        self.assertEqual(2, sheet1.cell(2, 26).value)
        self.assertEqual(10, sheet1.cell(2, 27).value)
        self.assertEqual(1.5, sheet1.cell(2, 28).value)
        self.assertEqual("11554549", sheet2.cell(2, 1).value)
        self.assertEqual("测", sheet2.cell(2, 2).value)
        self.assertEqual("SF123", sheet2.cell(2, 16).value)
        self.assertEqual(2, sheet2.cell(2, 17).value)
        tmp.cleanup()

    def test_blank_template_writes_master_and_dated_copy(self):
        tmp = tempfile.TemporaryDirectory()
        root = Path(tmp.name)
        (root / "files" / "templates").mkdir(parents=True)
        (root / "files" / "outputs").mkdir(parents=True)
        result = write_blank_dropship_template(root=root, today=date(2026, 8, 17), include_dated=True)
        self.assertTrue(Path(result["template"]).exists())
        self.assertEqual("代发订单模板.xlsx", Path(result["template"]).name)
        self.assertEqual("260817-代发.xlsx", result["filename"])
        self.assertTrue(Path(result["dated"]).exists())
        tmp.cleanup()

    def test_blank_template_default_skips_dated_file(self):
        tmp = tempfile.TemporaryDirectory()
        root = Path(tmp.name)
        (root / "files" / "templates").mkdir(parents=True)
        (root / "files" / "outputs").mkdir(parents=True)
        result = write_blank_dropship_template(root=root, today=date(2026, 8, 17))
        self.assertTrue(Path(result["template"]).exists())
        self.assertNotIn("dated", result)
        self.assertFalse((root / "files" / "outputs" / "dropship" / "260817-代发.xlsx").exists())
        tmp.cleanup()


class ExportMapTests(unittest.TestCase):
    def test_rows_one_sku_one_line(self):
        rows = rows_from_orders([{
            "o_id": "1",
            "so_id": "SO",
            "shop_name": "店",
            "shop_site": "线下店铺",
            "l_id": "SF999",
            "receiver_name": "张*",
            "receiver_mobile": "139****0000",
            "receiver_state": "浙江",
            "receiver_city": "杭州",
            "receiver_district": "西湖",
            "receiver_address": "文一路",
            "items": [
                {"sku": "A", "name": "鞋", "qty": 2, "props": "黑", "price": 10, "cost": 8, "supplier": "&厂"},
                {"sku": "B", "name": "垫", "qty": 1, "price": 3, "cost": 2, "supplier": "&厂"},
            ],
        }], {
            "1": {"name": "张三", "mobile": "13900001111", "street": "文一路1号"},
        })
        self.assertEqual(2, len(rows))
        self.assertEqual("张三", rows[0]["收货人"])
        self.assertEqual("13900001111", rows[0]["手机"])
        self.assertEqual("浙江杭州西湖文一路1号", rows[0]["地址(包含省市区)"])
        self.assertEqual("A", rows[0]["商品编码"])
        self.assertEqual(2, rows[0]["数量"])
        self.assertEqual(10, rows[0]["商品裸价"])
        self.assertEqual(8, rows[0]["成本价"])
        self.assertEqual("&厂", rows[0]["供应商"])
        self.assertEqual("SF999", rows[0]["快递单号"])
        self.assertEqual(3, rows[1]["商品裸价"])
        self.assertEqual(SHEET1_COLUMNS, list(rows[0].keys()))
        stats = fill_stats(rows)
        self.assertEqual(2, stats["手机"])
        self.assertEqual(1, stats["orders"])
        self.assertEqual(2, stats["商品裸价"])
        self.assertEqual(2, stats["成本价"])

    def test_sku_facts_fill_empty_supplier_and_cost(self):
        orders = [{
            "o_id": "1",
            "items": [{"sku": "A", "price": 10, "qty": 1}],
        }]
        apply_sku_facts(orders, {"A": {"supplier": "&厂", "cost": 7.5, "name": "鞋"}})
        self.assertEqual("&厂", orders[0]["items"][0]["supplier"])
        self.assertEqual(7.5, orders[0]["items"][0]["cost"])
        self.assertEqual("鞋", orders[0]["items"][0]["name"])
        apply_sku_facts(orders, {"A": {"supplier": "覆盖", "cost": 1}})
        self.assertEqual("&厂", orders[0]["items"][0]["supplier"])
        self.assertEqual(7.5, orders[0]["items"][0]["cost"])

    def test_sku_facts_fill_from_style_when_sku_missing(self):
        orders = [{"o_id": "1", "items": [{"sku": "A-1", "style": "A", "qty": 1}]}]
        apply_sku_facts(orders, {"A": {"supplier": "&厂", "cost": 4, "supplier_style": "款"}})
        self.assertEqual("&厂", orders[0]["items"][0]["supplier"])
        self.assertEqual(4, orders[0]["items"][0]["cost"])
        self.assertEqual("款", orders[0]["items"][0]["supplier_style"])

    def test_public_export_result_drops_order_rows(self):
        result = public_export_result({
            "filename": "260817-代发.xlsx",
            "path": "out.xlsx",
            "dataCount": 1,
            "stats": {"orders": 1, "lines": 1, "供应商": 1},
            "orders": [{"receiver_name": "隐"}],
        })
        self.assertEqual("260817-代发.xlsx", result["filename"])
        self.assertEqual(1, result["orders"])
        self.assertIsInstance(result["orders"], int)


class FakeFrame:
    def __init__(self, url, values):
        self.url = url
        self.values = values
        self.waits = []

    def evaluate(self, expr, arg=None):
        if "FullSearch" in str(expr):
            return self.values.get("filter", {"ok": True, "questionChecked": True})
        return self.values.get("list", {})

    def wait_for_function(self, expr, timeout=0):
        self.waits.append(expr)

    def wait_for_timeout(self, ms):
        return None


class FakePortalPage:
    def __init__(self, *, url="https://www.erp321.com/epaas", frames=None, evaluate_map=None):
        self.url = url
        self.frames = frames or []
        self.evaluate_map = evaluate_map or {}
        self.gotos = []
        self.timeouts = []

    def goto(self, url, **kwargs):
        self.gotos.append(url)
        self.url = url

    def evaluate(self, expr, arg=None):
        text = str(expr)
        if "iframe" in text:
            return self.evaluate_map.get("open", {"ok": True, "src": arg, "count": 1})
        if "GetTopDataByBX" in text:
            return self.evaluate_map.get("top", {"GetTopDataByBX": "function", "href": self.url})
        return None

    def wait_for_timeout(self, ms):
        self.timeouts.append(ms)


class PageTests(unittest.TestCase):
    def test_ensure_embeds_list_in_epaas(self):
        frame = FakeFrame(LIST_URL, {})
        page = FakePortalPage(frames=[frame])
        ready = ensure_epaas_order_page(page)
        self.assertTrue(ready["ok"])
        self.assertTrue(ready["hasGetTop"])
        self.assertEqual(LIST_URL, ready["opened"]["src"])
        self.assertEqual(0, len(page.gotos))

    def test_login_redirect_is_hard_error(self):
        page = FakePortalPage(url="https://www.erp321.com/login.aspx")
        with self.assertRaisesRegex(ErpError, "登录页"):
            ensure_epaas_order_page(page)

    def test_filter_only_uses_question_type(self):
        frame = FakeFrame(LIST_URL, {
            "filter": {"ok": True, "questionChecked": True},
            "list": {
                "dataCount": 32, "dropshipCount": 32, "rowCount": 32, "pageSize": 200,
                "shopSites": {"淘宝天猫": 18}, "oIds": ["1"],
                "href": LIST_URL,
            },
        })
        result = filter_unscheduled_dropship(frame)
        self.assertEqual(QUESTION_TYPE, "代发订单未安排")
        self.assertEqual(32, result["list"]["dropshipCount"])
        self.assertEqual(["1"], result["list"]["oIds"])


class CollectTests(unittest.TestCase):
    def test_prepare_reuses_runtime_browser_lock(self):
        calls = []

        class Runtime:
            def run_browser(self, func, *args, **kwargs):
                calls.append(func.__name__)
                return {"ok": True, "dropshipCount": 2, "oIds": ["a", "b"]}

        result = prepare_dropship_list(Runtime())
        self.assertEqual(["_prepare_on_page"], calls)
        self.assertEqual(2, result["dropshipCount"])

    def test_missing_runtime_explains_login(self):
        with self.assertRaisesRegex(ErpError, "login"):
            prepare_dropship_list(None)

    def test_run_browser_logs_in_and_saves_cookie(self):
        class Session:
            def __init__(self):
                self.page = FakePortalPage(frames=[FakeFrame(LIST_URL, {
                    "filter": {"ok": True, "questionChecked": True},
                    "list": {"dataCount": 1, "dropshipCount": 1, "rowCount": 1, "pageSize": 200, "oIds": ["9"], "shopSites": {}},
                })])
                self.saved = 0
                self.logged = 0

            def login_if_needed(self, **kwargs):
                self.logged += 1
                return {"ok": True}

            def save_state(self):
                self.saved += 1

            def close(self):
                return None

        session = Session()
        runtime = DigitalRuntime({
            "enabled": True, "workerId": "erp-ai-procurement",
            "baseUrl": "https://www.erp321.com/epaas",
            "orderListUrl": "https://www.erp321.com/app/order/order/list.aspx",
            "headless": True, "writeDelayMs": 250,
            "storageStatePath": "unused.json",
        }, session=session)
        with patch("backend.dropship.collect.ensure_epaas_order_page", return_value={"ok": True, "hasGetTop": True, "href": LIST_URL}):
            result = prepare_dropship_list(runtime)
        self.assertEqual(1, session.logged)
        self.assertEqual(1, session.saved)
        self.assertEqual(["9"], result["oIds"])
        runtime.close()


class SchedulerTests(unittest.TestCase):
    def test_skips_filled_workbook_and_missing_runtime(self):
        from datetime import datetime
        from backend.dropship.scheduler import DailyDropshipScheduler, dropship_file_has_rows
        from backend.dropship.workbook import dropship_output_path, write_dropship_workbook

        tmp = tempfile.TemporaryDirectory()
        root = Path(tmp.name)
        today = date(2026, 8, 18)
        path = dropship_output_path(root=root, today=today)
        write_dropship_workbook([{
            "内部订单号": "11554549",
            "线上订单号": "SO-1",
            "收货人": "测",
            "数量": 1,
        }], path)
        self.assertTrue(dropship_file_has_rows(path))
        scheduler = DailyDropshipScheduler(
            runtime=None, root=root, send_time="14:00", enabled=False,
        )
        skipped = scheduler.run_once(today=today)
        self.assertTrue(skipped["skipped"])
        self.assertIn("不覆盖", skipped["reason"])
        empty_root = Path(tmp.name) / "empty"
        empty_root.mkdir()
        missing = DailyDropshipScheduler(
            runtime=None, root=empty_root, send_time="14:00", enabled=False,
        )
        failed = missing.run_once(today=today)
        self.assertTrue(failed["failed"])
        early = scheduler.tick(now=datetime(2026, 8, 18, 10, 0))
        self.assertEqual("未到准备时间", early["reason"])
        too_early = scheduler.tick(now=datetime(2026, 8, 18, 13, 40))
        self.assertEqual("未到准备时间", too_early["reason"])
        waiting = scheduler.tick(now=datetime(2026, 8, 18, 13, 50))
        self.assertEqual("表已备好，等待发送", waiting["reason"])
        tmp.cleanup()

    def test_prepares_workbook_before_send_time(self):
        from datetime import datetime
        from backend.dropship.scheduler import DailyDropshipScheduler

        tmp = tempfile.TemporaryDirectory()
        root = Path(tmp.name)
        called = []

        class Runtime:
            pass

        def fake_export(runtime, *, path=None, root=None, env_path=None):
            called.append(str(path))
            Path(path).parent.mkdir(parents=True, exist_ok=True)
            from backend.dropship.workbook import write_dropship_workbook
            write_dropship_workbook([{
                "内部订单号": "1", "线上订单号": "SO-1", "收货人": "测", "数量": 1,
            }], Path(path))
            return {"ok": True, "filename": Path(path).name, "path": str(path), "stats": {}}

        sender = RecordingSender()
        scheduler = DailyDropshipScheduler(
            runtime=Runtime(), root=root, send_time="14:00",
            prepare_lead_minutes=30, enabled=False, sender=sender,
            conversation_id="cid-drop",
        )
        with patch("backend.dropship.export.export_today_dropship", side_effect=fake_export):
            prepared = scheduler.tick(now=datetime(2026, 8, 18, 13, 40))
        self.assertTrue(prepared.get("prepared"))
        self.assertFalse(prepared.get("sent"))
        self.assertEqual(1, len(called))
        self.assertEqual([], sender.files)
        waiting = scheduler.tick(now=datetime(2026, 8, 18, 13, 50))
        self.assertEqual("表已备好，等待发送", waiting["reason"])
        self.assertEqual(1, len(called))
        tmp.cleanup()

    def test_notice_uses_written_cutoff(self):
        from backend.dropship.scheduler import dropship_notice_text
        from backend.dropship.workbook import (
            dropship_output_path, read_dropship_cutoff, write_dropship_cutoff,
            write_dropship_workbook,
        )

        tmp = tempfile.TemporaryDirectory()
        root = Path(tmp.name)
        path = dropship_output_path(root=root, today=date(2026, 8, 18))
        write_dropship_workbook([{"内部订单号": "1", "数量": 1}], path)
        write_dropship_cutoff(path, cutoff="2026-08-18 13:55")
        self.assertEqual("2026-08-18 13:55", read_dropship_cutoff(path))
        text = dropship_notice_text(
            today="2026-08-18", rows=1, filename=path.name, cutoff="2026-08-18 13:55",
        )
        self.assertEqual(
            "今日代发未安排 1 行，数据截至 2026-08-18 13:55。文件：260818-代发.xlsx",
            text,
        )
        tmp.cleanup()

    def test_notice_mentions_rate_limited_orders(self):
        from backend.dropship.export import receiver_rest_complete
        from backend.dropship.scheduler import dropship_notice_text

        rows = [
            {
                "内部订单号": "11559389", "收货人": "张*", "手机": "138****0000",
                "地址(包含省市区)": "杭州市**",
            },
            {
                "内部订单号": "11555749", "收货人": "*", "手机": "****",
                "地址(包含省市区)": "**",
            },
            {
                "内部订单号": "11554001", "收货人": "李四", "手机": "13800138000",
                "地址(包含省市区)": "杭州市西湖区文三路1号",
            },
        ]
        limited = ["11559389", "11555749"]
        self.assertTrue(receiver_rest_complete(rows, limited))
        self.assertFalse(receiver_rest_complete(rows, []))
        text = dropship_notice_text(
            today="2026-08-18", rows=3, filename="260818-代发.xlsx",
            cutoff="2026-08-18 14:22", rate_limited=limited, rest_complete=True,
        )
        self.assertIn("有 2 单揭收货被限流（11559389、11555749）", text)
        self.assertIn("其余收货人/手机/地址都齐了", text)
        incomplete = dropship_notice_text(
            today="2026-08-18", rows=3, filename="260818-代发.xlsx",
            cutoff="2026-08-18 14:22", rate_limited=limited, rest_complete=False,
        )
        self.assertIn("有 2 单揭收货被限流（11559389、11555749）", incomplete)
        self.assertNotIn("都齐了", incomplete)

    def test_run_once_exports_with_fake_runtime(self):
        from backend.dropship.scheduler import DailyDropshipScheduler

        tmp = tempfile.TemporaryDirectory()
        root = Path(tmp.name)
        today = date(2026, 8, 18)
        called = []

        class Runtime:
            pass

        def fake_export(runtime, *, path=None, root=None, env_path=None):
            called.append(str(path))
            Path(path).parent.mkdir(parents=True, exist_ok=True)
            Path(path).write_bytes(b"xlsx")
            return {"ok": True, "filename": Path(path).name, "path": str(path), "stats": {}}

        scheduler = DailyDropshipScheduler(
            runtime=Runtime(), root=root, send_time="14:00", enabled=False,
        )
        with patch("backend.dropship.export.export_today_dropship", side_effect=fake_export):
            result = scheduler.run_once(today=today, operator="tester")
        self.assertTrue(result["ok"])
        self.assertEqual(1, len(called))
        self.assertEqual("tester", result["operator"])
        self.assertFalse(result.get("sent"))
        tmp.cleanup()

    def test_sends_existing_workbook_to_dingtalk_without_export(self):
        from backend.agent.store import AgentStore
        from backend.agent.audit import AuditLog
        from backend.dropship.scheduler import DailyDropshipScheduler
        from backend.dropship.workbook import (
            dropship_output_path, write_dropship_cutoff, write_dropship_workbook,
        )

        tmp = tempfile.TemporaryDirectory()
        root = Path(tmp.name)
        today = date(2026, 8, 18)
        path = dropship_output_path(root=root, today=today)
        write_dropship_workbook([{
            "内部订单号": "11554549",
            "线上订单号": "SO-1",
            "收货人": "测",
            "数量": 1,
        }], path)
        write_dropship_cutoff(
            path, cutoff="2026-08-18 14:22",
            rate_limited=["11559389", "11555749"], rest_complete=True,
        )
        sender = RecordingSender()
        audit = AuditLog(AgentStore(root / "agent.sqlite3"))
        scheduler = DailyDropshipScheduler(
            runtime=None, root=root, send_time="14:00", enabled=False,
            sender=sender, audit=audit, conversation_id="cid-drop",
        )
        exported = []

        def boom(*args, **kwargs):
            exported.append(1)
            raise AssertionError("已填表不应再抓取")

        with patch("backend.dropship.export.export_today_dropship", side_effect=boom):
            result = scheduler.run_once(today=today)
        self.assertTrue(result.get("sent"))
        self.assertTrue(result.get("reused"))
        self.assertEqual(1, result.get("rows"))
        self.assertEqual("cid-drop", sender.files[0]["cid"])
        self.assertIn("代发未安排", sender.markdowns[0][0])
        self.assertIn("1 行", sender.markdowns[0][1])
        self.assertIn("数据截至", sender.markdowns[0][1])
        self.assertIn("有 2 单揭收货被限流（11559389、11555749）", sender.markdowns[0][1])
        self.assertIn("其余收货人/手机/地址都齐了", sender.markdowns[0][1])
        self.assertNotIn("测", sender.markdowns[0][1])
        self.assertEqual([], exported)
        again = scheduler.run_once(today=today)
        self.assertTrue(again.get("skipped"))
        self.assertIn("已经发过", again["reason"])
        self.assertEqual(1, len(sender.files))
        tmp.cleanup()

    def test_sends_oto_copy_to_bound_buyer(self):
        from backend.agent.store import AgentStore
        from backend.agent.audit import AuditLog
        from backend.dingtalk.identity import StaffDirectory
        from backend.dropship.scheduler import DailyDropshipScheduler
        from backend.dropship.workbook import dropship_output_path, write_dropship_workbook

        tmp = tempfile.TemporaryDirectory()
        root = Path(tmp.name)
        today = date(2026, 8, 18)
        path = dropship_output_path(root=root, today=today)
        write_dropship_workbook([{
            "内部订单号": "11554549",
            "线上订单号": "SO-1",
            "收货人": "测",
            "数量": 1,
        }], path)
        directory = StaffDirectory(AgentStore(root / "agent.sqlite3"))
        directory.upsert("姚安安（安安）", dingtalk_user_id="u-anan")
        sender = RecordingSender()
        scheduler = DailyDropshipScheduler(
            runtime=None, root=root, send_time="14:00", enabled=False,
            sender=sender, audit=AuditLog(AgentStore(root / "audit.sqlite3")),
            conversation_id="cid-drop", directory=directory, oto_buyers="安安",
        )
        result = scheduler.run_once(today=today)
        self.assertTrue(result.get("sent"))
        self.assertEqual(1, result.get("otoUserCount"))
        self.assertEqual(["u-anan"], sender.files[-1]["oto"])
        self.assertEqual(2, len(sender.files))
        missing = DailyDropshipScheduler(
            runtime=None, root=root, send_time="14:00", enabled=False,
            sender=RecordingSender(), audit=AuditLog(AgentStore(root / "audit2.sqlite3")),
            conversation_id="cid-drop", directory=StaffDirectory(AgentStore(root / "empty.sqlite3")),
            oto_buyers="安安",
        )
        failed = missing.run_once(today=today)
        self.assertTrue(failed.get("failed"))
        self.assertIn("安安", failed["reason"])
        tmp.cleanup()
