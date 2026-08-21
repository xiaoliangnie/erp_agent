# -*- coding: utf-8 -*-
import unittest
from unittest.mock import patch

import tempfile
from pathlib import Path

from openpyxl import load_workbook

from backend.spu_plan import (
    CATEGORY_LINES,
    SKIP_WARNING_TAGS,
    SNAPSHOT_TABLE,
    TOTAL_COLUMNS,
    TURNOVER_ALERT_DAYS,
    daily_avg,
    daily_avg_baihuo,
    format_alert_text,
    obsolete_label,
    parse_product_attrs,
    product_in_baihuo_scope,
    product_in_scope,
    replenish_qty,
    save_style_snapshot,
    sku_code_status,
    skip_stockout_warning,
    style_stock,
    style_warning,
    style_year,
    write_style_workbook,
)
from backend.spu_plan.channel import (
    is_offline_shop, offline_department, shop_id_from_raw_so, so_id_from_raw_so,
)
from backend.spu_plan.service import SNAPSHOT_COLUMNS, _sale_channel, build_style_alerts
from backend.spu_plan.shops import load_shop_groups, shop_groups_from_records


class FormulaTests(unittest.TestCase):
    def test_daily_avg_four_windows(self):
        self.assertEqual(10.0, daily_avg(10, 30, 70, 150))
        self.assertEqual(0.0, daily_avg(0, 0, 0, 0))
        from backend.spu_plan.service import _inventory_in_qty
        self.assertEqual(300.0, _inventory_in_qty({"source_payload": '{"in_qty": 300}'}))
        self.assertEqual(0.0, _inventory_in_qty({"source_payload": "{}"}))
        self.assertEqual(1.0, daily_avg_baihuo(7, 15, 30))
        self.assertEqual(0.0, daily_avg_baihuo(0, 0, 0))
        self.assertEqual(0.3, daily_avg_baihuo(0, 0, 30))
        from backend.spu_plan import monthly_sales_baihuo
        self.assertEqual(10.0, monthly_sales_baihuo(0, 0, 30))
        self.assertEqual(29.3, monthly_sales_baihuo(7, 15, 30))
        self.assertEqual(TURNOVER_ALERT_DAYS, 30)
        self.assertEqual("21", style_year("BH21401306-027"))
        self.assertEqual(600, replenish_qty(10, 0))
        self.assertEqual(0, replenish_qty(10, 600))
        self.assertEqual(-5, replenish_qty(1, 65))
        self.assertEqual(300, replenish_qty(10, 0, cover_days=30))
        self.assertEqual(-70, replenish_qty(1, 100, cover_days=30))

    def test_order_qty_rounds_up(self):
        from backend.spu_plan import order_qty
        self.assertIsNone(order_qty(None))
        self.assertEqual(0, order_qty(0))
        self.assertEqual(0, order_qty(-5))
        self.assertEqual(10, order_qty(1))
        self.assertEqual(40, order_qty(40))
        self.assertEqual(50, order_qty(41))
        # 有起订量按起订量的倍数
        self.assertEqual(150, order_qty(101, moq=150))
        self.assertEqual(300, order_qty(151, moq=150))

    def test_obsolete_label_from_tags(self):
        self.assertEqual("", obsolete_label([]))
        self.assertEqual("", obsolete_label(["热销", "自营鞋品"]))
        self.assertEqual("淘汰", obsolete_label(["店加", "淘汰品"]))
        self.assertEqual("有升级", obsolete_label(["有升级"]))
        self.assertEqual("清仓", obsolete_label(["清仓品"]))
        self.assertEqual("淘汰、有升级", obsolete_label(["有升级", "淘汰品"]))

    def test_sku_code_status(self):
        self.assertEqual("断码", sku_code_status(0, 10))
        self.assertEqual("断码", sku_code_status(0.4, 0))
        self.assertEqual("缺码", sku_code_status(3, 7))
        self.assertEqual("充足", sku_code_status(7, 7))
        self.assertEqual("充足", sku_code_status(8, 7))

    def test_style_stock_and_warning(self):
        self.assertEqual(8.0, style_stock(10, 4, 2))
        hit = style_warning(
            qty=10, occupy=4, inbound=2,
            sales_1=10, sales_3=30, sales_7=70, sales_15=150,
            skip_warning=False,
        )
        self.assertEqual(10.0, hit["dailyAvg"])
        self.assertEqual(8.0, hit["onHand"])
        self.assertTrue(hit["stockout"])
        tagged = style_warning(
            qty=10, occupy=4, inbound=2,
            sales_1=10, sales_3=30, sales_7=70, sales_15=150,
            skip_warning=True,
        )
        self.assertFalse(tagged["stockout"])
        self.assertTrue(tagged["skipWarning"])
        idle = style_warning(
            qty=0, occupy=0, inbound=0,
            sales_1=0, sales_3=0, sales_7=0, sales_15=0,
            skip_warning=False,
        )
        self.assertIsNone(idle["turnoverDays"])
        self.assertFalse(idle["stockout"])


class RosterTests(unittest.TestCase):
    def test_category_lines_are_erp_text(self):
        self.assertEqual(("鞋类", "通勤裤", "服装-非通勤裤"), CATEGORY_LINES)
        self.assertIn("清仓品", SKIP_WARNING_TAGS)

    def test_parse_moq_from_remark(self):
        from backend.spu_plan.roster import parse_moq
        self.assertEqual(0, parse_moq(None))
        self.assertEqual(0, parse_moq("采购价不含税"))
        self.assertEqual(2000, parse_moq("起订2000"))
        self.assertEqual(300, parse_moq("起订量：300"))
        self.assertEqual(500, parse_moq("合布起订500（可组合）"))

    def test_parse_and_scope(self):
        attrs = parse_product_attrs({
            "sku_id": "SKU-01",
            "i_id": "BH25001",
            "name": "作训鞋",
            "labels": "热销,清仓品",
            "enabled": 1,
            "source_payload": {
                "other_3": "鞋类",
                "other_10": "一件代发",
                "other_2": "四季",
            },
        })
        self.assertEqual("BH25001", attrs["styleId"])
        self.assertEqual("鞋类", attrs["categoryLine"])
        self.assertEqual("一件代发", attrs["productionMode"])
        self.assertTrue(product_in_scope(attrs))
        self.assertTrue(skip_stockout_warning(attrs["labels"]))

    def test_out_of_scope_and_tag_match(self):
        pants = parse_product_attrs({
            "sku_id": "A", "i_id": "S1", "enabled": 1,
            "source_payload": {"other_3": "服饰-非通勤"},
        })
        self.assertFalse(product_in_scope(pants))
        apparel = parse_product_attrs({
            "sku_id": "B", "i_id": "S2", "enabled": 1,
            "source_payload": {"other_3": "服装-非通勤裤", "other_10": "外供"},
        })
        self.assertTrue(product_in_scope(apparel))
        self.assertFalse(skip_stockout_warning("新品,活动"))
        self.assertTrue(skip_stockout_warning("有升级"))
        self.assertTrue(skip_stockout_warning(["淘汰品"]))
        self.assertFalse(product_in_scope({"styleId": "", "categoryLine": "鞋类"}))

    def test_baihuo_scope_is_label_not_category_line(self):
        goods = parse_product_attrs({
            "sku_id": "CUP-01", "i_id": "WC-01", "enabled": 1,
            "labels": "店加,自营百货",
            "source_payload": {"other_3": "文创百货", "other_10": ""},
        })
        self.assertTrue(product_in_baihuo_scope(goods))
        self.assertFalse(product_in_scope(goods))
        shoe = parse_product_attrs({
            "sku_id": "SKU-01", "i_id": "XZ-01", "enabled": 1,
            "labels": "自营鞋品",
            "source_payload": {"other_3": "鞋类", "other_10": "自营"},
        })
        self.assertTrue(product_in_scope(shoe))
        self.assertFalse(product_in_baihuo_scope(shoe))
        self.assertFalse(product_in_baihuo_scope({
            "skuId": "", "styleId": "WC-01", "labels": ["自营百货"],
        }))


class ChannelTests(unittest.TestCase):
    def test_offline_shops_are_four_departments(self):
        self.assertEqual("消防", offline_department("消防-吴欣宇-蜀黍家消防业务经理"))
        self.assertEqual("消防", offline_department("消防-小冲-蜀黍家消防渠道负责人"))
        self.assertEqual("公安", offline_department("公安-康康-蜀黍家大客户经理（广东）"))
        self.assertEqual("交警", offline_department("交警-周周-蜀黍家交警业务经理"))
        self.assertEqual("交警", offline_department("张雨-蜀黍家交警业务经理"))
        self.assertEqual("渠道", offline_department("KA渠道-店家订单"))
        self.assertTrue(is_offline_shop("张雨-蜀黍家交警业务经理"))
        self.assertFalse(is_offline_shop("抖音-蜀黍家潮玩文创专卖店"))
        self.assertFalse(is_offline_shop("KA门店-代销商业-上海公安部三所"))
        self.assertFalse(is_offline_shop("私域增长-蜀黍家店长-公安礼品"))
        self.assertFalse(is_offline_shop("张雨-蜀黍家B端业务经理"))
        self.assertFalse(is_offline_shop("{线下}"))
        self.assertFalse(is_offline_shop(""))
        self.assertEqual("内部", offline_department("{内部店铺}"))
        self.assertEqual("内部", offline_department("A内部店铺", "内部店铺"))
        self.assertEqual("渠道", offline_department("KA门店-货架-上海公安部三所", "渠道业务部"))
        self.assertEqual("公安", offline_department("黄维杰-蜀黍家大客户经理", "公安业务部"))
        self.assertEqual("渠道", offline_department("扶光-蜀黍家消防客户经理", "渠道业务部"))
        self.assertEqual("", offline_department("私域增长-蜀黍家店长-公安礼品", "私域运营部"))
        self.assertEqual("", offline_department("抖音-蜀黍家潮玩文创专卖店", "蜀黍家直播部"))
        self.assertEqual("10368931", shop_id_from_raw_so("10368931:E20260820220036033506183"))
        self.assertEqual("E20260820220036033506183", so_id_from_raw_so("10368931:E20260820220036033506183"))
        self.assertEqual("", shop_id_from_raw_so("not-a-shop"))

    def test_shop_setting_group_beats_shop_name(self):
        groups = shop_groups_from_records([
            {"shop_id": 18842068, "shop_name": "KA门店-货架-上海公安部三所", "group_name": "渠道业务部"},
            {"shop_id": 17441473, "shop_name": "私域增长-蜀黍家店长-公安礼品", "group_name": "私域运营部"},
        ])
        self.assertEqual("渠道业务部", groups.group_name("18842068"))
        self.assertEqual(
            "offline",
            _sale_channel({"shop_id": 18842068}, {}, groups),
        )
        self.assertEqual(
            "offline",
            _sale_channel({"raw_so_id": "18842068:E20260821001"}, {}, groups),
        )
        self.assertEqual(
            "online",
            _sale_channel({"shop_id": 17441473}, {}, groups),
        )

    def test_shop_group_cache_without_network(self):
        with tempfile.TemporaryDirectory() as tmp:
            cache = Path(tmp) / "shop-groups.json"
            cache.write_text(
                '{"fetched_at":"2026-08-21T11:00:00+08:00","shops":['
                '{"shop_id":1,"shop_name":"KA-货架","group_name":"渠道业务部"}'
                "]}",
                encoding="utf-8",
            )
            groups = load_shop_groups(cache=cache, fetch=False)
            self.assertEqual("渠道业务部", groups.group_name("1"))


class AlertTextTests(unittest.TestCase):
    def test_format_empty_and_hits(self):
        empty = format_alert_text({
            "today": "2026-08-19", "styleCount": 2,
            "stockoutCount": 0, "brokenStyleCount": 0, "shortStyleCount": 0,
            "styles": [{"styleId": "A", "name": "甲", "stockout": False,
                        "brokenSkus": 0, "shortSkus": 0, "turnoverDays": None}],
        })
        self.assertIn("当前没有缺货", empty)
        text = format_alert_text({
            "today": "2026-08-19", "styleCount": 1,
            "stockoutCount": 1, "brokenStyleCount": 1, "shortStyleCount": 0,
            "styles": [{
                "styleId": "BH25001", "name": "作训鞋", "stockout": True,
                "brokenSkus": 2, "shortSkus": 0, "turnoverDays": 12.5,
                "sales60": 240,
            }],
        })
        self.assertIn("缺货 周转12.5", text)
        self.assertIn("断码2", text)
        self.assertIn("60天", text)
        baihuo_hit = format_alert_text({
            "today": "2026-08-21", "board": "baihuo", "styleCount": 1,
            "stockoutCount": 1, "brokenStyleCount": 0, "shortStyleCount": 0,
            "styles": [{
                "styleId": "CUP01", "name": "杯", "stockout": True,
                "brokenSkus": 0, "shortSkus": 0, "turnoverDays": 8,
                "sales30": 45, "inQty": 300,
            }],
        })
        self.assertIn("30天45", baihuo_hit)
        self.assertIn("进货仓300", baihuo_hit)
        self.assertNotIn("60天", baihuo_hit)
        self.assertNotIn("断码", baihuo_hit)
        self.assertNotIn("缺码", baihuo_hit)
        baihuo = format_alert_text({
            "today": "2026-08-21", "board": "baihuo", "styleCount": 0,
            "stockoutCount": 0, "brokenStyleCount": 0, "shortStyleCount": 0,
            "styles": [],
        })
        self.assertIn("自营百货", baihuo)
        self.assertIn("当前没有缺货或待补货", baihuo)
        self.assertNotIn("断码", baihuo)


class BuildAlertsTests(unittest.TestCase):
    def test_missing_inventory_row_counts_as_zero(self):
        """接口不给从未入库的 SKU 行；示例表按 0 记断码，这里同口径。"""
        products = [{
            "skuId": "SKU-01", "styleId": "S1", "name": "鞋",
            "categoryLine": "鞋类", "productionMode": "自营", "labels": [],
        }]
        with patch("backend.spu_plan.service.load_products", return_value=products), \
                patch("backend.spu_plan.service._load_inventory", return_value={}), \
                patch(
                    "backend.spu_plan.service._load_sales_windows",
                    return_value={"SKU-01": {"1": 10, "3": 30, "7": 70, "15": 150}},
                ), \
                patch("backend.spu_plan.service.load_in_transit", return_value={}):
            result = build_style_alerts("unused.env", today="2026-08-19")
        style = result["styles"][0]
        self.assertEqual(1, style["missingInventory"])
        self.assertEqual(1, style["brokenSkus"])
        self.assertEqual(10.0, style["dailyAvg"])
        self.assertEqual(0.0, style["onHand"])
        self.assertTrue(style["stockout"])
        self.assertEqual("缺货", style["stockoutLabel"])
        self.assertEqual(600, style["replenishQty"])
        self.assertIn("无库存记录", style["remark"])

    def test_apparel_keeps_channel_split_and_sixty_day_cover(self):
        products = [{
            "skuId": "SKU-01", "styleId": "S1", "name": "鞋",
            "categoryLine": "鞋类", "productionMode": "自营", "labels": [],
        }]
        windows = {
            "1": 10, "3": 30, "7": 70, "15": 150, "30": 200,
            "online": {"7": 40, "15": 90, "30": 120},
            "offline": {"7": 30, "15": 60, "30": 80},
            "shops": {
                "抖音店": {
                    "qty7": 40, "qty15": 90, "qty30": 120, "channel": "online",
                    "shopName": "抖音店",
                },
            },
        }
        with patch("backend.spu_plan.service.load_products", return_value=products), \
                patch("backend.spu_plan.service._load_inventory", return_value={
                    "SKU-01": {"qty": 20, "occupy": 0, "inbound": 0, "inQty": 0},
                }), \
                patch(
                    "backend.spu_plan.service._load_sales_windows",
                    return_value={"SKU-01": windows},
                ), \
                patch("backend.spu_plan.service.load_in_transit", return_value={}):
            result = build_style_alerts("unused.env", today="2026-08-21")
        style = result["styles"][0]
        self.assertEqual(40.0, style["sales7Online"])
        self.assertEqual(80.0, style["sales30Offline"])
        self.assertEqual("抖音店", style["saleShops"][0]["shopName"])
        self.assertEqual(580, style["replenishQty"])

    def test_baihuo_daily_avg_uses_three_windows(self):
        """百货日均 = (7×4 + 15×2 + 30) / 3 / 30。"""
        products = [{
            "skuId": "SKU-01", "styleId": "S1", "name": "杯",
            "categoryLine": "文创百货", "productionMode": "", "labels": ["自营百货"],
        }]
        windows = {
            "1": 0, "3": 0, "7": 0, "14": 0, "15": 0, "30": 30,
            "prev7": 0, "days": [0.0] * 30,
        }
        with patch("backend.spu_plan.service.load_products", return_value=products), \
                patch(
                    "backend.spu_plan.service._load_inventory",
                    return_value={"SKU-01": {"qty": 100, "occupy": 0, "inbound": 0, "inQty": 300}},
                ), \
                patch(
                    "backend.spu_plan.service._load_sales_windows",
                    return_value={"SKU-01": windows},
                ), \
                patch("backend.spu_plan.service.load_in_transit", return_value={}):
            result = build_style_alerts("unused.env", today="2026-08-21", board="baihuo")
        style = result["styles"][0]
        self.assertEqual("baihuo", result["board"])
        self.assertEqual(0.3, style["dailyAvg"])
        self.assertEqual(10.0, style["monthlySales"])
        self.assertEqual("SKU-01", style["styleId"])
        self.assertEqual("杯", style["name"])
        self.assertEqual(30.0, style["sales30"])
        self.assertEqual(0, style["brokenSkus"])
        self.assertEqual(0, style["shortSkus"])
        self.assertEqual(100.0, style["onHand"])
        self.assertEqual(300.0, style["inQty"])
        # 日均 0.3 × 30 − 100 = −91；进货仓不进总库存
        self.assertEqual(-91, style["replenishQty"])

    def test_baihuo_rows_are_sku_not_style(self):
        products = [
            {"skuId": "SKU-A", "styleId": "BH-1", "name": "小酷公仔",
             "categoryLine": "文创百货", "productionMode": "", "labels": ["自营百货"]},
            {"skuId": "SKU-B", "styleId": "BH-1", "name": "小帅公仔",
             "categoryLine": "文创百货", "productionMode": "", "labels": ["自营百货"]},
        ]
        inventory = {
            "SKU-A": {"qty": 10, "occupy": 0, "inbound": 0, "inQty": 0},
            "SKU-B": {"qty": 4, "occupy": 0, "inbound": 0, "inQty": 0},
        }
        windows_a = {
            "1": 0, "3": 0, "7": 4, "14": 8, "15": 9, "30": 13,
            "prev7": 0, "days": [0.0] * 30,
            "online": {"7": 3, "15": 8, "30": 12},
            "offline": {"7": 1, "15": 1, "30": 1},
        }
        windows_b = {
            "1": 0, "3": 0, "7": 2, "14": 6, "15": 7, "30": 8,
            "prev7": 0, "days": [0.0] * 30,
            "online": {"7": 0, "15": 0, "30": 0},
            "offline": {"7": 2, "15": 7, "30": 8},
        }
        with patch("backend.spu_plan.service.load_products", return_value=products), \
                patch("backend.spu_plan.service._load_inventory", return_value=inventory), \
                patch(
                    "backend.spu_plan.service._load_sales_windows",
                    return_value={"SKU-A": windows_a, "SKU-B": windows_b},
                ), \
                patch("backend.spu_plan.service.load_in_transit", return_value={}):
            result = build_style_alerts("unused.env", today="2026-08-21", board="baihuo")
        ids = [item["styleId"] for item in result["styles"]]
        self.assertEqual(["SKU-A", "SKU-B"], ids)
        by_id = {item["styleId"]: item for item in result["styles"]}
        self.assertEqual("小酷公仔", by_id["SKU-A"]["name"])
        self.assertEqual(3.0, by_id["SKU-A"]["sales7Online"])
        self.assertEqual(1.0, by_id["SKU-A"]["sales7Offline"])
        self.assertEqual(8.0, by_id["SKU-A"]["sales15Online"])
        self.assertEqual(12.0, by_id["SKU-A"]["sales30Online"])
        self.assertEqual(1.0, by_id["SKU-A"]["sales30Offline"])
        self.assertEqual(8.0, by_id["SKU-A"]["sales14"])
        self.assertEqual(8.0, by_id["SKU-B"]["sales30Offline"])

    def test_tagged_or_non_own_styles_are_excluded(self):
        """2026-08-20 拍板：只出自营；带清仓/淘汰/有升级的整款不进表。"""
        products = [
            {"skuId": "SKU-01", "styleId": "S1", "name": "清仓鞋",
             "categoryLine": "鞋类", "productionMode": "自营", "labels": ["清仓品"]},
            {"skuId": "SKU-02", "styleId": "S2", "name": "代发鞋",
             "categoryLine": "鞋类", "productionMode": "一件代发", "labels": []},
            {"skuId": "SKU-03", "styleId": "S3", "name": "自营鞋",
             "categoryLine": "鞋类", "productionMode": "自营", "labels": ["热销"]},
        ]
        inventory = {
            key: {"qty": 5.0, "occupy": 0.0, "inbound": 0.0}
            for key in ("SKU-01", "SKU-02", "SKU-03")
        }
        with patch("backend.spu_plan.service.load_products", return_value=products), \
                patch("backend.spu_plan.service._load_inventory", return_value=inventory), \
                patch("backend.spu_plan.service._load_sales_windows", return_value={}), \
                patch("backend.spu_plan.service.load_in_transit", return_value={}):
            result = build_style_alerts("unused.env", today="2026-08-19")
        self.assertEqual(["S3"], [item["styleId"] for item in result["styles"]])

    def test_baihuo_board_uses_label_and_skips_clearance(self):
        """自营百货看板按标签圈款，不要求生产模式=自营；清仓整款仍剔除。"""
        products = [
            {"skuId": "G1", "styleId": "B1", "name": "杯子",
             "categoryLine": "文创百货", "productionMode": "", "labels": ["自营百货"]},
            {"skuId": "G2", "styleId": "B2", "name": "清仓杯",
             "categoryLine": "文创百货", "productionMode": "自营",
             "labels": ["自营百货", "清仓品"]},
            {"skuId": "S1", "styleId": "S1", "name": "鞋",
             "categoryLine": "鞋类", "productionMode": "自营", "labels": []},
        ]
        inventory = {
            key: {"qty": 8.0, "occupy": 0.0, "inbound": 0.0}
            for key in ("G1", "G2", "S1")
        }
        with patch("backend.spu_plan.service.load_products", return_value=products), \
                patch("backend.spu_plan.service._load_inventory", return_value=inventory), \
                patch("backend.spu_plan.service._load_sales_windows", return_value={}), \
                patch("backend.spu_plan.service.load_in_transit", return_value={}):
            result = build_style_alerts("unused.env", today="2026-08-21", board="baihuo")
        self.assertEqual("baihuo", result["board"])
        self.assertEqual(["G1"], [item["styleId"] for item in result["styles"]])
        self.assertEqual("杯子", result["styles"][0]["name"])

    def test_week_over_week_ratio(self):
        products = [{
            "skuId": "SKU-01", "styleId": "S1", "name": "鞋",
            "categoryLine": "鞋类", "productionMode": "自营", "labels": [],
            "moq": 150,
        }]
        inventory = {"SKU-01": {"qty": 100.0, "occupy": 0.0, "inbound": 0.0}}
        with patch("backend.spu_plan.service.load_products", return_value=products), \
                patch("backend.spu_plan.service._load_inventory", return_value=inventory), \
                patch(
                    "backend.spu_plan.service._load_sales_windows",
                    return_value={"SKU-01": {
                        "1": 2, "3": 6, "7": 14, "15": 30, "prev7": 7,
                    }},
                ), \
                patch("backend.spu_plan.service.load_in_transit", return_value={}):
            result = build_style_alerts("unused.env", today="2026-08-19")
        style = result["styles"][0]
        self.assertEqual(7.0, style["salesPrev7"])
        self.assertEqual(1.0, style["wowRatio"])  # (14-7)/7
        self.assertEqual(30, len(style["salesDaily"]))
        self.assertEqual(150, style["moq"])
        # 建议 20 件，起订 150 → 按起订量抬到 150
        self.assertEqual(20, style["replenishQty"])
        self.assertEqual(150, style["orderQty"])
        no_prev = style_warning  # noqa: F841 保持导入使用
        with patch("backend.spu_plan.service.load_products", return_value=products), \
                patch("backend.spu_plan.service._load_inventory", return_value=inventory), \
                patch(
                    "backend.spu_plan.service._load_sales_windows",
                    return_value={"SKU-01": {"7": 14}},
                ), \
                patch("backend.spu_plan.service.load_in_transit", return_value={}):
            result = build_style_alerts("unused.env", today="2026-08-19")
        self.assertIsNone(result["styles"][0]["wowRatio"])


class _SnapshotCursor:
    def __init__(self, conn):
        self.conn = conn

    def __enter__(self):
        return self

    def __exit__(self, *args):
        return False

    def execute(self, sql, args=None):
        text = " ".join(str(sql).split())
        self.conn.calls.append((text, tuple(args or ())))
        upper = text.upper()
        if upper.startswith("DELETE") and "NOT IN" in upper:
            keep = {str(item) for item in (args or ())}
            self.conn.rows = {key: value for key, value in self.conn.rows.items() if key in keep}
            return
        if upper.startswith("DELETE"):
            self.conn.rows.clear()

    def executemany(self, sql, rows):
        self.conn.calls.append((str(sql), list(rows)))
        for row in rows:
            self.conn.rows[str(row[0])] = tuple(row)


class _SnapshotConn:
    def __init__(self, rows=None):
        self.rows = dict(rows or {})
        self.calls = []
        self.commits = 0
        self.rollbacks = 0

    def __enter__(self):
        return self

    def __exit__(self, *args):
        return False

    def cursor(self):
        return _SnapshotCursor(self)

    def commit(self):
        self.commits += 1

    def rollback(self):
        self.rollbacks += 1


def _style_row(style_id, **overrides):
    item = {
        "styleId": style_id, "name": style_id, "categoryLine": "鞋类",
        "productionMode": "自营", "season": "四季", "category": "作训鞋",
        "salePrice": 99, "costPrice": 40, "year": "25",
        "skuCount": 2, "labels": ["热销"],
        "sales1": 1, "sales3": 3, "sales7": 7, "sales15": 15,
        "sales30": 30, "sales45": 45, "sales60": 60,
        "brokenSkus": 0, "shortSkus": 0,
        "turnoverDays": 12.5, "stockoutLabel": "",
        "replenishQty": 10, "dailyAvg": 10,
        "onHand": 8, "qty": 10, "occupy": 4, "inbound": 2,
        "remark": "", "missingInventory": 0,
    }
    item.update(overrides)
    return item


class SnapshotTests(unittest.TestCase):
    def test_table_name_matches_mirror_ddl(self):
        from backend.realtime_mirror import (
            BAIHUO_STYLE_SNAPSHOT_TABLE, SCHEMA_SQL, SPU_STYLE_SNAPSHOT_TABLE,
        )
        from backend.spu_plan import BAIHUO_SNAPSHOT_TABLE
        self.assertEqual(SPU_STYLE_SNAPSHOT_TABLE, SNAPSHOT_TABLE)
        self.assertEqual(BAIHUO_STYLE_SNAPSHOT_TABLE, BAIHUO_SNAPSHOT_TABLE)
        self.assertEqual("style_id", SNAPSHOT_COLUMNS[0])
        self.assertTrue(any(SNAPSHOT_TABLE in sql for sql in SCHEMA_SQL))
        self.assertTrue(any(BAIHUO_SNAPSHOT_TABLE in sql for sql in SCHEMA_SQL))
        from backend.spu_plan.service import _snapshot_tuple
        row = _snapshot_tuple(_style_row("S1"), "2026-08-21 12:00:00")
        self.assertEqual(len(SNAPSHOT_COLUMNS), len(row))
        self.assertIn("sales_14", SNAPSHOT_COLUMNS)
        self.assertIn("sales_7_online", SNAPSHOT_COLUMNS)

    def test_upsert_overwrites_and_deletes_styles_out_of_scope(self):
        conn = _SnapshotConn({
            "OLD": ("OLD",),
            "KEEP": ("KEEP", "旧名"),
        })
        result = {
            "styles": [
                _style_row("KEEP", name="新名", qty=20),
                _style_row("NEW", name="新品"),
            ],
        }
        with patch("backend.realtime_mirror.ensure_schema"), \
                patch("backend.spu_plan.service.connect", return_value=conn):
            written = save_style_snapshot("unused.env", result)
        self.assertEqual(2, written)
        self.assertEqual(1, conn.commits)
        self.assertNotIn("OLD", conn.rows)
        self.assertIn("KEEP", conn.rows)
        self.assertIn("NEW", conn.rows)
        self.assertEqual("新名", conn.rows["KEEP"][2])
        self.assertEqual(20, conn.rows["KEEP"][SNAPSHOT_COLUMNS.index("qty")])
        delete_sql = next(sql for sql, _args in conn.calls if sql.upper().startswith("DELETE"))
        self.assertIn("NOT IN", delete_sql)
        insert_sql = next(sql for sql, _args in conn.calls if "INSERT" in sql.upper())
        self.assertIn("ON DUPLICATE KEY UPDATE", insert_sql)

    def test_empty_result_clears_snapshot(self):
        conn = _SnapshotConn({"GONE": ("GONE",)})
        with patch("backend.realtime_mirror.ensure_schema"), \
                patch("backend.spu_plan.service.connect", return_value=conn):
            written = save_style_snapshot("unused.env", {"styles": []})
        self.assertEqual(0, written)
        self.assertEqual({}, conn.rows)
        self.assertEqual(1, conn.commits)
        delete_sql = next(
            sql for sql, _args in conn.calls if not sql.upper().startswith("ALTER")
        )
        self.assertTrue(delete_sql.upper().startswith("DELETE"))

    def test_baihuo_save_uses_separate_table(self):
        conn = _SnapshotConn({})
        with patch("backend.realtime_mirror.ensure_schema"), \
                patch("backend.spu_plan.service.connect", return_value=conn):
            save_style_snapshot("unused.env", {
                "board": "baihuo",
                "styles": [_style_row("B1", name="杯")],
            }, board="baihuo")
        insert_sql = next(sql for sql, _args in conn.calls if "INSERT" in sql.upper())
        self.assertIn("baihuo_style_snapshot", insert_sql)
        delete_sql = next(sql for sql, _args in conn.calls if sql.upper().startswith("DELETE"))
        self.assertIn("baihuo_style_snapshot", delete_sql)
        self.assertIn("NOT IN", delete_sql)


class _ReadCursor:
    def __init__(self, rows):
        self._rows = rows

    def __enter__(self):
        return self

    def __exit__(self, *args):
        return False

    def execute(self, sql, args=None):
        self._last = str(sql)

    def fetchall(self):
        return list(self._rows)

    def fetchone(self):
        return {"n": 1}


class _ReadConn:
    def __init__(self, rows):
        self.rows = rows

    def __enter__(self):
        return self

    def __exit__(self, *args):
        return False

    def cursor(self):
        return _ReadCursor(self.rows)


class SnapshotReadTests(unittest.TestCase):
    def test_load_style_snapshot_maps_rows(self):
        from decimal import Decimal
        from backend.spu_plan import load_style_snapshot

        rows = [{
            "style_id": "S1", "category_line": "鞋类", "name": "作训鞋",
            "sku_count": 3, "sales_1": Decimal("2"), "sales_7": Decimal("14"),
            "sales_prev7": Decimal("7"), "wow_ratio": Decimal("1.0"),
            "sales_60": Decimal("120"), "sales_90": Decimal("180"),
            "in_qty": Decimal("300"),
            "daily_avg": Decimal("2.0"),
            "turnover_days": Decimal("12.5"), "stockout_label": "缺货",
            "broken_skus": 1, "short_skus": 0, "on_hand": Decimal("25"),
            "qty": Decimal("30"), "occupy": Decimal("10"), "inbound": Decimal("5"),
            "replenish_qty": 95, "remark": "", "computed_at": "2026-08-20 11:00:00",
            "year": "25", "season": "夏季", "category": "T恤",
            "sale_price": Decimal("108"), "cost_price": Decimal("41.5"),
            "labels": "店加，新品", "missing_inventory": 2, "production_mode": "自营",
        }]
        with patch("backend.spu_plan.service.connect", return_value=_ReadConn(rows)), \
                patch("backend.spu_plan.service.fetch_last_suppliers", return_value={"S1": "佰特"}):
            payload = load_style_snapshot("unused.env")
        self.assertTrue(payload["ok"])
        self.assertEqual(1, payload["styleCount"])
        self.assertEqual(1, payload["stockoutCount"])
        self.assertEqual("2026-08-20 11:00:00", payload["computedAt"])
        style = payload["styles"][0]
        self.assertEqual("S1", style["styleId"])
        self.assertTrue(style["stockout"])
        self.assertEqual(1.0, style["wowRatio"])
        self.assertEqual(12.5, style["turnoverDays"])
        self.assertEqual(95, style["replenishQty"])
        self.assertEqual(180.0, style["sales90"])
        self.assertEqual(300.0, style["inQty"])
        self.assertEqual("夏季", style["season"])
        self.assertEqual("T恤", style["category"])
        self.assertEqual(["店加", "新品"], style["labels"])
        self.assertEqual(2, style["missingInventory"])
        self.assertEqual("佰特", style["lastSupplier"])


class SchedulerTests(unittest.TestCase):
    def test_daily_tick_runs_once_after_run_time(self):
        from datetime import datetime
        from backend.spu_plan.scheduler import DailySpuSnapshotScheduler

        scheduler = DailySpuSnapshotScheduler(env_path="unused.env", run_time="09:00")
        calls = []
        scheduler.run_once = lambda today="": calls.append(today) or {"ok": True, "today": today}

        early = scheduler.tick(now=datetime(2026, 8, 20, 8, 59))
        self.assertTrue(early.get("skipped"))
        scheduler.tick(now=datetime(2026, 8, 20, 9, 0))
        scheduler.last_run = "2026-08-20"
        again = scheduler.tick(now=datetime(2026, 8, 20, 10, 0))
        self.assertEqual("今日已重算", again.get("reason"))
        self.assertEqual(["2026-08-20"], calls)

    def test_plan_regen_skips_without_source(self):
        from backend.spu_plan.scheduler import DailySpuSnapshotScheduler

        scheduler = DailySpuSnapshotScheduler(env_path="unused.env", plan_source="")
        scheduler._run_plan()  # 未配置源文件：不报错、不留错误
        self.assertEqual("", scheduler.plan_last_error)
        missing = DailySpuSnapshotScheduler(
            env_path="unused.env", plan_source=r"Z:\不存在\订货表.xlsx",
        )
        missing._run_plan()
        self.assertIn("找不到", missing.plan_last_error)

    def test_plan_source_upload_validates_and_saves(self):
        import io as io_module
        from backend.spu_plan.plan_source import (
            PlanSourceError, PlanSourceUpdater, validate_plan_workbook,
        )

        with self.assertRaises(PlanSourceError):
            validate_plan_workbook(b"")
        with self.assertRaises(PlanSourceError):
            validate_plan_workbook(b"not an xlsx")

        with tempfile.TemporaryDirectory() as tmp:
            src = Path(tmp) / "上传.xlsx"
            _make_source_workbook(src)
            data = src.read_bytes()
            checked = validate_plan_workbook(data)
            self.assertGreaterEqual(checked["styles"], 1)

            dest = Path(tmp) / "config" / "重点产品订货表.xlsx"
            updater = PlanSourceUpdater(env_path="unused.env", source_path=str(dest))
            done = []
            with patch(
                "backend.spu_plan.production_plan.build_production_plan",
                return_value={"styles": 3, "added": [], "dropped": [],
                              "missingDemand": [], "xlsx": "x.xlsx", "today": ""},
            ):
                result = updater.update(data, origin="test", notify=done.append)
                self.assertTrue(result["started"])
                # 等后台线程收尾
                for _ in range(50):
                    if not updater.state["running"]:
                        break
                    import time as time_module
                    time_module.sleep(0.05)
            self.assertTrue(dest.exists())
            self.assertEqual(data, dest.read_bytes())
            self.assertEqual("", updater.state["lastError"])
            self.assertTrue(done and "重生成完成" in done[0])
            # 缺工作表的文件被拒
            bad = io_module.BytesIO()
            from openpyxl import Workbook
            book = Workbook(); book.active.title = "别的表"; book.save(bad)
            with self.assertRaises(PlanSourceError):
                updater.update(bad.getvalue(), origin="test")

    def test_run_once_skips_when_lock_busy(self):
        import threading
        from backend.spu_plan.scheduler import DailySpuSnapshotScheduler

        lock = threading.Lock()
        scheduler = DailySpuSnapshotScheduler(env_path="unused.env", lock=lock)
        lock.acquire()
        try:
            result = scheduler.run_once(today="2026-08-20")
        finally:
            lock.release()
        self.assertTrue(result.get("skipped"))
        self.assertIn("重算", result.get("reason") or "")

    def test_run_once_writes_apparel_and_baihuo(self):
        from backend.spu_plan.scheduler import DailySpuSnapshotScheduler

        calls = []
        empty = {
            "styles": [], "today": "2026-08-21",
            "styleCount": 0, "stockoutCount": 0, "brokenStyleCount": 0,
            "shortStyleCount": 0,
        }

        def fake_build(env, board="apparel"):
            calls.append(board)
            return {**empty, "board": board}

        scheduler = DailySpuSnapshotScheduler(env_path="unused.env")
        with patch("backend.spu_plan.scheduler.build_style_alerts", side_effect=fake_build), \
                patch("backend.spu_plan.scheduler.save_style_snapshot", return_value=0), \
                patch("backend.spu_plan.scheduler.write_style_workbook"), \
                patch("backend.spu_plan.scheduler.style_workbook_path", return_value=Path("x.xlsx")):
            result = scheduler.run_once(today="2026-08-21")
        self.assertTrue(result.get("ok"))
        self.assertEqual(["apparel", "baihuo"], calls)


def _make_source_workbook(path):
    """最小化的员工订货工作簿：两张订货表 + 库存 + 生产计划表。"""
    from openpyxl import Workbook
    from openpyxl.utils import column_index_from_string as ci

    book = Workbook()
    key = book.active
    key.title = "重点产品订货"
    key.cell(4, 2, "S1")
    # EW..FG = 2026-04 .. 2027-02 的最终需求
    for offset, value in enumerate([0, 100, 110, 120, 130, 140, 150, 160, 170, 180, 190]):
        key.cell(4, ci("EW") + offset, value)

    hot = book.create_sheet("爆品订货")
    hot.cell(4, 2, "S2")
    # DK..DX = 2026-04 .. 2027-05 的总计需求
    for offset, value in enumerate([0, 0, 0, 0, 800, 900, 1000, 1100, 1200, 1300, 1400, 0, 0, 0]):
        hot.cell(4, ci("DK") + offset, value)

    stock = book.create_sheet("库存")
    stock.cell(3, 1, "S1"); stock.cell(3, 2, 55); stock.cell(3, 3, 20)
    stock.cell(4, 1, "S2"); stock.cell(4, 2, 700); stock.cell(4, 3, 0)

    plan = book.create_sheet("生产计划表")
    plan.cell(2, 6, "期初库存7.31")
    rows = [
        ("刃海", "服装非通勤裤", "S1", "训练袜", "常青款", 66, 77, 88),
        ("清松", "通勤裤", "S2", "通勤裤", "/", 111, 222, 333),
        ("清松", "通勤裤", "S3", "订货表没有的款", "/", None, None, None),
    ]
    for offset, row in enumerate(rows):
        r = 6 + offset
        plan.cell(r, 1, row[0]); plan.cell(r, 2, row[1]); plan.cell(r, 3, row[2])
        plan.cell(r, 4, row[3]); plan.cell(r, 5, row[4])
        plan.cell(r, 15, row[5]); plan.cell(r, 18, row[6]); plan.cell(r, 21, row[7])
    book.save(path)


class ProductionPlanTests(unittest.TestCase):
    def test_month_helpers(self):
        from backend.spu_plan.production_plan import month_seq, shift_month
        self.assertEqual([(2026, 11), (2026, 12), (2027, 1)], month_seq((2026, 11), 3))
        self.assertEqual((2027, 1), shift_month((2026, 8), 5))
        self.assertEqual((2026, 5), shift_month((2026, 8), -3))

    def test_month_verdict_matches_excel(self):
        from backend.spu_plan.alerts import (
            collect_month_alerts, format_plan_alert_markdown, month_gap, month_verdict,
            push_plan_workbook,
        )
        self.assertEqual("库存满足", month_verdict(100, 0, 90))
        self.assertEqual("及时入库", month_verdict(80, 30, 90))
        self.assertEqual("需补货", month_verdict(80, 5, 90))
        self.assertEqual("", month_verdict(100, 0, None))
        # Excel 用 > 不是 >=：期初刚好等于需求、没有在途 → 需补货
        self.assertEqual("需补货", month_verdict(90, 0, 90))
        self.assertEqual("及时入库", month_verdict(90, 1, 90))
        self.assertEqual(10.0, month_gap(80, 10, 100))
        self.assertEqual(0.0, month_gap(80, 30, 100))

        source = {
            "styles": [
                {"styleId": "S1", "owner": "刃海", "line": "鞋类", "name": "有缺口"},
                {"styleId": "S2", "owner": "清松", "line": "服装", "name": "等在途"},
                {"styleId": "S3", "owner": "清松", "line": "服装", "name": "够用"},
                {"styleId": "S4", "owner": "清松", "line": "服装", "name": "零需求不推"},
            ],
            "demands": {
                "S1": {(2026, 8): 100.0},
                "S2": {(2026, 8): 80.0},
                "S3": {(2026, 8): 50.0},
                "S4": {(2026, 8): 0.0},
            },
            "opening": {
                "S1": (40.0, 10.0), "S2": (50.0, 40.0),
                "S3": (80.0, 0.0), "S4": (0.0, 0.0),
            },
        }
        from datetime import date as date_cls
        alerts = collect_month_alerts(source, {}, date_cls(2026, 8, 20))
        self.assertEqual(["S1"], [row["styleId"] for row in alerts["replenish"]])
        self.assertEqual(["S2"], [row["styleId"] for row in alerts["inbound"]])
        self.assertEqual(50.0, alerts["replenish"][0]["gap"])
        text = format_plan_alert_markdown(alerts)
        self.assertIn("需补货 1 款", text)
        self.assertIn("及时入库 1 款", text)
        self.assertIn("S1", text)
        self.assertNotIn("S3", text)
        self.assertNotIn("S4", text)

        class FakeSender:
            configured = True
            app_ready = True
            group_conversation_id = "cid"
            def __init__(self):
                self.markdown = []
                self.files = []
            def send_markdown(self, title, text):
                self.markdown.append((title, text))
                return {"channel": "app"}
            def upload_media(self, path, filetype="file"):
                return {"mediaId": "mid"}
            def send_file(self, conversation_id, media_id, name, file_type="xlsx"):
                self.files.append((conversation_id, media_id, name, file_type))
                return {"channel": "app"}

        sender = FakeSender()
        missing = push_plan_workbook("missing.xlsx", sender=sender, force=True)
        self.assertTrue(missing.get("skipped"))
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "2608-生产计划表.xlsx"
            path.write_bytes(b"xlsx")
            sent = push_plan_workbook(path, sender=sender, force=True, operator="test")
        self.assertTrue(sent["sent"])
        self.assertEqual([], sender.markdown)
        self.assertEqual([("cid", "mid", "2608-生产计划表.xlsx", "xlsx")], sender.files)

    def test_net_sales_subtracts_received_returns(self):
        from backend.spu_plan.production_plan import (
            apply_monthly_returns, counts_as_net_return,
        )
        # 阿宝 XZ26401302-028：出库 4135 − 实收 643 = 示例 3492
        self.assertEqual(
            {(2026, 7): 3492.0},
            apply_monthly_returns({(2026, 7): 4135.0}, {(2026, 7): 643.0}),
        )
        self.assertEqual(
            {(2026, 7): -10.0},
            apply_monthly_returns({}, {(2026, 7): 10.0}),
        )
        self.assertTrue(counts_as_net_return("普通退货", "普通退货", "Confirmed", 2))
        self.assertTrue(counts_as_net_return("换货", "", "Confirmed", 1))
        self.assertFalse(counts_as_net_return("仅退款", "", "Confirmed", 2))
        self.assertFalse(counts_as_net_return("普通退货", "仅退款", "Confirmed", 2))
        self.assertFalse(counts_as_net_return("投诉", "", "Confirmed", 2))
        self.assertFalse(counts_as_net_return("普通退货", "", "Cancelled", 2))
        self.assertFalse(counts_as_net_return("普通退货", "", "Confirmed", 0))

        from datetime import date as date_cls
        from backend.spu_plan.production_plan import pick_past_net, returns_cover_month
        self.assertFalse(returns_cover_month(None, (2026, 7)))
        self.assertFalse(returns_cover_month(date_cls(2026, 6, 7), (2026, 7)))
        self.assertFalse(returns_cover_month(date_cls(2026, 7, 15), (2026, 7)))
        self.assertTrue(returns_cover_month(date_cls(2026, 7, 31), (2026, 7)))
        self.assertTrue(returns_cover_month("2026-08-01 12:00:00", (2026, 7)))
        self.assertEqual(
            88,
            pick_past_net((2026, 7), {(2026, 7): 4135}, 88, date_cls(2026, 6, 7)),
        )
        self.assertEqual(
            3492,
            pick_past_net((2026, 7), {(2026, 7): 3492}, 88, date_cls(2026, 7, 31)),
        )
        self.assertEqual(
            66,
            pick_past_net((2026, 5), {(2026, 5): 1}, 66, date_cls(2026, 8, 1)),
        )

    def test_merge_roster_follows_tag(self):
        from backend.spu_plan.production_plan import merge_roster
        tagged = [
            {"styleId": "S2", "line": "通勤裤", "name": "通勤裤（镜像名）"},
            {"styleId": "S9", "line": "鞋类", "name": "新打标款"},
        ]
        source = [
            {"styleId": "S1", "owner": "刃海", "line": "鞋类", "name": "已摘标",
             "node": "", "pastNet": [1, 2, 3]},
            {"styleId": "S2", "owner": "清松", "line": "通勤裤", "name": "通勤裤（表内名）",
             "node": "常青款", "pastNet": [4, 5, 6]},
        ]
        merged = merge_roster(tagged, source)
        self.assertEqual(["S9", "S2"], [s["styleId"] for s in merged["styles"]])  # 鞋类在前
        keep = merged["styles"][1]
        self.assertEqual("清松", keep["owner"])
        self.assertEqual("通勤裤（表内名）", keep["name"])
        self.assertEqual([4, 5, 6], keep["pastNet"])
        new = merged["styles"][0]
        self.assertEqual("", new["owner"])
        self.assertEqual("新打标款", new["name"])
        self.assertEqual(["S9"], merged["added"])
        self.assertEqual(["S1"], merged["dropped"])

    def test_read_source_and_write_plan(self):
        from datetime import date as date_cls
        from openpyxl import load_workbook
        from backend.spu_plan.production_plan import (
            read_source_plan, write_production_plan,
        )

        with tempfile.TemporaryDirectory() as tmp:
            src = Path(tmp) / "订货表.xlsx"
            _make_source_workbook(src)
            source = read_source_plan(src)
            self.assertEqual(["S1", "S2", "S3"], [s["styleId"] for s in source["styles"]])
            self.assertEqual((55.0, 20.0), source["opening"]["S1"])
            # 需求按月对齐：2026-08 = EW+4 / DK+4
            self.assertEqual(130.0, source["demands"]["S1"][(2026, 8)])
            self.assertEqual(800.0, source["demands"]["S2"][(2026, 8)])
            # 11 月按 11 月取（员工原表错引 12 月，这里必须对齐）
            self.assertEqual(160.0, source["demands"]["S1"][(2026, 11)])
            self.assertNotIn("S3", source["demands"])

            live = {
                "S1": {"qty": 50.0, "transit": 10.0,
                       "byMonth": {(2026, 8): 40.0, (2026, 7): 70.0}},
                "S2": {"qty": 900.0, "transit": 0.0, "byMonth": {(2026, 8): 500.0}},
                "S3": {"qty": 5.0, "transit": 0.0, "byMonth": {}},
            }
            out = Path(tmp) / "生产计划表.xlsx"
            write_production_plan(source, live, out, today=date_cls(2026, 8, 20))
            book = load_workbook(out)
            ws = book["生产计划表"]
            self.assertEqual("K5", ws.freeze_panes)
            # 表头月份组：当月 8 月在 W 列（5固定+2期初+3实时+3判定+3×3历史 = 22 列后）
            self.assertEqual("8月", ws.cell(2, 23).value)
            self.assertEqual(130, ws.cell(5, 23).value)   # S1 当月需求
            self.assertEqual(40, ws.cell(5, 10).value)    # S1 当月净销量（镜像）
            self.assertEqual(88, ws.cell(5, 21).value)    # 7 月售后未齐，沿用手贴
            self.assertEqual(66, ws.cell(5, 15).value)    # 5 月净销量沿用手贴值
            self.assertIn("7月售后未齐", str(ws.cell(1, 8).value))
            book.close()
            write_production_plan(
                source, live, out, today=date_cls(2026, 8, 20),
                returns_through=date_cls(2026, 7, 31),
            )
            book = load_workbook(out)
            ws = book["生产计划表"]
            self.assertEqual(70, ws.cell(5, 21).value)    # 售后齐了才走镜像净销
            self.assertNotIn("7月售后未齐", str(ws.cell(1, 8).value))
            self.assertEqual(160, ws.cell(5, 32).value)   # 11 月需求 = 160（修正后）
            self.assertTrue(str(ws.cell(5, 24).value).startswith("=IF("))   # 计划入库数公式
            self.assertTrue(str(ws.cell(5, 11).value).startswith("=IFERROR("))  # 销售进度
            # S3 在订货表里没有需求：需求列为空
            self.assertIsNone(ws.cell(7, 23).value)
            # 历史月折叠：5 月整组 + 6 月需求/净销收起，6/7 月售罄率露出；订货节点折进名称
            self.assertFalse(ws.sheet_properties.outlinePr.summaryRight)
            self.assertEqual(1, ws.column_dimensions["N"].outlineLevel)
            self.assertTrue(ws.column_dimensions["N"].hidden)
            self.assertTrue(ws.column_dimensions["P"].hidden)   # 5 月售罄率一并收起
            self.assertFalse(ws.column_dimensions["S"].hidden)  # 6 月售罄率
            self.assertEqual(0, ws.column_dimensions["S"].outlineLevel)
            self.assertEqual(1, ws.column_dimensions["T"].outlineLevel)
            self.assertTrue(ws.column_dimensions["T"].hidden)
            self.assertFalse(ws.column_dimensions["V"].hidden)  # 7 月售罄率
            self.assertTrue(ws.column_dimensions["M"].collapsed)
            self.assertEqual(1, ws.column_dimensions["E"].outlineLevel)
            self.assertTrue(ws.column_dimensions["E"].hidden)
            self.assertTrue(ws.column_dimensions["D"].collapsed)
            self.assertIn("FFF3CE", str(ws.cell(5, 9).fill.fgColor.rgb or ""))
            # 「是/否」「库存满足」用 containsText + bgColor，否则 Excel 不显示底色
            enough_rules = []
            verdict_rules = []
            for rng, rules in ws.conditional_formatting._cf_rules.items():
                target = str(getattr(rng, "sqref", rng))
                if target.startswith("M"):
                    enough_rules.extend(rules)
                if target.startswith("X"):
                    verdict_rules.extend(rules)
            self.assertTrue(enough_rules)
            self.assertTrue(verdict_rules)
            self.assertEqual("containsText", enough_rules[0].type)
            self.assertIn("C9EDD1", str(enough_rules[0].dxf.fill.bgColor.rgb or ""))
            self.assertEqual("containsText", verdict_rules[0].type)
            self.assertTrue(any(
                str(rule.dxf.fill.bgColor.rgb or "").endswith("C9EDD1")
                or str(rule.dxf.fill.bgColor.rgb or "").endswith("FCE3A1")
                or str(rule.dxf.fill.bgColor.rgb or "").endswith("FBC8CF")
                for rule in verdict_rules
            ))
            book.close()


class WorkbookTests(unittest.TestCase):
    def test_baihuo_workbook_filename(self):
        from datetime import date
        from backend.spu_plan.workbook import style_workbook_filename

        self.assertTrue(style_workbook_filename(date(2026, 8, 21)).endswith("鞋服SPU总表.xlsx"))
        self.assertTrue(
            style_workbook_filename(date(2026, 8, 21), board="baihuo").endswith("自营百货总表.xlsx")
        )

    def test_writes_total_sheet_headers(self):
        result = {
            "today": "2026-08-19",
            "styles": [{
                "styleId": "BH21401306-027", "name": "作训鞋",
                "categoryLine": "鞋类", "productionMode": "自营",
                "season": "四季", "category": "作训鞋",
                "salePrice": 99, "costPrice": 40, "year": "21",
                "skuCount": 8, "labels": ["热销"],
                "sales1": 1, "sales3": 3, "sales7": 7, "sales15": 15,
                "sales30": 30, "sales45": 45, "sales60": 60,
                "salesPrev7": 4, "wowRatio": 0.75,
                "brokenSkus": 1, "shortSkus": 2,
                "turnoverDisplay": 12.5, "turnoverDays": 12.5,
                "stockout": True, "stockoutLabel": "缺货",
                "replenishQty": 40, "dailyAvg": 10,
                "onHand": 8, "qty": 10, "occupy": 4, "inbound": 2,
                "remark": "",
            }],
        }
        with tempfile.TemporaryDirectory() as tmp:
            path = write_style_workbook(result, Path(tmp) / "总表.xlsx")
            book = load_workbook(path, read_only=False, data_only=True)
            sheet = book["总表"]
            headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
            self.assertEqual(list(TOTAL_COLUMNS), headers)
            row = [cell.value for cell in next(sheet.iter_rows(min_row=2, max_row=2))]
            self.assertEqual("鞋类", row[0])
            self.assertEqual("BH21401306-027", row[1])
            self.assertEqual(60, row[18])
            self.assertEqual(4, row[19])   # 前7天合计
            self.assertEqual(7, row[20])   # 近7天合计
            self.assertEqual(0.75, row[21])
            self.assertEqual("缺货", row[25])
            self.assertEqual(40, row[26])
            self.assertNotIn("25年净销量", TOTAL_COLUMNS)
            self.assertEqual(
                ["前7天合计", "近7天合计", "周环比销量"], TOTAL_COLUMNS[19:22],
            )
            # 版式对齐员工示例：橙表头、微软雅黑、冻结 Q2、条件格式
            self.assertEqual("Q2", sheet.freeze_panes)
            head = sheet.cell(1, 1)
            self.assertEqual("微软雅黑", head.font.name)
            self.assertTrue(head.font.bold)
            self.assertIn("F4B482", str(head.fill.fgColor.rgb))
            self.assertEqual(47.0, sheet.row_dimensions[1].height)
            self.assertEqual(20.0, sheet.row_dimensions[2].height)
            data_cell = sheet.cell(2, 1)
            self.assertEqual(8.0, data_cell.font.size)
            self.assertEqual("thin", data_cell.border.left.style)
            self.assertGreaterEqual(len(list(sheet.conditional_formatting)), 3)
            book.close()

    def test_baihuo_workbook_inserts_14_and_channel_columns(self):
        result = {
            "board": "baihuo",
            "today": "2026-08-21",
            "styles": [{
                "styleId": "CUP01", "name": "杯",
                "categoryLine": "文创百货", "productionMode": "",
                "season": "", "category": "",
                "salePrice": 20, "costPrice": 8, "year": "26",
                "skuCount": 1, "labels": ["自营百货"],
                "sales1": 0, "sales3": 0, "sales7": 4, "sales14": 10, "sales15": 12,
                "sales30": 20, "sales45": 30, "sales60": 40,
                "sales7Online": 3, "sales7Offline": 1,
                "sales15Online": 8, "sales15Offline": 4,
                "sales30Online": 12, "sales30Offline": 8,
                "salesPrev7": 0, "wowRatio": None,
                "brokenSkus": 0, "shortSkus": 0,
                "turnoverDisplay": 20, "turnoverDays": 20,
                "stockout": False, "stockoutLabel": "",
                "replenishQty": 0, "dailyAvg": 0.8,
                "onHand": 50, "qty": 50, "occupy": 0, "inbound": 0,
                "remark": "",
            }],
        }
        with tempfile.TemporaryDirectory() as tmp:
            path = write_style_workbook(result, Path(tmp) / "百货.xlsx")
            book = load_workbook(path, read_only=True, data_only=True)
            sheet = book["总表"]
            headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
            row = [cell.value for cell in next(sheet.iter_rows(min_row=2, max_row=2))]
            self.assertIn("14天销量", headers)
            self.assertIn("进货仓库存", headers)
            self.assertIn("商品编码", headers)
            self.assertIn("商品名称", headers)
            self.assertNotIn("款式编码", headers)
            self.assertNotIn("品名", headers)
            self.assertNotIn("已断码数量", headers)
            self.assertNotIn("即将缺码数量", headers)
            self.assertNotIn("90天销量", headers)
            self.assertEqual("14天销量", headers[headers.index("7天销量") + 1])
            self.assertEqual(
                ["线上7天", "线下7天", "线上15天", "线下15天", "线上30天", "线下30天"],
                headers[headers.index("30天销量") + 1:headers.index("30天销量") + 7],
            )
            self.assertEqual(10, row[headers.index("14天销量")])
            self.assertEqual(3, row[headers.index("线上7天")])
            self.assertEqual(8, row[headers.index("线下30天")])
            book.close()


class AnalyzeCacheTests(unittest.TestCase):
    def test_sales_payload_uses_7_to_14_label(self):
        from backend.spu_plan.analyze import (
            baihuo_notes_payload, category_peers_payload, dispatch_tool, evidence_pack,
            season_notes_payload, style_sales_payload,
        )
        style = {
            "styleId": "S1", "name": "加绒通勤裤", "categoryLine": "通勤裤",
            "skuCount": 4, "sales1": 1, "sales3": 3, "sales7": 10,
            "salesPrev7": 8, "wowRatio": 0.25, "sales15": 15, "sales30": 30,
            "sales45": 45, "sales60": 60, "salesDaily": [1, 2], "dailyAvg": 2.5,
            "qty": 20, "occupy": 2, "inbound": 5, "onHand": 23,
            "turnoverDays": 9, "stockout": True, "brokenSkus": 1, "shortSkus": 0,
            "replenishQty": 40, "orderQty": 50, "moq": 50, "remark": "起订50",
        }
        pack = style_sales_payload(style, {"computedAt": "2026-08-20 09:00:00"})
        self.assertIn("7~14天", pack["销量窗口_不含当天"])
        self.assertNotIn("前7天_第8到14天", pack["销量窗口_不含当天"])
        self.assertEqual(8, pack["销量窗口_不含当天"]["7~14天"])
        self.assertEqual("四窗口加权", pack["日均口径"])
        cup = {
            "styleId": "CUP01", "name": "文创纪念杯", "categoryLine": "文创百货",
            "skuCount": 2, "sales1": 0, "sales7": 0, "sales14": 10, "sales15": 0, "sales30": 30,
            "salesDaily": [0] * 29 + [1],
            "dailyAvg": 0.3, "monthlySales": 10.0, "qty": 40, "occupy": 0, "inbound": 0, "onHand": 40,
            "turnoverDays": 133, "stockout": False, "brokenSkus": 0, "shortSkus": 0,
            "labels": ["自营百货"], "category": "杯子", "inQty": 300,
        }
        baihuo_pack = style_sales_payload(cup, {"computedAt": "2026-08-21 09:00:00", "board": "baihuo"})
        self.assertEqual("三窗折月", baihuo_pack["日均口径"])
        self.assertIn("三窗折月", baihuo_pack)
        self.assertEqual(10.0, baihuo_pack["三窗折月"]["月销量"])
        self.assertEqual(30, baihuo_pack["销量窗口_不含当天"]["近30天"])
        self.assertNotIn("近90天", baihuo_pack["销量窗口_不含当天"])
        self.assertNotIn("穿着季节", baihuo_pack)
        self.assertEqual("CUP01", baihuo_pack["商品编码"])
        self.assertEqual("文创纪念杯", baihuo_pack["商品名称"])
        self.assertEqual("", baihuo_pack["款式编码"])
        self.assertNotIn("断码SKU数", baihuo_pack)
        self.assertNotIn("缺码SKU数", baihuo_pack)
        self.assertNotIn("断码占比", baihuo_pack["研判要点"])
        self.assertIn("少补", baihuo_pack["表内公式说明"])
        from datetime import date as date_cls
        gift = baihuo_notes_payload(cup, today=date_cls(2026, 8, 21))
        self.assertTrue(any("慢动销" in item for item in gift["notes"]))
        self.assertTrue(any("礼品" in item or "文创" in item for item in gift["notes"]))
        self.assertTrue(any("进货仓" in item for item in gift["notes"]))
        self.assertFalse(any("断码" in item or "缺码" in item for item in gift["notes"]))
        self.assertEqual(300, baihuo_pack["库存"]["进货仓库存"])
        notes = season_notes_payload(style, today=date_cls(2026, 8, 20))
        self.assertEqual("冬季", notes["穿着季节"])
        self.assertEqual("前置备货", notes["阶段"])
        self.assertTrue(any("前置备货" in item for item in notes["notes"]))
        summer = season_notes_payload(
            {
                "styleId": "FS2616Z01-016",
                "name": "TZ603专用款冰爽酷无痕压胶7A抑菌短袖T恤",
                "categoryLine": "服装-非通勤裤",
                "season": "夏季",
                "category": "T恤",
            },
            today=date_cls(2026, 8, 20),
        )
        self.assertEqual("夏季", summer["穿着季节"])
        self.assertEqual("尾季清货", summer["阶段"])
        self.assertTrue(any("不要跟表上补货建议" in item for item in summer["notes"]))
        tagged = season_notes_payload(
            {
                "styleId": "FS2516Z01-016",
                "name": "TZ603专用款冰爽酷无痕压胶7A抑菌短袖T恤",
                "categoryLine": "服装-非通勤裤",
                "season": "夏季",
                "category": "T恤",
                "labels": ["新品"],
                "year": "25",
            },
            today=date_cls(2026, 8, 20),
        )
        self.assertTrue(any("新品" in item for item in tagged["notes"]))
        self.assertTrue(any("旧年款" in item for item in tagged["notes"]))
        ev = evidence_pack({
            "salesDaily": [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 10, 0, 0, 0, 0, 0, 1],
            "sales30": 11, "sales7": 3, "wowRatio": 0.5, "skuCount": 16,
            "brokenSkus": 15, "onHand": 1, "dailyAvg": 0.9, "inbound": 0,
            "qty": 2, "occupy": 2, "missingInventory": 1,
            "salePrice": 108, "costPrice": 41.5, "labels": ["新品"], "year": "26",
        })
        self.assertEqual("脉冲", ev["近30天形态"])
        self.assertEqual("量码都紧", ev["库存结构"])
        self.assertEqual(0.616, ev["资料毛利率"])
        self.assertTrue(ev["订单占有偏高"])
        self.assertEqual(1, ev["无库存记录SKU数"])
        peers = category_peers_payload(
            {"styles": [
                {**style, "category": "通勤裤"},
                {"styleId": "S2", "name": "对照", "categoryLine": "通勤裤",
                 "category": "其他", "sales7": 80, "salesPrev7": 10, "wowRatio": 1,
                 "turnoverDays": 12, "stockout": False},
                {"styleId": "S3", "name": "同品类", "categoryLine": "通勤裤",
                 "category": "通勤裤", "sales7": 20, "salesPrev7": 8, "wowRatio": 0,
                 "turnoverDays": 10, "stockout": False},
            ]},
            {**style, "category": "通勤裤"}, limit=3,
        )
        self.assertEqual(["S3", "S2"], [row["款式编码"] for row in peers["对照款"]])
        miss = dispatch_tool("spu_style_sales", {"styleId": "NO"}, snapshot={"styles": [style]}, today=date_cls(2026, 8, 20))
        self.assertIn("不在", miss["error"])

    def test_same_day_cache_skips_llm_and_next_day_is_stale(self):
        from datetime import date as date_cls
        from backend.spu_plan.analyze import (
            load_cached_analysis, run_style_analysis, save_cached_analysis,
        )
        snapshot = {"computedAt": "2026-08-20 09:00:00", "styles": [{
            "styleId": "S1", "name": "训练袜", "categoryLine": "鞋类",
            "sales7": 10, "salesPrev7": 4, "salesDaily": [],
        }]}

        class FakeLLM:
            configured = True
            model = "fake"
            def __init__(self):
                self.calls = 0
            def chat(self, messages, *, tools=None, tool_choice="auto"):
                self.calls += 1
                if self.calls == 1:
                    return {
                        "content": "",
                        "tool_calls": [{
                            "id": "c1",
                            "function": {"name": "spu_style_sales", "arguments": '{"styleId":"S1"}'},
                        }],
                    }
                return {"content": "趋势：近7天强于7~14天。\n库存：周转紧。\n建议：先补断码。\n复核：无", "tool_calls": []}

        with tempfile.TemporaryDirectory() as tmp:
            llm = FakeLLM()
            first = run_style_analysis(
                "S1", snapshot=snapshot, llm=llm,
                today=date_cls(2026, 8, 20), root=tmp,
            )
            self.assertFalse(first["cached"])
            self.assertGreaterEqual(llm.calls, 2)
            self.assertIn("7~14天", first["analysis"])
            llm.calls = 0
            again = run_style_analysis(
                "S1", snapshot=snapshot, llm=llm,
                today=date_cls(2026, 8, 20), root=tmp,
            )
            self.assertTrue(again["cached"])
            self.assertFalse(again["stale"])
            self.assertEqual(0, llm.calls)
            stale = load_cached_analysis("S1", today=date_cls(2026, 8, 21), root=tmp)
            self.assertTrue(stale["stale"])
            self.assertEqual(first["analysis"], stale["analysis"])
            from backend.spu_plan.analyze import load_day_analyses
            day_map = load_day_analyses(today=date_cls(2026, 8, 20), root=tmp)
            self.assertEqual(first["analysis"], day_map["S1"]["analysis"])
            self.assertFalse(day_map["S1"]["stale"])


if __name__ == "__main__":
    unittest.main()
