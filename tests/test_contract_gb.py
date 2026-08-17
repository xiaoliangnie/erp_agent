# -*- coding: utf-8 -*-
"""合同执行标准：排序、覆盖解析、Excel 列布局。离线，不连库。"""
import tempfile
import unittest
from pathlib import Path

from backend.contracts import (
    normalize_gb_overrides,
    pick_saved_standard,
    resolve_line_gb,
    saved_gb_option,
)
from backend.gb_standards import category_tokens, rank_standards


class RankStandardsTests(unittest.TestCase):
    def test_strips_category_ops_codes(self):
        self.assertEqual(["毛绒"], category_tokens("毛绒（04）", ""))
        self.assertIn("衬衫", category_tokens("衬衫", "男装"))

    def test_current_product_standard_with_name_token_ranks_first(self):
        rows = [
            {
                "standard_no": "GB/T 9832-2026", "name_cn": "毛绒、布制玩具",
                "status": "即将实施", "std_type": "产品", "issue_date": "2026-05-25",
            },
            {
                "standard_no": "GB/T 9832-2007", "name_cn": "毛绒、布制玩具",
                "status": "现行", "std_type": "产品", "issue_date": "2007-06-14",
            },
            {
                "standard_no": "GB 6675.1-2014", "name_cn": "玩具安全 第1部分：基本规范",
                "status": "现行", "std_type": "产品", "issue_date": "2014-05-06",
            },
            {
                "standard_no": "FZ/T 01057.1-2007", "name_cn": "纺织纤维鉴别试验方法",
                "status": "现行", "std_type": "方法", "issue_date": "2007-11-01",
            },
        ]
        ranked = rank_standards(rows, product_name="小熊", category="毛绒（04）")
        self.assertEqual("GB/T 9832-2007", ranked[0]["standard_no"])
        self.assertEqual("GB/T 9832-2026", ranked[1]["standard_no"])
        self.assertNotEqual("FZ/T 01057.1-2007", ranked[0]["standard_no"])


class GbOverrideTests(unittest.TestCase):
    def lookup(self, standard_no):
        catalog = {
            "GB/T 9832-2007": {
                "samr_id": "old", "standard_no": "GB/T 9832-2007",
                "name_cn": "毛绒、布制玩具", "status": "现行",
            },
            "GB/T 9832-2026": {
                "samr_id": "new", "standard_no": "GB/T 9832-2026",
                "name_cn": "毛绒、布制玩具", "status": "即将实施",
            },
            "GB/T 1-1993": {
                "samr_id": "dead", "standard_no": "GB/T 1-1993",
                "name_cn": "已废止示例", "status": "废止",
            },
        }
        return catalog.get(standard_no)

    def test_normalize_rejects_non_object(self):
        with self.assertRaises(ValueError):
            normalize_gb_overrides(["GB/T 9832-2026"])

    def test_line_save_beats_sku_save(self):
        picked = pick_saved_standard(
            "101", "SKU-A",
            {"101": {"standard_no": "GB/T 9832-2026"}},
            {"SKU-A": {"standard_no": "GB/T 9832-2007"}},
        )
        self.assertEqual("GB/T 9832-2026", picked["standard_no"])

    def test_sku_save_used_when_line_has_none(self):
        picked = pick_saved_standard(
            "101", "SKU-A", {},
            {"SKU-A": {"standard_no": "GB/T 9832-2007"}},
        )
        self.assertEqual("GB/T 9832-2007", picked["standard_no"])

    def test_poi_override_wins_and_does_not_auto_pick(self):
        gb = resolve_line_gb(
            poi_id="101", sku="SKU-A",
            gb_overrides={"101": "GB/T 9832-2026"},
            line_saves={"101": {"standard_no": "GB/T 9832-2007"}},
            sku_saves={},
            lookup=self.lookup,
        )
        self.assertEqual("GB/T 9832-2026", gb["standard_no"])
        self.assertEqual("new", gb["samr_id"])

    def test_empty_override_clears_saved_value(self):
        gb = resolve_line_gb(
            poi_id="101", sku="SKU-A",
            gb_overrides={"101": ""},
            line_saves={"101": {"standard_no": "GB/T 9832-2007"}},
            sku_saves={},
            lookup=self.lookup,
        )
        self.assertEqual("", gb["standard_no"])

    def test_no_override_and_no_save_stays_empty(self):
        gb = resolve_line_gb(
            poi_id="101", sku="SKU-A",
            gb_overrides={}, line_saves={}, sku_saves={},
            lookup=self.lookup,
        )
        self.assertEqual("", gb["standard_no"])

    def test_saved_option_uses_catalog_status_for_badge(self):
        option = saved_gb_option(
            {"samr_id": "dead", "standard_no": "GB/T 1-1993", "name_cn": "旧名"},
            {"samr_id": "dead", "standard_no": "GB/T 1-1993", "name_cn": "已废止示例", "status": "废止"},
        )
        self.assertEqual("废止", option["status"])
        self.assertEqual("已废止示例", option["nameCn"])
        missing = saved_gb_option({"standard_no": "GB/T 9", "name_cn": "未入库"})
        self.assertEqual("已保存", missing["status"])
        self.assertEqual("GB/T 9", missing["standardNo"])

    def test_unknown_standard_fails_closed(self):
        with self.assertRaises(ValueError) as ctx:
            resolve_line_gb(
                poi_id="101", sku="SKU-A",
                gb_overrides={"101": "GB/T 9999-9999"},
                line_saves={}, sku_saves={},
                lookup=self.lookup,
            )
        self.assertIn("不在国标目录中", str(ctx.exception))

    def test_retired_standard_fails_closed(self):
        with self.assertRaises(ValueError) as ctx:
            resolve_line_gb(
                poi_id="101", sku="SKU-A",
                gb_overrides={"101": "GB/T 1-1993"},
                line_saves={}, sku_saves={},
                lookup=self.lookup,
            )
        self.assertIn("废止", str(ctx.exception))

    def test_sku_keyed_override_is_accepted(self):
        gb = resolve_line_gb(
            poi_id="101", sku="SKU-A",
            gb_overrides={"SKU-A": "GB/T 9832-2007"},
            line_saves={}, sku_saves={},
            lookup=self.lookup,
        )
        self.assertEqual("GB/T 9832-2007", gb["standard_no"])


class ExcelLayoutTests(unittest.TestCase):
    def setUp(self):
        self._temp = tempfile.TemporaryDirectory(prefix="contract-gb-")
        self.dir = Path(self._temp.name)

    def tearDown(self):
        self._temp.cleanup()

    def _write(self, model, name="layout.xlsx"):
        from backend.contract_workbook import write_contract_workbook
        path = self.dir / name
        write_contract_workbook(model, path)
        return path

    def test_execution_standard_column_sits_after_barcode(self):
        from openpyxl import load_workbook
        from backend.contract_workbook import layout_for, unit_price_header

        png = self.dir / "dot.png"
        png.write_bytes(
            b"\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR\x00\x00\x00\x01\x00\x00\x00\x01"
            b"\x08\x06\x00\x00\x00\x1f\x15\xc4\x89\x00\x00\x00\nIDATx\x9cc\x00\x01"
            b"\x00\x00\x05\x00\x01\r\n-\xb4\x00\x00\x00\x00IEND\xaeB`\x82"
        )
        model = _sample_model(items=[
            _sample_item(imagePath=str(png), remark="包体32+2个魔术贴标3.45"),
            _sample_item(
                poiId="102", sku="SKU-B", nationalCode="ABC-01",
                gbStandard="", imagePath=None, quantity=2, unitPrice=8.5,
            ),
        ])
        path = self._write(model, "two-items.xlsx")
        wb = load_workbook(path)
        ws = wb["合同"]
        layout = layout_for(2)
        headers = [ws.cell(8, col).value for col in range(1, 17)]
        self.assertEqual(16, len(headers))
        self.assertEqual("国标码", headers[4])
        self.assertEqual("品名", headers[5])
        self.assertEqual("执行标准", headers[6])
        self.assertLess(headers.index("国标码"), headers.index("品名"))
        self.assertLess(headers.index("品名"), headers.index("执行标准"))
        self.assertEqual(unit_price_header(model["invoice"]), headers[13])
        self.assertIn("13%", headers[13])
        self.assertIn("专票", headers[13])
        self.assertEqual("6901234567890", ws["E9"].value)
        self.assertFalse(str(ws["E9"].value).startswith("="))
        self.assertEqual("@", ws["E9"].number_format)
        self.assertEqual("小熊", ws["F9"].value)
        self.assertEqual("GB/T 9832-2007", ws["G9"].value)
        self.assertEqual("ABC-01", ws["E10"].value)
        self.assertFalse(ws["G10"].value)
        self.assertEqual(10, ws["L9"].value)
        self.assertEqual("=N9*L9", ws["O9"].value)
        self.assertEqual("=N10*L10", ws["O10"].value)
        self.assertEqual(f"=SUM(O{layout.item_start}:O{layout.item_end})", ws["O11"].value)
        merged = {str(item) for item in ws.merged_cells.ranges}
        self.assertIn("A1:P2", merged)
        self.assertIn("A6:A7", merged)
        self.assertIn("B6:H7", merged)
        self.assertIn(f"A{layout.total_row}:N{layout.total_row}", merged)
        self.assertEqual("收货信息", ws["A6"].value)
        self.assertEqual("center", ws["B4"].alignment.horizontal)
        self.assertEqual("center", ws["B5"].alignment.horizontal)
        self.assertEqual("center", ws["J5"].alignment.horizontal)
        self.assertEqual("center", ws["B6"].alignment.horizontal)
        self.assertEqual(1, len(ws._images))
        self.assertEqual("暂无图片", ws["B10"].value)
        self.assertIsNone(ws["B9"].value)
        self.assertIsNone(ws["P9"].value)
        self.assertEqual("包体32+2个魔术贴标3.45", ws["Q9"].value)
        self.assertEqual("备注", ws["Q8"].value)
        self.assertFalse(bool(ws.column_dimensions["Q"].hidden))
        self.assertIn(f"$A$1:$P${layout.final_row}", ws.print_area or "")
        wb.close()

    def test_decimal_quantity_is_written_not_truncated(self):
        from decimal import Decimal
        from openpyxl import load_workbook

        model = _sample_model(items=[_sample_item(quantity=Decimal("100.6"))])
        path = self._write(model, "qty-decimal.xlsx")
        wb = load_workbook(path)
        self.assertEqual(100.6, float(wb["合同"]["L9"].value))
        self.assertNotEqual(100, wb["合同"]["L9"].value)
        self.assertEqual("=N9*L9", wb["合同"]["O9"].value)
        wb.close()

    def test_one_item_puts_delivery_address_on_row_14(self):
        from openpyxl import load_workbook
        from backend.contract_workbook import layout_for

        model = _sample_model()
        path = self._write(model, "one-item.xlsx")
        wb = load_workbook(path)
        ws = wb["合同"]
        layout = layout_for(1)
        self.assertEqual(14, layout.address_row)
        self.assertEqual("送货地址", ws["A14"].value)
        self.assertEqual("鄂州仓：默认收货信息", ws["B14"].value)
        self.assertEqual("鄂州仓：默认收货信息", ws["B6"].value)
        self.assertEqual("收货信息", ws["A6"].value)
        wb.close()

    def test_blank_template_keeps_fixed_fields_and_empty_po(self):
        from openpyxl import load_workbook
        from backend.contracts import DEFAULT_RECEIVING_INFO, write_blank_contract_template

        path = write_blank_contract_template(self.dir / "采购合同模板.xlsx")
        wb = load_workbook(path)
        ws = wb["合同"]
        self.assertEqual("杭 州 无 际 云 帆 采 购 单", ws["A1"].value)
        self.assertEqual("需方：", ws["A5"].value)
        self.assertEqual("杭州无际云帆文化传媒有限公司", ws["B5"].value)
        self.assertEqual("收货信息", ws["A6"].value)
        self.assertEqual(DEFAULT_RECEIVING_INFO, ws["B6"].value)
        self.assertIsNone(ws["B4"].value)
        self.assertFalse(ws["J5"].value)
        self.assertEqual("国标码", ws["E8"].value)
        self.assertEqual("品名", ws["F8"].value)
        self.assertEqual("执行标准", ws["G8"].value)
        self.assertEqual("单价", ws["N8"].value)
        self.assertIsNone(ws["A9"].value)
        self.assertEqual("=N9*L9", ws["O9"].value)
        self.assertEqual("送货地址", ws["A14"].value)
        self.assertEqual(DEFAULT_RECEIVING_INFO, ws["B14"].value)
        self.assertEqual("包装", ws["A11"].value)
        self.assertEqual("检验标准", ws["A13"].value)
        wb.close()

    def test_no_invoice_header_and_empty_items_fail(self):
        from openpyxl import load_workbook
        from backend.contract_workbook import write_contract_workbook

        model = _sample_model(invoice_type="no_invoice", tax_rate=0)
        path = self._write(model, "no-invoice.xlsx")
        wb = load_workbook(path)
        self.assertEqual("单价（不开票）", wb["合同"]["N8"].value)
        wb.close()
        with self.assertRaises(ValueError):
            write_contract_workbook({**model, "items": []}, self.dir / "empty.xlsx")

    def test_missing_image_file_writes_failure_placeholder(self):
        from openpyxl import load_workbook

        model = _sample_model(items=[_sample_item(imagePath="/tmp/not-a-real-product.png")])
        path = self._write(model, "bad-image.xlsx")
        wb = load_workbook(path)
        self.assertEqual("图片读取失败", wb["合同"]["B9"].value)
        self.assertEqual(0, len(wb["合同"]._images))
        wb.close()

    def test_webp_is_converted_and_embedded(self):
        from openpyxl import load_workbook
        from PIL import Image

        webp = self.dir / "dot.webp"
        Image.new("RGB", (8, 8), (255, 0, 0)).save(webp, "WEBP")
        model = _sample_model(items=[_sample_item(imagePath=str(webp))])
        path = self._write(model, "webp.xlsx")
        wb = load_workbook(path)
        self.assertEqual(1, len(wb["合同"]._images))
        self.assertIsNone(wb["合同"]["B9"].value)
        wb.close()


def _sample_item(**overrides):
    item = {
        "poiId": "101",
        "sku": "SKU-A",
        "styleCode": "ST-A",
        "nationalCode": "6901234567890",
        "gbStandard": "GB/T 9832-2007",
        "name": "小熊",
        "category": "毛绒",
        "virtualCategory": "",
        "materialProcess": "短毛绒",
        "packaging": "袋装",
        "quantity": 10,
        "unit": "个",
        "unitPrice": 21.7,
        "remark": "",
        "imagePath": None,
    }
    item.update(overrides)
    return item


def _sample_model(*, invoice_type="special_invoice", tax_rate=13, items=None, **overrides):
    labels = {"no_invoice": "不开票", "normal_invoice": "普票", "special_invoice": "专票"}
    model = {
        "purchaseOrderNo": "604264",
        "orderDate": "2026-06-05",
        "deliveryDate": "2026-07-09",
        "projectLead": "",
        "buyer": {
            "company_name": "杭州无际云帆文化传媒有限公司",
            "warehouse_name": "鄂州仓",
            "contact": "蜀黍家 13385711803",
        },
        "supplier": {
            "shortName": "佰特",
            "legalName": "测试供应商有限公司",
            "address": "测试地址",
            "contact": "张三 13800000000",
        },
        "invoice": {
            "type": invoice_type,
            "label": labels[invoice_type],
            "taxRate": tax_rate,
        },
        "items": items or [_sample_item()],
        "packagingTerms": "包装条款",
        "paymentTerms": "100% - 现结\n采购单号604264",
        "inspectionStandards": "检验条款",
        "receivingInfo": "鄂州仓：默认收货信息",
        "deliveryAddress": "鄂州仓：默认收货信息",
        "terms": ["1、条款一", "2、条款二"],
        "applicant": "采购员",
    }
    model.update(overrides)
    return model


if __name__ == "__main__":
    unittest.main()
